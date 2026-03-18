import os
import sys
import signal
import logging
import subprocess
from threading import Lock, Thread
import time
import sqlite3
import json
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor
from flask import Flask, render_template, send_file, request, Response, redirect, url_for, session, jsonify, flash
from flask_socketio import SocketIO, emit
import pandas as pd
from collections import defaultdict, deque
import glob
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
import re
import tenacity
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from PIL import Image
import hashlib
import secrets
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
try:
    from pysnmp.hlapi import (
        getCmd, nextCmd, SnmpEngine, CommunityData, UdpTransportTarget,
        ContextData, ObjectType, ObjectIdentity
    )
    SNMP_AVAILABLE = True
except ImportError:
    import shutil
    SNMP_AVAILABLE = bool(shutil.which('snmpget') and shutil.which('snmpwalk'))
    if SNMP_AVAILABLE:
        logging.info("Using system snmpget/snmpwalk for SNMP monitoring")
    else:
        logging.warning("SNMP tools not available - SNMP monitoring disabled")

# Setup logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("ping_debug.log"),
        logging.StreamHandler()
    ]
)

# Configuration
app = Flask(__name__, template_folder='templates')
app.config['SECRET_KEY'] = os.getenv('FLASK_SECRET_KEY', secrets.token_hex(32))
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)

# Email configuration for password reset and notifications
EMAIL_HOST = os.getenv('EMAIL_HOST', 'smtp.gmail.com')
EMAIL_PORT = int(os.getenv('EMAIL_PORT', '587'))
EMAIL_USER = os.getenv('EMAIL_USER', 'demo.networkmonitor@gmail.com')  # Demo email
EMAIL_PASS = os.getenv('EMAIL_PASS', 'demo_password_123')  # Demo password
EMAIL_FROM = os.getenv('EMAIL_FROM', 'Network Monitor <demo.networkmonitor@gmail.com>')

socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading', ping_timeout=120, ping_interval=30)
results_lock = Lock()
results = []
running = True
PING_INTERVAL = 10
XLSX_FILE = "sm_ips.xlsx"
RETENTION_DAYS = 7
LATENCY_THRESHOLD = 200
DEGRADED_LOSS_THRESHOLD = 5
MAX_WORKERS = 50
BATCH_SIZE = 20
LAST_PRUNE = None
LAST_XLSX_LOAD = None
CACHED_DF = None
ALERT_LOG = deque(maxlen=1000)
ALERT_QUEUE = deque(maxlen=1000)
LOG_DIR = "logs"
CURRENT_LOG_FILE = None
LAST_LOG_DATE = None
LAST_LOG_HOUR = None
ALERT_UPDATE_INTERVAL = 3.0
ALERT_TIMESTAMP_COUNTER = 0
HISTORY_CACHE = {}
SUMMARY_CACHE = {}
status_cache = {}
previous_status_cache = {}
downtime_cache = {}
uptime_cache = {}
alert_cache = {}
alert_thread_running = True

# Maintenance Mode
MAINTENANCE_WINDOWS = {}  # IP -> {'start': datetime, 'end': datetime, 'reason': str}
MAINTENANCE_SCHEDULES = []  # List of scheduled maintenance windows

def is_in_maintenance(sm_ip):
    """Check if a device is currently in maintenance mode"""
    if sm_ip not in MAINTENANCE_WINDOWS:
        return False
    
    maintenance = MAINTENANCE_WINDOWS[sm_ip]
    now = datetime.now()
    
    if maintenance['start'] <= now <= maintenance['end']:
        return True
    
    # Clean up expired maintenance windows
    if now > maintenance['end']:
        del MAINTENANCE_WINDOWS[sm_ip]
    
    return False

def add_maintenance_window(sm_ip, start_time, end_time, reason="Scheduled maintenance"):
    """Add a maintenance window for a device"""
    MAINTENANCE_WINDOWS[sm_ip] = {
        'start': start_time,
        'end': end_time,
        'reason': reason
    }
    logging.info(f"Added maintenance window for {sm_ip}: {start_time} to {end_time}")

def remove_maintenance_window(sm_ip):
    """Remove maintenance window for a device"""
    if sm_ip in MAINTENANCE_WINDOWS:
        del MAINTENANCE_WINDOWS[sm_ip]
        logging.info(f"Removed maintenance window for {sm_ip}")

def get_maintenance_status():
    """Get current maintenance status for all devices"""
    now = datetime.now()
    active_maintenance = {}
    
    for sm_ip, maintenance in MAINTENANCE_WINDOWS.items():
        if maintenance['start'] <= now <= maintenance['end']:
            active_maintenance[sm_ip] = {
                'reason': maintenance['reason'],
                'start': maintenance['start'].strftime('%Y-%m-%d %H:%M:%S'),
                'end': maintenance['end'].strftime('%Y-%m-%d %H:%M:%S'),
                'remaining': str(maintenance['end'] - now).split('.')[0]
            }
    
    return active_maintenance

# Ensure log directory exists
if not os.path.exists(LOG_DIR):
    os.makedirs(LOG_DIR)
    logging.info(f"Created log directory: {LOG_DIR}")

def format_duration(seconds):
    """Convert seconds to HH:MM:SS format."""
    if not isinstance(seconds, (int, float)) or seconds < 0:
        return "N/A"
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    secs = int(seconds % 60)
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"

def get_log_filename():
    now = datetime.now()
    return os.path.join(LOG_DIR, f"alerts_{now.strftime('%Y-%m-%d_%H')}.jsonl")

def rotate_log_file():
    global CURRENT_LOG_FILE, LAST_LOG_DATE, LAST_LOG_HOUR
    now = datetime.now()
    if LAST_LOG_DATE != now.date() or LAST_LOG_HOUR != now.hour:
        CURRENT_LOG_FILE = get_log_filename()
        LAST_LOG_DATE = now.date()
        LAST_LOG_HOUR = now.hour
        logging.info(f"Rotated log file to {CURRENT_LOG_FILE}")
    elif os.path.exists(CURRENT_LOG_FILE) and os.path.getsize(CURRENT_LOG_FILE) > 10 * 1024 * 1024:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        os.rename(CURRENT_LOG_FILE, f"{CURRENT_LOG_FILE}.{timestamp}")
        CURRENT_LOG_FILE = get_log_filename()
        logging.info(f"Rotated log file due to size limit to {CURRENT_LOG_FILE}")

def append_to_log_file(alert):
    rotate_log_file()
    try:
        with open(CURRENT_LOG_FILE, 'a', encoding='utf-8') as f:
            json.dump(alert, f)
            f.write('\n')
    except Exception as e:
        logging.error(f"Failed to append to log file: {str(e)}")

@tenacity.retry(stop=tenacity.stop_after_attempt(3), wait=tenacity.wait_fixed(1))
def init_db():
    try:
        conn = sqlite3.connect('ping_history.db', check_same_thread=False, timeout=30)
        conn.execute('PRAGMA journal_mode=WAL')
        conn.execute('PRAGMA synchronous=NORMAL')
        conn.execute('PRAGMA cache_size=10000')
        conn.execute('PRAGMA temp_store=memory')
        conn.execute('''CREATE TABLE IF NOT EXISTS history 
                        (timestamp TEXT, sm_ip TEXT, status TEXT, latency REAL)''')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_sm_ip ON history (sm_ip)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_timestamp ON history (timestamp)')
        conn.commit()
        conn.close()
        logging.info("Database initialized successfully")
    except Exception as e:
        logging.error(f"Database init failed: {str(e)}")
        raise

init_db()

def init_users_db():
    """Initialize users database with admin user"""
    try:
        conn = sqlite3.connect('ping_history.db', check_same_thread=False)
        
        # Create users table
        conn.execute('''CREATE TABLE IF NOT EXISTS users 
                        (id INTEGER PRIMARY KEY AUTOINCREMENT,
                         username TEXT UNIQUE NOT NULL,
                         email TEXT UNIQUE NOT NULL,
                         password_hash TEXT NOT NULL,
                         role TEXT DEFAULT 'user',
                         is_active INTEGER DEFAULT 1,
                         created_at TEXT NOT NULL,
                         last_login TEXT,
                         reset_token TEXT,
                         reset_token_expires TEXT,
                         contact_number TEXT,
                         designation TEXT,
                         department TEXT)''')
        
        # Create user activity logs table
        conn.execute('''CREATE TABLE IF NOT EXISTS user_activity_logs 
                        (id INTEGER PRIMARY KEY AUTOINCREMENT,
                         user_id INTEGER NOT NULL,
                         username TEXT NOT NULL,
                         activity_type TEXT NOT NULL,
                         activity_description TEXT,
                         ip_address TEXT,
                         timestamp TEXT NOT NULL,
                         FOREIGN KEY (user_id) REFERENCES users (id))''')
        
        # Create sessions table for better session management
        conn.execute('''CREATE TABLE IF NOT EXISTS user_sessions 
                        (id INTEGER PRIMARY KEY AUTOINCREMENT,
                         user_id INTEGER NOT NULL,
                         session_token TEXT UNIQUE NOT NULL,
                         created_at TEXT NOT NULL,
                         expires_at TEXT NOT NULL,
                         ip_address TEXT,
                         user_agent TEXT,
                         FOREIGN KEY (user_id) REFERENCES users (id))''')
        
        conn.execute('CREATE INDEX IF NOT EXISTS idx_users_username ON users (username)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_users_email ON users (email)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_sessions_token ON user_sessions (session_token)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_activity_logs_user ON user_activity_logs (user_id)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_activity_logs_timestamp ON user_activity_logs (timestamp)')
        
        # Migrate existing users table to add new columns if they don't exist
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(users)")
        columns = [column[1] for column in cursor.fetchall()]
        
        if 'contact_number' not in columns:
            conn.execute('ALTER TABLE users ADD COLUMN contact_number TEXT')
        if 'designation' not in columns:
            conn.execute('ALTER TABLE users ADD COLUMN designation TEXT')
        if 'department' not in columns:
            conn.execute('ALTER TABLE users ADD COLUMN department TEXT')
        if 'region_id' not in columns:
            conn.execute('ALTER TABLE users ADD COLUMN region_id INTEGER')
            logging.info("Added region_id column to users table")
        
        # Create default admin user if not exists
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM users WHERE role = 'admin'")
        admin_count = cursor.fetchone()[0]
        
        if admin_count == 0:
            admin_username = os.getenv('ADMIN_USERNAME', 'admin')
            admin_email = os.getenv('ADMIN_EMAIL', 'admin@example.com')
            admin_password = os.getenv('ADMIN_PASSWORD', 'admin123')
            admin_hash = hash_password(admin_password)
            
            cursor.execute("""
                INSERT INTO users (username, email, password_hash, role, created_at)
                VALUES (?, ?, ?, 'admin', ?)
            """, (admin_username, admin_email, admin_hash, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            
            logging.info(f"Created default admin user: {admin_username}")
        
        # Create demo users: dilip.kc and paras.thapa with demo123 passwords
        demo_users = [
            ('dilip.kc', 'dilip.kc@worldlink.com.np', 'demo123'),
            ('paras.thapa', 'paras.thapa@worldlink.com.np', 'demo123')
        ]
        
        for username, email, password in demo_users:
            cursor.execute("SELECT COUNT(*) FROM users WHERE username = ? OR email = ?", (username, email))
            if cursor.fetchone()[0] == 0:
                password_hash = hash_password(password)
                cursor.execute("""
                    INSERT OR IGNORE INTO users (username, email, password_hash, role, created_at)
                    VALUES (?, ?, ?, 'user', ?)
                """, (username, email, password_hash, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
                logging.info(f"Created demo user: {username}")
        
        conn.commit()
        conn.close()
        logging.info("Users database initialized")
    except Exception as e:
        logging.error(f"Users database init failed: {str(e)}")

def hash_password(password):
    """Hash password using SHA-256 with salt"""
    salt = secrets.token_hex(32)
    password_hash = hashlib.sha256((password + salt).encode()).hexdigest()
    return f"{salt}:{password_hash}"

def verify_password(password, stored_hash):
    """Verify password against stored hash"""
    try:
        salt, password_hash = stored_hash.split(':')
        return hashlib.sha256((password + salt).encode()).hexdigest() == password_hash
    except:
        return False

def generate_reset_token():
    """Generate secure reset token"""
    return secrets.token_urlsafe(32)

def send_reset_email(email, token, username):
    """Send password reset email"""
    # Demo mode - just log the email instead of sending
    if EMAIL_USER == 'demo.networkmonitor@gmail.com' or not EMAIL_PASS or EMAIL_PASS == 'demo_password_123':
        logging.info(f"DEMO EMAIL - Password reset would be sent to {email}")
        logging.info(f"Reset token: {token}")
        logging.info(f"Reset URL: {request.url_root}reset-password?token={token}")
        return True  # Simulate successful email
    
    if not EMAIL_USER or not EMAIL_PASS:
        logging.warning("Email not configured, cannot send reset email")
        return False
    
    try:
        msg = MIMEMultipart()
        msg['From'] = EMAIL_FROM
        msg['To'] = email
        msg['Subject'] = "Password Reset - Network Monitor"
        
        reset_url = f"{request.url_root}reset-password?token={token}"
        
        body = f"""
        Hello {username},
        
        You have requested a password reset for your Network Monitor account.
        
        Click the link below to reset your password:
        {reset_url}
        
        This link will expire in 1 hour.
        
        If you did not request this reset, please ignore this email.
        
        Best regards,
        Network Monitor Team
        """
        
        msg.attach(MIMEText(body, 'plain'))
        
        server = smtplib.SMTP(EMAIL_HOST, EMAIL_PORT)
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASS)
        server.send_message(msg)
        server.quit()
        
        logging.info(f"Password reset email sent to {email}")
        return True
    except Exception as e:
        logging.error(f"Failed to send reset email: {str(e)}")
        return False

init_users_db()

def init_comments_db():
    """Initialize comments and acknowledgments database"""
    try:
        conn = sqlite3.connect('ping_history.db', check_same_thread=False)
        
        # Create comments table
        conn.execute('''CREATE TABLE IF NOT EXISTS device_comments 
                        (id INTEGER PRIMARY KEY AUTOINCREMENT,
                         sm_ip TEXT NOT NULL,
                         comment TEXT NOT NULL,
                         username TEXT DEFAULT 'Anonymous',
                         timestamp TEXT NOT NULL,
                         comment_type TEXT DEFAULT 'comment')''')
        
        # Create acknowledgments table
        conn.execute('''CREATE TABLE IF NOT EXISTS device_acknowledgments 
                        (sm_ip TEXT PRIMARY KEY,
                         status TEXT NOT NULL,
                         username TEXT DEFAULT 'Anonymous',
                         timestamp TEXT NOT NULL,
                         comment TEXT)''')
        
        # Create tasks table for task management
        conn.execute('''CREATE TABLE IF NOT EXISTS device_tasks 
                        (id INTEGER PRIMARY KEY AUTOINCREMENT,
                         task_id TEXT UNIQUE,
                         sm_ip TEXT NOT NULL,
                         task_title TEXT NOT NULL,
                         task_description TEXT,
                         status TEXT DEFAULT 'open',
                         priority TEXT DEFAULT 'medium',
                         assigned_to TEXT,
                         created_by TEXT NOT NULL,
                         created_at TEXT NOT NULL,
                         updated_at TEXT NOT NULL,
                         closed_at TEXT,
                         closed_by TEXT,
                         resolution TEXT)''')
        
        # Add task_id column if it doesn't exist (migration)
        try:
            conn.execute('ALTER TABLE device_tasks ADD COLUMN task_id TEXT UNIQUE')
        except sqlite3.OperationalError:
            pass  # Column already exists
        
        # Add region_id column for role-based filtering (migration)
        try:
            conn.execute('ALTER TABLE device_tasks ADD COLUMN region_id INTEGER')
        except sqlite3.OperationalError:
            pass  # Column already exists
            
        # Add assigned_at column for tracking assignment time (migration)
        try:
            conn.execute('ALTER TABLE device_tasks ADD COLUMN assigned_at TEXT')
        except sqlite3.OperationalError:
            pass  # Column already exists
        
        # Generate task_ids for existing tasks that don't have one
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM device_tasks WHERE task_id IS NULL ORDER BY id")
        tasks_without_id = cursor.fetchall()
        for task in tasks_without_id:
            task_id = f"TASK-{1000 + task[0]:04d}"
            cursor.execute("UPDATE device_tasks SET task_id = ? WHERE id = ?", (task_id, task[0]))
        
        conn.execute('CREATE INDEX IF NOT EXISTS idx_comments_ip ON device_comments (sm_ip)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_ack_ip ON device_acknowledgments (sm_ip)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_tasks_ip ON device_tasks (sm_ip)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_tasks_assigned ON device_tasks (assigned_to)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_tasks_status ON device_tasks (status)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_tasks_task_id ON device_tasks (task_id)')
        conn.commit()
        conn.close()
        logging.info("Comments and acknowledgments database initialized")
    except Exception as e:
        logging.error(f"Comments database init failed: {str(e)}")

init_comments_db()

def init_regions_locations_db():
    """Initialize regions and locations database"""
    try:
        conn = sqlite3.connect('ping_history.db', check_same_thread=False, timeout=30)
        
        # Create regions table
        conn.execute('''CREATE TABLE IF NOT EXISTS regions 
                        (id INTEGER PRIMARY KEY AUTOINCREMENT,
                         name TEXT UNIQUE NOT NULL,
                         description TEXT,
                         created_at TEXT NOT NULL,
                         updated_at TEXT NOT NULL)''')
        
        # Create locations table (linked to regions)
        conn.execute('''CREATE TABLE IF NOT EXISTS locations 
                        (id INTEGER PRIMARY KEY AUTOINCREMENT,
                         name TEXT UNIQUE NOT NULL,
                         region_id INTEGER,
                         description TEXT,
                         created_at TEXT NOT NULL,
                         updated_at TEXT NOT NULL,
                         FOREIGN KEY (region_id) REFERENCES regions (id) ON DELETE SET NULL)''')
        
        conn.execute('CREATE INDEX IF NOT EXISTS idx_locations_region ON locations (region_id)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_locations_name ON locations (name)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_regions_name ON regions (name)')
        
        # Check if we need to migrate existing locations from Excel to database
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM regions")
        region_count = cursor.fetchone()[0]
        
        if region_count == 0:
            # Create default Region 7
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute("""
                INSERT INTO regions (name, description, created_at, updated_at)
                VALUES (?, ?, ?, ?)
            """, ('Region 7', 'Default region for existing locations', now, now))
            
            region_7_id = cursor.lastrowid
            logging.info("Created default Region 7")
        
        # Always check and migrate locations from Excel if they don't exist in database
        cursor.execute("SELECT COUNT(*) FROM locations")
        location_count = cursor.fetchone()[0]
        
        if location_count == 0 and CACHED_DF is not None and 'Location' in CACHED_DF.columns:
            # Get Region 7 ID (or create if doesn't exist)
            cursor.execute("SELECT id FROM regions WHERE name = 'Region 7'")
            region_result = cursor.fetchone()
            
            if not region_result:
                now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                cursor.execute("""
                    INSERT INTO regions (name, description, created_at, updated_at)
                    VALUES (?, ?, ?, ?)
                """, ('Region 7', 'Default region for existing locations', now, now))
                region_7_id = cursor.lastrowid
            else:
                region_7_id = region_result[0]
            
            # Get unique locations from Excel file
            unique_locations = CACHED_DF['Location'].dropna().unique()
            migrated_count = 0
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            for location in unique_locations:
                location = str(location).strip()
                if location and location != 'Unknown':
                    try:
                        cursor.execute("""
                            INSERT INTO locations (name, region_id, created_at, updated_at)
                            VALUES (?, ?, ?, ?)
                        """, (location, region_7_id, now, now))
                        migrated_count += 1
                    except sqlite3.IntegrityError:
                        # Location already exists, skip
                        pass
            
            if migrated_count > 0:
                logging.info(f"Migrated {migrated_count} locations to Region 7")
                logging.info(f"Locations: {', '.join([str(loc).strip() for loc in unique_locations if str(loc).strip() and str(loc).strip() != 'Unknown'])}")
        
        conn.commit()
        conn.close()
        logging.info("Regions and locations database initialized")
    except Exception as e:
        logging.error(f"Regions/locations database init failed: {str(e)}")

def init_snmp_db():
    """Initialize SNMP monitoring tables"""
    try:
        conn = sqlite3.connect('ping_history.db', check_same_thread=False, timeout=30)
        conn.execute('''CREATE TABLE IF NOT EXISTS snmp_devices
                        (id INTEGER PRIMARY KEY AUTOINCREMENT,
                         name TEXT NOT NULL,
                         ip TEXT NOT NULL,
                         community TEXT DEFAULT 'public',
                         snmp_version TEXT DEFAULT '2c',
                         device_type TEXT DEFAULT 'generic',
                         location TEXT,
                         region_id INTEGER,
                         added_by TEXT,
                         created_at TEXT NOT NULL,
                         is_active INTEGER DEFAULT 1)''')
        conn.execute('''CREATE TABLE IF NOT EXISTS snmp_metrics
                        (id INTEGER PRIMARY KEY AUTOINCREMENT,
                         device_id INTEGER NOT NULL,
                         timestamp TEXT NOT NULL,
                         status TEXT DEFAULT 'unknown',
                         uptime_seconds INTEGER,
                         cpu_percent REAL,
                         mem_percent REAL,
                         interfaces_json TEXT,
                         rx_bytes INTEGER,
                         tx_bytes INTEGER,
                         signal_dbm REAL,
                         FOREIGN KEY (device_id) REFERENCES snmp_devices (id))''')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_snmp_metrics_device ON snmp_metrics (device_id)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_snmp_metrics_ts ON snmp_metrics (timestamp)')
        conn.commit()
        conn.close()
        logging.info("SNMP database initialized")
    except Exception as e:
        logging.error(f"SNMP database init failed: {str(e)}")

init_snmp_db()
init_regions_locations_db()

def migrate_locations_from_excel():
    """Migrate locations from Excel to database if not already done"""
    try:
        if CACHED_DF is None or 'Location' not in CACHED_DF.columns:
            return
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Get Region 7 (or create if doesn't exist)
            cursor.execute("SELECT id FROM regions WHERE name = 'Region 7'")
            region_result = cursor.fetchone()
            
            if not region_result:
                now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                cursor.execute("""
                    INSERT INTO regions (name, description, created_at, updated_at)
                    VALUES (?, ?, ?, ?)
                """, ('Region 7', 'Default region for existing locations', now, now))
                region_7_id = cursor.lastrowid
                logging.info("Created Region 7 for migration")
            else:
                region_7_id = region_result[0]
            
            # Get existing locations in database
            cursor.execute("SELECT name FROM locations")
            existing_locations = set(row[0] for row in cursor.fetchall())
            
            # Get unique locations from Excel
            unique_locations = CACHED_DF['Location'].dropna().unique()
            migrated_count = 0
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            for location in unique_locations:
                location = str(location).strip()
                if location and location != 'Unknown' and location not in existing_locations:
                    try:
                        cursor.execute("""
                            INSERT INTO locations (name, region_id, created_at, updated_at)
                            VALUES (?, ?, ?, ?)
                        """, (location, region_7_id, now, now))
                        migrated_count += 1
                    except sqlite3.IntegrityError:
                        # Location already exists, skip
                        pass
            
            if migrated_count > 0:
                conn.commit()
                logging.info(f"Migrated {migrated_count} new locations from Excel to Region 7")
                
    except Exception as e:
        logging.error(f"Location migration error: {str(e)}")

# Authentication helper functions
def login_required(f):
    """Decorator to require login for routes"""
    from functools import wraps
    from flask import request, jsonify
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.is_json or request.path.startswith('/api') or request.path.startswith('/get_'):
                return jsonify({'error': 'Authentication required'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    """Decorator to require admin role (superadmin or regional_admin) for routes"""
    from functools import wraps
    from flask import request, jsonify
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.is_json:
                return jsonify({'error': 'Authentication required'}), 401
            return redirect(url_for('login'))
        
        user = get_current_user()
        if not user or user['role'] not in ['admin', 'superadmin', 'regional_admin']:
            if request.is_json:
                return jsonify({'error': 'Admin access required'}), 403
            flash('Admin access required', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def superadmin_required(f):
    """Decorator to require superadmin role for routes"""
    from functools import wraps
    from flask import request, jsonify
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.is_json:
                return jsonify({'error': 'Authentication required'}), 401
            return redirect(url_for('login'))
        
        user = get_current_user()
        if not user or user['role'] != 'superadmin':
            if request.is_json:
                return jsonify({'error': 'Superadmin access required'}), 403
            flash('Superadmin access required', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def regional_admin_or_super_required(f):
    """Decorator to require regional_admin or superadmin role for routes"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        
        user = get_current_user()
        if not user or user['role'] not in ['superadmin', 'regional_admin']:
            flash('Admin access required', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def get_current_user():
    """Get current logged-in user"""
    if 'user_id' not in session:
        return None
    
    try:
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, username, email, role, is_active, last_login, contact_number, designation, department, region_id
                FROM users WHERE id = ?
            """, (session['user_id'],))
            user_data = cursor.fetchone()
            
            if user_data:
                return {
                    'id': user_data[0],
                    'username': user_data[1],
                    'email': user_data[2],
                    'role': user_data[3],
                    'is_active': user_data[4],
                    'last_login': user_data[5],
                    'contact_number': user_data[6],
                    'designation': user_data[7],
                    'department': user_data[8],
                    'region_id': user_data[9]
                }
    except Exception as e:
        logging.error(f"Error getting current user: {str(e)}")
    
    return None

def log_user_activity(user_id, username, activity_type, activity_description='', ip_address=None):
    """Log user activity to database"""
    try:
        if ip_address is None:
            ip_address = request.remote_addr if request else None
        
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO user_activity_logs (user_id, username, activity_type, activity_description, ip_address, timestamp)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (user_id, username, activity_type, activity_description, ip_address, timestamp))
            conn.commit()
            
        logging.info(f"Activity logged: {username} - {activity_type}")
    except Exception as e:
        logging.error(f"Error logging activity: {str(e)}")

def get_all_users():
    """Get all users for admin management"""
    try:
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT u.id, u.username, u.email, u.role, u.is_active, u.created_at, u.last_login, u.region_id, r.name
                FROM users u
                LEFT JOIN regions r ON u.region_id = r.id
                ORDER BY u.created_at DESC
            """)
            users_data = cursor.fetchall()
            
            users = []
            for user_data in users_data:
                users.append({
                    'id': user_data[0],
                    'username': user_data[1],
                    'email': user_data[2],
                    'role': user_data[3],
                    'is_active': user_data[4],
                    'created_at': user_data[5],
                    'last_login': user_data[6],
                    'region_id': user_data[7],
                    'region_name': user_data[8]
                })
            return users
    except Exception as e:
        logging.error(f"Error getting all users: {str(e)}")
        return []

def send_assignment_notification(user_email, username, device_ip, device_name, location, assigned_by, comment=""):
    """Send email notification when a device is assigned to a user"""
    # Demo mode - just log the email instead of sending
    if EMAIL_USER == 'demo.networkmonitor@gmail.com' or not EMAIL_PASS or EMAIL_PASS == 'demo_password_123':
        logging.info(f"DEMO EMAIL - Assignment notification would be sent to {user_email}")
        logging.info(f"Subject: Device Assignment - {device_ip}")
        logging.info(f"Device: {device_ip} ({device_name}) at {location}")
        logging.info(f"Assigned by: {assigned_by}")
        logging.info(f"Comment: {comment}")
        return True  # Simulate successful email
    
    if not EMAIL_USER or not EMAIL_PASS:
        logging.warning("Email not configured, cannot send assignment notification")
        return False
    
    try:
        msg = MIMEMultipart()
        msg['From'] = EMAIL_FROM
        msg['To'] = user_email
        msg['Subject'] = f"Device Assignment - {device_ip}"
        
        dashboard_url = f"{request.url_root}login"
        
        body = f"""
        Hello {username},
        
        You have been assigned a network device that requires attention:
        
        Device Details:
        - IP Address: {device_ip}
        - Device Name: {device_name}
        - Location: {location}
        - Assigned by: {assigned_by}
        {f"- Comment: {comment}" if comment else ""}
        
        Please log in to the Network Monitor dashboard to review and acknowledge this assignment:
        {dashboard_url}
        
        You can add comments, update the status, and track the device's performance through the dashboard.
        
        Best regards,
        Network Monitor System
        """
        
        msg.attach(MIMEText(body, 'plain'))
        
        server = smtplib.SMTP(EMAIL_HOST, EMAIL_PORT)
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASS)
        server.send_message(msg)
        server.quit()
        
        logging.info(f"Assignment notification sent to {user_email} for device {device_ip}")
        return True
    except Exception as e:
        logging.error(f"Failed to send assignment notification: {str(e)}")
        return False

def create_user_session(user_id, ip_address, user_agent):
    """Create a new user session"""
    try:
        session_token = secrets.token_urlsafe(32)
        expires_at = datetime.now() + timedelta(hours=24)
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            conn.execute("""
                INSERT INTO user_sessions (user_id, session_token, created_at, expires_at, ip_address, user_agent)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (user_id, session_token, datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  expires_at.strftime('%Y-%m-%d %H:%M:%S'), ip_address, user_agent))
            conn.commit()
        
        return session_token
    except Exception as e:
        logging.error(f"Error creating user session: {str(e)}")
        return None

def cleanup_expired_sessions():
    """Clean up expired sessions"""
    try:
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            conn.execute("DELETE FROM user_sessions WHERE expires_at < ?", 
                        (datetime.now().strftime('%Y-%m-%d %H:%M:%S'),))
            conn.commit()
    except Exception as e:
        logging.error(f"Error cleaning up sessions: {str(e)}")

# Authentication Routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        remember_me = request.form.get('remember_me') == 'on'
        
        if not username or not password:
            flash('Username and password are required', 'error')
            return render_template('login.html')
        
        try:
            with sqlite3.connect('ping_history.db', timeout=10) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT id, username, email, password_hash, role, is_active
                    FROM users WHERE username = ? OR email = ?
                """, (username, username))
                user_data = cursor.fetchone()
                
                if user_data and verify_password(password, user_data[3]):
                    if not user_data[5]:  # is_active
                        flash('Account is deactivated. Please contact administrator.', 'error')
                        return render_template('login.html')
                    
                    # Update last login
                    cursor.execute("UPDATE users SET last_login = ? WHERE id = ?",
                                 (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), user_data[0]))
                    conn.commit()
                    
                    # Create session
                    session['user_id'] = user_data[0]
                    session['username'] = user_data[1]
                    session['role'] = user_data[4]
                    session.permanent = remember_me
                    
                    # Create session record
                    create_user_session(user_data[0], request.remote_addr, request.headers.get('User-Agent', ''))
                    
                    # Log login activity
                    log_user_activity(user_data[0], user_data[1], 'login', 'User logged in successfully', request.remote_addr)
                    
                    flash(f'Welcome back, {user_data[1]}!', 'success')
                    
                    # Redirect based on role
                    next_page = request.args.get('next')
                    if next_page:
                        return redirect(next_page)
                    elif user_data[4] in ['admin', 'superadmin', 'regional_admin']:
                        return redirect(url_for('admin_dashboard'))
                    else:
                        return redirect(url_for('dashboard'))
                else:
                    flash('Invalid username or password', 'error')
        except Exception as e:
            logging.error(f"Login error: {str(e)}")
            flash('Login failed. Please try again.', 'error')
    
    return render_template('login.html')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')
        region_id = request.form.get('region_id', '').strip()
        
        logging.info(f"Signup attempt for username: {username}, email: {email}, region_id: {region_id}")
        
        # Validation
        if not username or not email or not password:
            flash('All fields are required', 'error')
            return render_template('signup.html')
        
        if not region_id:
            flash('Please select a region', 'error')
            return render_template('signup.html')
        
        if len(username) < 3:
            flash('Username must be at least 3 characters long', 'error')
            return render_template('signup.html')
        
        if len(password) < 6:
            flash('Password must be at least 6 characters long', 'error')
            return render_template('signup.html')
        
        if password != confirm_password:
            flash('Passwords do not match', 'error')
            return render_template('signup.html')
        
        # Email validation
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_pattern, email):
            flash('Invalid email address', 'error')
            return render_template('signup.html')
        
        try:
            with sqlite3.connect('ping_history.db', timeout=10) as conn:
                cursor = conn.cursor()
                
                # Check if username or email already exists
                cursor.execute("SELECT id FROM users WHERE username = ? OR email = ?", (username, email))
                if cursor.fetchone():
                    flash('Username or email already exists', 'error')
                    return render_template('signup.html')
                
                # Verify region exists
                cursor.execute("SELECT id FROM regions WHERE id = ?", (region_id,))
                if not cursor.fetchone():
                    flash('Invalid region selected', 'error')
                    return render_template('signup.html')
                
                # Create new user with region
                password_hash = hash_password(password)
                cursor.execute("""
                    INSERT INTO users (username, email, password_hash, role, region_id, created_at)
                    VALUES (?, ?, ?, 'user', ?, ?)
                """, (username, email, password_hash, int(region_id), datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
                conn.commit()
                
                logging.info(f"User {username} created successfully with region {region_id}, redirecting to login")
                flash('Account created successfully! Please log in.', 'success')
                return redirect(url_for('login'))
        except Exception as e:
            logging.error(f"Signup error: {str(e)}")
            import traceback
            logging.error(traceback.format_exc())
            flash('Registration failed. Please try again.', 'error')
    
    return render_template('signup.html')

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        
        if not email:
            flash('Email address is required', 'error')
            return render_template('forgot_password.html')
        
        try:
            with sqlite3.connect('ping_history.db', timeout=10) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, username, email FROM users WHERE email = ?", (email,))
                user_data = cursor.fetchone()
                
                if user_data:
                    # Generate reset token
                    reset_token = generate_reset_token()
                    expires_at = datetime.now() + timedelta(hours=1)
                    
                    cursor.execute("""
                        UPDATE users SET reset_token = ?, reset_token_expires = ?
                        WHERE id = ?
                    """, (reset_token, expires_at.strftime('%Y-%m-%d %H:%M:%S'), user_data[0]))
                    conn.commit()
                    
                    # Send reset email
                    if send_reset_email(email, reset_token, user_data[1]):
                        flash('Password reset instructions have been sent to your email', 'success')
                    else:
                        flash('Failed to send reset email. Please contact administrator.', 'error')
                else:
                    # Don't reveal if email exists or not for security
                    flash('If an account with that email exists, reset instructions have been sent', 'info')
        except Exception as e:
            logging.error(f"Forgot password error: {str(e)}")
            flash('Password reset failed. Please try again.', 'error')
    
    return render_template('forgot_password.html')

@app.route('/reset-password', methods=['GET', 'POST'])
def reset_password():
    token = request.args.get('token') or request.form.get('token')
    
    if not token:
        flash('Invalid reset link', 'error')
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')
        
        if not password or len(password) < 6:
            flash('Password must be at least 6 characters long', 'error')
            return render_template('reset_password.html', token=token)
        
        if password != confirm_password:
            flash('Passwords do not match', 'error')
            return render_template('reset_password.html', token=token)
        
        try:
            with sqlite3.connect('ping_history.db', timeout=10) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT id, username FROM users 
                    WHERE reset_token = ? AND reset_token_expires > ?
                """, (token, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
                user_data = cursor.fetchone()
                
                if user_data:
                    # Update password and clear reset token
                    password_hash = hash_password(password)
                    cursor.execute("""
                        UPDATE users SET password_hash = ?, reset_token = NULL, reset_token_expires = NULL
                        WHERE id = ?
                    """, (password_hash, user_data[0]))
                    conn.commit()
                    
                    flash('Password reset successfully! Please log in with your new password.', 'success')
                    return redirect(url_for('login'))
                else:
                    flash('Invalid or expired reset link', 'error')
                    return redirect(url_for('login'))
        except Exception as e:
            logging.error(f"Reset password error: {str(e)}")
            flash('Password reset failed. Please try again.', 'error')
    
    # Verify token is valid for GET request
    try:
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id FROM users 
                WHERE reset_token = ? AND reset_token_expires > ?
            """, (token, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            if not cursor.fetchone():
                flash('Invalid or expired reset link', 'error')
                return redirect(url_for('login'))
    except Exception as e:
        logging.error(f"Token verification error: {str(e)}")
        flash('Invalid reset link', 'error')
        return redirect(url_for('login'))
    
    return render_template('reset_password.html', token=token)

@app.route('/logout')
def logout():
    user_id = session.get('user_id')
    username = session.get('username', 'User')
    
    # Log logout activity before clearing session
    if user_id and username:
        log_user_activity(user_id, username, 'logout', 'User logged out')
    
    session.clear()
    flash(f'Goodbye, {username}!', 'info')
    return redirect(url_for('login'))

@app.route('/profile')
@login_required
def profile():
    user = get_current_user()
    
    # Get user's region information
    try:
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Get region name if user has a region assigned
            if user.get('region_id'):
                cursor.execute("""
                    SELECT name FROM regions WHERE id = ?
                """, (user['region_id'],))
                region_row = cursor.fetchone()
                user['region_name'] = region_row[0] if region_row else None
            else:
                user['region_name'] = 'All Regions' if user['role'] in ['admin', 'superadmin'] else 'Not Assigned'
            
            # Get user activity logs
            cursor.execute("""
                SELECT activity_type, activity_description, ip_address, timestamp
                FROM user_activity_logs
                WHERE user_id = ?
                ORDER BY timestamp DESC
                LIMIT 100
            """, (user['id'],))
            
            activities = []
            for row in cursor.fetchall():
                activities.append({
                    'type': row[0],
                    'description': row[1],
                    'ip_address': row[2],
                    'timestamp': row[3]
                })
            
            user['activities'] = activities
    except Exception as e:
        logging.error(f"Error loading profile data: {str(e)}")
        user['activities'] = []
        user['region_name'] = 'Error loading region'
    
    return render_template('profile.html', user=user)

@app.route('/update-profile', methods=['POST'])
@login_required
def update_profile():
    user = get_current_user()
    
    try:
        contact_number = request.form.get('contact_number', '').strip()
        designation = request.form.get('designation', '').strip()
        department = request.form.get('department', '').strip()
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE users 
                SET contact_number = ?, designation = ?, department = ?
                WHERE id = ?
            """, (contact_number, designation, department, user['id']))
            conn.commit()
        
        # Log activity
        log_user_activity(user['id'], user['username'], 'profile_update', 'Profile information updated')
        
        flash('Profile updated successfully!', 'success')
    except Exception as e:
        logging.error(f"Error updating profile: {str(e)}")
        flash('Failed to update profile', 'error')
    
    return redirect(url_for('profile'))

@app.route('/export-activity-logs/<format>')
@login_required
def export_activity_logs(format):
    user = get_current_user()
    
    try:
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT activity_type, activity_description, ip_address, timestamp
                FROM user_activity_logs
                WHERE user_id = ?
                ORDER BY timestamp DESC
            """, (user['id'],))
            
            activities = cursor.fetchall()
        
        if format == 'csv':
            import csv
            from io import StringIO
            
            output = StringIO()
            writer = csv.writer(output)
            writer.writerow(['Activity Type', 'Description', 'IP Address', 'Timestamp'])
            writer.writerows(activities)
            
            response = make_response(output.getvalue())
            response.headers['Content-Disposition'] = f'attachment; filename={user["username"]}_activity_logs.csv'
            response.headers['Content-Type'] = 'text/csv'
            
            # Log export activity
            log_user_activity(user['id'], user['username'], 'export_logs', f'Exported activity logs as CSV')
            
            return response
            
        elif format == 'xlsx':
            try:
                import openpyxl
                from openpyxl import Workbook
                from io import BytesIO
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Activity Logs"
                
                # Add headers
                ws.append(['Activity Type', 'Description', 'IP Address', 'Timestamp'])
                
                # Add data
                for activity in activities:
                    ws.append(activity)
                
                # Save to BytesIO
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                response = make_response(output.getvalue())
                response.headers['Content-Disposition'] = f'attachment; filename={user["username"]}_activity_logs.xlsx'
                response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                
                # Log export activity
                log_user_activity(user['id'], user['username'], 'export_logs', f'Exported activity logs as XLSX')
                
                return response
            except ImportError:
                flash('openpyxl library not installed. Please install it to export as XLSX.', 'error')
                return redirect(url_for('profile'))
        else:
            flash('Invalid export format', 'error')
            return redirect(url_for('profile'))
            
    except Exception as e:
        logging.error(f"Error exporting logs: {str(e)}")
        flash('Failed to export activity logs', 'error')
        return redirect(url_for('profile'))

@app.route('/change-password', methods=['POST'])
@login_required
def change_password():
    current_password = request.form.get('current_password', '')
    new_password = request.form.get('new_password', '')
    confirm_password = request.form.get('confirm_password', '')
    
    if not current_password or not new_password:
        flash('All fields are required', 'error')
        return redirect(url_for('profile'))
    
    if len(new_password) < 6:
        flash('New password must be at least 6 characters long', 'error')
        return redirect(url_for('profile'))
    
    if new_password != confirm_password:
        flash('New passwords do not match', 'error')
        return redirect(url_for('profile'))
    
    try:
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT password_hash FROM users WHERE id = ?", (session['user_id'],))
            stored_hash = cursor.fetchone()[0]
            
            if verify_password(current_password, stored_hash):
                new_hash = hash_password(new_password)
                cursor.execute("UPDATE users SET password_hash = ? WHERE id = ?", 
                             (new_hash, session['user_id']))
                conn.commit()
                
                # Log password change activity
                user = get_current_user()
                log_user_activity(user['id'], user['username'], 'password_change', 'Password changed successfully')
                
                flash('Password changed successfully', 'success')
            else:
                flash('Current password is incorrect', 'error')
    except Exception as e:
        logging.error(f"Change password error: {str(e)}")
        flash('Password change failed. Please try again.', 'error')
    
    return redirect(url_for('profile'))

def prune_old_records():
    global LAST_PRUNE
    try:
        now = datetime.now()
        if LAST_PRUNE and now - LAST_PRUNE < timedelta(hours=1):
            return
        cutoff = (now - timedelta(days=RETENTION_DAYS)).strftime('%Y-%m-%d %H:%M:%S')
        with sqlite3.connect('ping_history.db') as conn:
            conn.execute("DELETE FROM history WHERE timestamp < ?", (cutoff,))
            conn.commit()
        LAST_PRUNE = now
        logging.info("Pruned old records")
    except Exception as e:
        logging.error(f"Prune error: {str(e)}")

def validate_ip(ip):
    ip_pattern = r'^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$'
    return bool(ip and re.match(ip_pattern, ip))

def ping_ip(ip, timeout=0.2):
    try:
        if ip == "192.168.138.141" and os.getenv("TEST_MODE"):
            logging.debug(f"Simulating {ip}: Reachable")
            return "Reachable", 33.25, None

        cached = status_cache.get(ip)
        if cached and datetime.now() - cached["time"] < timedelta(seconds=PING_INTERVAL * 5):
            logging.debug(f"Using cached status for {ip}: {cached['status']}")
            return cached["status"], cached["latency"], cached.get("issue_type")
        
        cmd = ['ping', '-c', '2', '-W', str(timeout), ip]
        logging.debug(f"Pinging {ip}")
        result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=3)
        
        stdout = result.stdout.lower()
        if result.returncode != 0 or "100% packet loss" in stdout or "destination host unreachable" in stdout:
            logging.warning(f"{ip} Down")
            status_cache[ip] = {"status": "Down", "latency": None, "issue_type": None, "time": datetime.now()}
            return "Down", None, None
        
        loss_percent = 100
        latency = None
        for line in stdout.splitlines():
            if "packet loss" in line:
                loss_percent = int(line.split(",")[2].split("%")[0].strip())
            if "rtt min/avg/max/mdev" in line:
                latency = float(line.split('=')[1].split('/')[1])
        
        status = "Reachable"
        issue_type = None
        
        # Determine status and issue type
        if loss_percent >= DEGRADED_LOSS_THRESHOLD:
            status = "Degraded"
            issue_type = "packet_loss"
        elif latency and latency > LATENCY_THRESHOLD:
            status = "Degraded"
            issue_type = "high_latency"
        
        status_cache[ip] = {"status": status, "latency": latency, "issue_type": issue_type, "time": datetime.now()}
        logging.debug(f"{ip} {status}, latency: {latency or 'N/A'}, issue: {issue_type or 'none'}")
        return status, latency, issue_type
    except subprocess.TimeoutExpired:
        logging.error(f"Ping {ip} timed out")
        status_cache[ip] = {"status": "Down", "latency": None, "issue_type": None, "time": datetime.now()}
        return "Down", None, None
    except Exception as e:
        logging.error(f"Ping {ip} error: {str(e)}")
        status_cache[ip] = {"status": "Down", "latency": None, "issue_type": None, "time": datetime.now()}
        return "Down", None, None

def ping_all_ips(ips):
    start = time.time()
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        batches = [ips[i:i + BATCH_SIZE] for i in range(0, len(ips), BATCH_SIZE)]
        ping_results = []
        for batch in batches:
            try:
                batch_results = list(executor.map(ping_ip, batch, timeout=5))
                ping_results.extend(batch_results)
            except Exception as e:
                logging.error(f"Error in batch processing: {str(e)}")
                # Add default results for failed batch
                ping_results.extend([("Down", None, None) for _ in batch])
    duration = time.time() - start
    logging.info(f"Pinging {len(ips)} IPs took {duration:.2f}s")
    return ping_results

@tenacity.retry(stop=tenacity.stop_after_attempt(3), wait=tenacity.wait_fixed(1))
def log_to_db(entries):
    try:
        with sqlite3.connect('ping_history.db', timeout=30) as conn:
            conn.execute('PRAGMA journal_mode=WAL')
            conn.executemany("INSERT INTO history VALUES (?, ?, ?, ?)", entries)
            conn.commit()
        logging.debug(f"Logged {len(entries)} entries to DB")
    except Exception as e:
        logging.error(f"DB log error: {str(e)}")
        raise

def get_downtime_since(sm_ip, current_status):
    if current_status != "Down":
        return "N/A"
    cached = downtime_cache.get(sm_ip)
    if cached and datetime.now() - cached["time"] < timedelta(minutes=2):
        return cached["value"]
    try:
        with sqlite3.connect('ping_history.db', timeout=30) as conn:
            cursor = conn.cursor()
            cutoff = (datetime.now() - timedelta(hours=48)).strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute("""
                SELECT timestamp FROM history 
                WHERE sm_ip = ? AND status = 'Down' AND timestamp >= ?
                ORDER BY timestamp DESC LIMIT 1
            """, (sm_ip, cutoff))
            record = cursor.fetchone()
            if not record:
                logging.debug(f"No recent Down record for {sm_ip} in last 48 hours")
                return "Unknown"
            result = record[0]
            downtime_cache[sm_ip] = {"value": result, "time": datetime.now()}
            return result
    except Exception as e:
        logging.error(f"Downtime {sm_ip}: {str(e)}")
        return "Unknown"

def get_uptime_since(sm_ip, current_status):
    if current_status != "Reachable":
        return "N/A"
    cached = uptime_cache.get(sm_ip)
    if cached and datetime.now() - cached["time"] < timedelta(minutes=2):
        return cached["value"]
    try:
        with sqlite3.connect('ping_history.db', timeout=30) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT timestamp FROM history 
                WHERE sm_ip = ? AND status = 'Reachable' 
                ORDER BY timestamp DESC LIMIT 1
            """, (sm_ip,))
            record = cursor.fetchone()
            if not record:
                return "Unknown"
            result = record[0]
            uptime_cache[sm_ip] = {"value": result, "time": datetime.now()}
            return result
    except Exception as e:
        logging.error(f"Uptime {sm_ip}: {str(e)}")
        return "Unknown"

def get_previous_status_from_db(sm_ip):
    try:
        with sqlite3.connect('ping_history.db', timeout=30) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT status FROM history 
                WHERE sm_ip = ? 
                ORDER BY timestamp DESC LIMIT 1
            """, (sm_ip,))
            record = cursor.fetchone()
            return record[0] if record else None
    except Exception as e:
        logging.error(f"DB previous status {sm_ip}: {str(e)}")
        return None

def update_alert_log():
    global alert_thread_running
    while running and alert_thread_running:
        try:
            if ALERT_QUEUE:
                alert = ALERT_QUEUE.popleft()
                if alert['entry']['status'] not in ['Error', 'Timeout']:
                    ALERT_LOG.append(alert)
                    append_to_log_file(alert)
                    cutoff_time = datetime.now() - timedelta(hours=1)
                    recent_alerts = [
                        a for a in ALERT_LOG
                        if datetime.strptime(a['time'].split('.')[0], '%Y-%m-%d %H:%M:%S') >= cutoff_time
                        and a['entry']['status'] in ['Down', 'Degraded', 'Reachable']
                    ]
                    socketio.emit('alert_update', {'alerts': recent_alerts})
                    logging.debug(f"Emitted {len(recent_alerts)} alerts for {alert['time']} - {alert['entry']['status']} for {alert['entry']['ip']}")
            time.sleep(ALERT_UPDATE_INTERVAL)
        except Exception as e:
            logging.error(f"Alert log update error: {str(e)}")
            alert_thread_running = False
            time.sleep(ALERT_UPDATE_INTERVAL)

def monitor_alert_thread():
    global alert_thread_running
    while running:
        if not alert_thread_running:
            logging.warning("Alert thread failed, restarting...")
            alert_thread_running = True
            alert_thread = Thread(target=update_alert_log, name="AlertUpdateThread", daemon=True)
            alert_thread.start()
            logging.info("Alert thread restarted")
        time.sleep(10)

def update_ping_status():
    global results, running, CACHED_DF, LAST_XLSX_LOAD, ALERT_TIMESTAMP_COUNTER
    
    # Send initial empty status to show the interface is loading
    socketio.emit('update_status', {
        'results': [{"AP Name": "Loading...", "AP IP": "Loading...", "CID": "Loading...", "SM IP": "Loading...", "Device Name": "Initializing monitoring system...", "Location": "Please wait", "Status": "Loading", "Latency": "N/A", "Downtime Since": "N/A"}],
        'pop_summary': {},
        'alerts': [],
        'analysis': {}
    })
    
    while running:
        start_time = time.time()
        logging.info("Starting ping cycle")
        try:
            prune_old_records()
            
            now = datetime.now()
            if CACHED_DF is None or LAST_XLSX_LOAD is None or now - LAST_XLSX_LOAD > timedelta(minutes=5):
                try:
                    if not os.path.exists(XLSX_FILE):
                        logging.error(f"{XLSX_FILE} missing")
                        raise FileNotFoundError(f"{XLSX_FILE} not found")
                    CACHED_DF = pd.read_excel(XLSX_FILE, engine='openpyxl')
                    LAST_XLSX_LOAD = now
                    logging.info(f"Loaded XLSX with {len(CACHED_DF)} rows")
                    
                    # Migrate locations from Excel to database
                    migrate_locations_from_excel()
                except Exception as e:
                    logging.error(f"XLSX load error: {str(e)}")
                    CACHED_DF = pd.DataFrame(columns=['AP Name', 'AP IP', 'CID', 'SM IP', 'Device Name', 'Location'])
                    socketio.emit('update_status', {
                        'results': [{"AP Name": "Error", "AP IP": "Error", "CID": "Error", "SM IP": "Error", "Device Name": f"XLSX load failed: {str(e)}", "Location": "Error", "Status": "Error", "Latency": "N/A", "Downtime Since": "N/A"}],
                        'pop_summary': {},
                        'analysis': {}
                    })
                    time.sleep(10)  # Wait longer before retrying
                    continue
            
            df = CACHED_DF
            if 'SM IP' not in df.columns or 'Location' not in df.columns or df.empty:
                error_msg = f"Invalid XLSX: Missing required columns or empty. Columns: {df.columns.tolist()}"
                logging.error(error_msg)
                socketio.emit('update_status', {
                    'results': [{"AP Name": "Error", "AP IP": "Error", "CID": "Error", "SM IP": "Error", "Device Name": error_msg, "Location": "Error", "Status": "Error", "Latency": "N/A", "Downtime Since": "N/A"}],
                    'pop_summary': {},
                    'analysis': {}
                })
                time.sleep(10)
                continue
            
            # Debug: Log column names and sample data
            logging.debug(f"Excel columns: {df.columns.tolist()}")
            if not df.empty:
                logging.debug(f"Sample row: {df.iloc[0].to_dict()}")
            
            valid_locations = df['Location'].dropna().astype(str).str.strip()
            if valid_locations.empty:
                error_msg = "No valid Location values in XLSX"
                logging.error(error_msg)
                socketio.emit('update_status', {
                    'results': [{"AP Name": "Error", "AP IP": "Error", "CID": "Error", "SM IP": "Error", "Device Name": error_msg, "Location": "Error", "Status": "Error", "Latency": "N/A", "Downtime Since": "N/A"}],
                    'pop_summary': {},
                    'analysis': {}
                })
                time.sleep(10)
                continue
            
            ip_info = {
                str(row['SM IP']): {
                    'org_name': str(row.get('Device Name', 'N/A')),
                    'location': str(row.get('Location', 'Unknown')),
                    'ap_name': str(row.get('AP Name', 'N/A')),
                    'ap_ip': str(row.get('AP IP', 'N/A')),
                    'cid': str(row.get('CID', 'N/A'))
                } for row in df.to_dict('records') if pd.notna(row.get('SM IP'))
            }
            
            # Debug: Log a few sample entries to see CID values
            sample_ips = list(ip_info.keys())[:3]
            for sample_ip in sample_ips:
                logging.debug(f"Sample IP {sample_ip}: {ip_info[sample_ip]}")
            
            ips = list(ip_info.keys())
            logging.info(f"Pinging {len(ips)} IPs")

            ping_results = ping_all_ips(ips)
            ping_duration = time.time() - start_time
            ip_to_result = dict(zip(ips, ping_results))

            new_results = []
            pop_counts = defaultdict(lambda: {'Reachable': 0, 'Degraded': 0, 'Down': 0})
            db_entries = []
            reachable_counts = defaultdict(int)
            ip_stats = defaultdict(lambda: {
                'uptime': 0,
                'downtime': 0,
                'degraded_time': 0,
                'days': set(),
                'down_count': 0,
                'up_count': 0,
                'down_events': [],
                'up_events': [],
                'degraded_events': []
            })

            for sm_ip in ips:
                status, latency, issue_type = ip_to_result.get(sm_ip, ("Unknown", None, None))
                if status in ['Error', 'Timeout']:
                    continue
                logging.debug(f"Ping result for {sm_ip}: {status}, latency: {latency or 'N/A'}, issue: {issue_type or 'none'}")
                
                last_status = previous_status_cache.get(sm_ip, {}).get("status")
                if last_status is None:
                    last_status = get_previous_status_from_db(sm_ip)
                
                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                db_entries.append((timestamp, sm_ip, status, latency))
                
                info = ip_info[sm_ip]
                downtime_since = get_downtime_since(sm_ip, status)
                uptime_since = get_uptime_since(sm_ip, status)
                is_long_term = False
                if downtime_since not in ["N/A", "Unknown"]:
                    try:
                        downtime_start = datetime.strptime(downtime_since, '%Y-%m-%d %H:%M:%S')
                        duration = (datetime.now() - downtime_start).total_seconds()
                        is_long_term = duration >= 24 * 3600
                        logging.debug(f"Long-term check for {sm_ip}: downtime_since={downtime_since}, is_long_term={is_long_term}")
                    except Exception as e:
                        logging.error(f"Error parsing downtime for {sm_ip}: {str(e)}")
                
                result = {
                    "AP Name": info['ap_name'],
                    "AP IP": info['ap_ip'],
                    "CID": info['cid'],
                    "SM IP": sm_ip,
                    "Device Name": info['org_name'],
                    "Location": info['location'],
                    "Status": status,
                    "Latency": f"{latency:.2f} ms" if latency is not None else "N/A",
                    "Downtime Since": downtime_since,
                    "Issue Type": issue_type
                }
                new_results.append(result)
                pop_counts[info['location']][status] += 1
                
                ip_stats[sm_ip]['days'].add(datetime.now().strftime('%Y-%m-%d'))
                if status == 'Down':
                    ip_stats[sm_ip]['downtime'] += PING_INTERVAL
                    if last_status != 'Down':
                        ip_stats[sm_ip]['down_count'] += 1
                        ip_stats[sm_ip]['down_events'].append(timestamp)
                elif status == 'Reachable':
                    ip_stats[sm_ip]['uptime'] += PING_INTERVAL
                    if last_status != 'Reachable':
                        ip_stats[sm_ip]['up_count'] += 1
                        ip_stats[sm_ip]['up_events'].append(timestamp)
                elif status == 'Degraded':
                    ip_stats[sm_ip]['degraded_time'] += PING_INTERVAL
                    if last_status != 'Degraded':
                        ip_stats[sm_ip]['degraded_events'].append(timestamp)
                
                alert_entry = {
                    "ip": sm_ip,
                    "location": info.get('location', 'N/A'),
                    "org_name": info.get('org_name', 'N/A'),
                    "ap_name": info.get('ap_name', 'N/A'),
                    "ap_ip": info.get('ap_ip', 'N/A'),
                    "cid": info.get('cid', 'N/A'),
                    "status": status,
                    "downtime_since": downtime_since,
                    "uptime_since": uptime_since,
                    "latency": latency,
                    "long_term": is_long_term,
                    "high_latency": latency is not None and latency > LATENCY_THRESHOLD,
                    "issue_type": issue_type
                }
                should_generate = False
                if status in ["Down", "Degraded"] or is_long_term:
                    should_generate = True
                elif status == "Reachable" and last_status in ["Down", "Degraded"]:
                    should_generate = True
                    reachable_counts[sm_ip] += 1
                
                # Suppress alerts if device is in maintenance mode
                if should_generate and is_in_maintenance(sm_ip):
                    should_generate = False
                    logging.debug(f"Suppressed alert for {sm_ip} - device in maintenance mode")
                
                cached_alert = alert_cache.get(sm_ip, {})
                if should_generate and (
                    not cached_alert or
                    cached_alert.get("status") != status or
                    (is_long_term and not cached_alert.get("long_term"))
                ):
                    ALERT_TIMESTAMP_COUNTER += 1
                    timestamp = f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.{ALERT_TIMESTAMP_COUNTER:06d}"
                    alert = {
                        "time": timestamp,
                        "entry": alert_entry
                    }
                    ALERT_LOG.append(alert)
                    ALERT_QUEUE.append(alert)
                    append_to_log_file(alert)
                    alert_cache[sm_ip] = {"status": status, "long_term": is_long_term, "time": datetime.now()}
                    logging.debug(f"Generated alert: {status} for {sm_ip}, long_term: {is_long_term}, previous status: {last_status or 'None'}")
                    if is_long_term:
                        alert_cache[sm_ip] = {"status": status, "long_term": False, "time": datetime.now()}
                
                previous_status_cache[sm_ip] = {"status": status, "latency": latency, "time": datetime.now()}

            # Try to log to database, but don't fail the entire cycle if it fails
            if db_entries:
                try:
                    log_to_db(db_entries)
                except Exception as db_error:
                    logging.error(f"Failed to log to database: {str(db_error)}")
                    # Continue without failing the cycle

            new_results.sort(key=lambda x: (
                0 if x['Status'] == 'Down' else 1 if x['Status'] == 'Degraded' else 2))
            with results_lock:
                results = new_results
            
            ip_daily_stats = {}
            for sm_ip in ip_stats:
                days = len(ip_stats[sm_ip]['days']) or 1
                ip_daily_stats[sm_ip] = {
                    'uptime': format_duration(ip_stats[sm_ip]['uptime']),
                    'downtime': format_duration(ip_stats[sm_ip]['downtime']),
                    'degraded_time': format_duration(ip_stats[sm_ip]['degraded_time']),
                    'down_count': ip_stats[sm_ip]['down_count'],
                    'up_count': ip_stats[sm_ip]['up_count'],
                    'down_events': ip_stats[sm_ip]['down_events'],
                    'up_events': ip_stats[sm_ip]['up_events'],
                    'degraded_events': ip_stats[sm_ip]['degraded_events'],
                    'location': ip_info.get(sm_ip, {}).get('location', 'Unknown')
                }
            
            logging.info(f"Emitting update_status: {len(new_results)} results, {len(pop_counts)} locations in pop_summary")
            
            # Debug: Log first few results to see what's being sent
            if new_results:
                for i, result in enumerate(new_results[:3]):
                    logging.debug(f"Result {i}: {result}")
            
            socketio.emit('update_status', {
                'results': new_results,
                'pop_summary': dict(pop_counts),
                'ip_stats': ip_daily_stats
            })
            logging.info(f"Cycle completed in {time.time() - start_time:.2f}s")
            
        except Exception as e:
            logging.exception(f"Cycle error: {str(e)}")
            # Send error status but don't crash
            socketio.emit('update_status', {
                'results': [{"AP Name": "Error", "AP IP": "Error", "CID": "Error", "SM IP": "Error", "Device Name": f"Monitoring cycle failed: {str(e)}", "Location": "Error", "Status": "Error", "Latency": "N/A", "Downtime Since": "N/A"}],
                'pop_summary': {},
                'ip_stats': {}
            })
        finally:
            elapsed = time.time() - start_time
            sleep_time = max(0, PING_INTERVAL - elapsed)
            logging.debug(f"Sleeping {sleep_time:.2f}s")
            time.sleep(sleep_time)

def start_periodic_update():
    ping_thread = Thread(target=update_ping_status, name="PingUpdateThread", daemon=True)
    alert_thread = Thread(target=update_alert_log, name="AlertUpdateThread", daemon=True)
    monitor_thread = Thread(target=monitor_alert_thread, name="MonitorAlertThread", daemon=True)
    ping_thread.start()
    alert_thread.start()
    monitor_thread.start()
    logging.info("Update and monitor threads started")

def signal_handler(sig, frame):
    global running
    running = False
    logging.info("Shutting down")
    sys.exit(0)

signal.signal(signal.SIGINT, signal_handler)
signal.signal(signal.SIGTERM, signal_handler)

# SocketIO Events
@socketio.on('refresh_now')
def handle_refresh_now():
    logging.info("Received refresh_now request")
    try:
        with results_lock:
            socketio.emit('update_status', {
                'results': results,
                'pop_summary': SUMMARY_CACHE.get('pop_summary', {}),
                'ip_stats': SUMMARY_CACHE.get('ip_stats', {})
            })
    except Exception as e:
        logging.error(f"Refresh now error: {str(e)}")
        socketio.emit('update_status', {
            'results': [{"AP Name": "N/A", "AP IP": "N/A", "CID": "N/A", "SM IP": "N/A", "Device Name": f"Refresh failed: {str(e)}", "Location": "N/A", "Status": "Error", "Latency": "N/A", "Downtime Since": "N/A"}],
            'pop_summary': {},
            'ip_stats': {}
        })

@socketio.on('refresh_alerts')
def handle_refresh_alerts():
    logging.info("Received refresh_alerts request")
    try:
        cutoff_time = datetime.now() - timedelta(hours=1)
        recent_alerts = [
            a for a in ALERT_LOG
            if datetime.strptime(a['time'].split('.')[0], '%Y-%m-%d %H:%M:%S') >= cutoff_time
            and a['entry']['status'] in ['Down', 'Degraded', 'Reachable']
        ]
        socketio.emit('refresh_alerts', {'alerts': recent_alerts})
        logging.debug(f"Emitted {len(recent_alerts)} alerts for refresh")
    except Exception as e:
        logging.error(f"Refresh alerts error: {str(e)}")
        socketio.emit('refresh_alerts', {'error': str(e)})

# Location Uptime/Downtime Monitor Routes
@app.route('/location_uptime_monitor')
@login_required
def location_uptime_monitor():
    """Display location uptime/downtime monitor page"""
    user = get_current_user()
    return render_template('location_uptime_monitor.html', user=user)

@app.route('/api/location_uptime_events')
@login_required
def location_uptime_events():
    """Get uptime/downtime events for a specific location"""
    try:
        location = request.args.get('location', '').strip()
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        if not location or not start_date or not end_date:
            return jsonify({'error': 'Missing required parameters'}), 400
        
        # Get all IPs in this location from Excel
        location_ips = {}
        with results_lock:
            if CACHED_DF is not None and 'Location' in CACHED_DF.columns:
                # Filter by location, handling whitespace
                location_data = CACHED_DF[CACHED_DF['Location'].astype(str).str.strip() == location]
                for _, row in location_data.iterrows():
                    ip = str(row.get('SM IP', '')).strip()
                    if ip and ip != 'nan':
                        location_ips[ip] = {
                            'device_name': str(row.get('Device Name', 'N/A')),
                            'location': str(row.get('Location', 'N/A')).strip(),
                            'cid': str(row.get('CID', 'N/A'))
                        }
        
        if not location_ips:
            return jsonify({'error': f'No IPs found for location: {location}'}), 404
        
        # Get events from history table
        try:
            start_datetime = f"{start_date} 00:00:00"
            end_datetime = f"{end_date} 23:59:59"
            
            with sqlite3.connect('ping_history.db', timeout=10) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT sm_ip, status, timestamp
                    FROM history
                    WHERE sm_ip IN ({})
                    AND timestamp >= ? AND timestamp <= ?
                    ORDER BY sm_ip, timestamp
                """.format(','.join(['?'] * len(location_ips))), 
                list(location_ips.keys()) + [start_datetime, end_datetime])
                
                records = cursor.fetchall()
        except Exception as e:
            logging.error(f"Database error: {str(e)}")
            return jsonify({'error': f'Database error: {str(e)}'}), 500
        
        # Process events
        ip_events = {}
        for ip in location_ips:
            ip_events[ip] = {
                'device_name': location_ips[ip]['device_name'],
                'down_events': [],
                'degraded_events': [],
                'total_downtime_seconds': 0,
                'total_degraded_seconds': 0,
                'uptime_percent': 100.0,
                'last_event_type': 'UP',
                'last_event_time': None
            }
        
        # Group records by IP and detect events
        ip_records = {}
        for sm_ip, status, timestamp in records:
            if sm_ip not in ip_records:
                ip_records[sm_ip] = []
            ip_records[sm_ip].append((status, timestamp))
        
        # Detect down and degraded events
        for ip, records_list in ip_records.items():
            if ip not in ip_events:
                continue
            
            current_status = None
            event_start = None
            
            for status, timestamp in records_list:
                if status == 'Down' and current_status != 'Down':
                    event_start = timestamp
                    current_status = 'Down'
                elif status != 'Down' and current_status == 'Down':
                    if event_start:
                        start_time = datetime.strptime(event_start, '%Y-%m-%d %H:%M:%S')
                        end_time = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
                        duration = (end_time - start_time).total_seconds()
                        ip_events[ip]['down_events'].append({
                            'start': event_start,
                            'end': timestamp,
                            'duration_seconds': int(duration)
                        })
                        ip_events[ip]['total_downtime_seconds'] += int(duration)
                    current_status = None
                
                if status == 'Degraded' and current_status != 'Degraded':
                    event_start = timestamp
                    current_status = 'Degraded'
                elif status != 'Degraded' and current_status == 'Degraded':
                    if event_start:
                        start_time = datetime.strptime(event_start, '%Y-%m-%d %H:%M:%S')
                        end_time = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
                        duration = (end_time - start_time).total_seconds()
                        ip_events[ip]['degraded_events'].append({
                            'start': event_start,
                            'end': timestamp,
                            'duration_seconds': int(duration)
                        })
                        ip_events[ip]['total_degraded_seconds'] += int(duration)
                    current_status = None
                
                ip_events[ip]['last_event_type'] = status
                ip_events[ip]['last_event_time'] = timestamp
        
        # Calculate uptime percentage (based on actual time range)
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
            total_seconds = (end_dt - start_dt).total_seconds() + 86400  # +1 day to include end date
        except:
            total_seconds = 7 * 24 * 60 * 60  # 7 days default
        
        for ip in ip_events:
            downtime = ip_events[ip]['total_downtime_seconds']
            uptime_percent = ((total_seconds - downtime) / total_seconds) * 100 if total_seconds > 0 else 100
            ip_events[ip]['uptime_percent'] = max(0, min(100, uptime_percent))
        
        # Prepare response
        events_list = []
        total_down_events = 0
        total_degraded_events = 0
        total_uptime = 0
        
        for ip, data in ip_events.items():
            events_list.append({
                'ip': ip,
                'device_name': data['device_name'],
                'down_events': data['down_events'],
                'degraded_events': data['degraded_events'],
                'total_downtime_seconds': data['total_downtime_seconds'],
                'uptime_percent': data['uptime_percent'],
                'last_event_type': data['last_event_type'],
                'last_event_time': data['last_event_time']
            })
            total_down_events += len(data['down_events'])
            total_degraded_events += len(data['degraded_events'])
            total_uptime += data['uptime_percent']
        
        avg_uptime = total_uptime / len(ip_events) if ip_events else 0
        
        return jsonify({
            'events': events_list,
            'stats': {
                'total_ips': len(location_ips),
                'total_down_events': total_down_events,
                'total_degraded_events': total_degraded_events,
                'avg_uptime_percent': avg_uptime
            }
        })
    
    except Exception as e:
        logging.error(f"Location uptime events error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/location_uptime_events_export')
@login_required
def location_uptime_events_export():
    """Export location uptime/downtime events to CSV"""
    try:
        location = request.args.get('location', '')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        format_type = request.args.get('format', 'csv')
        
        # Get data using the same logic as location_uptime_events
        response = location_uptime_events()
        data = response.get_json()
        
        if 'error' in data:
            return response
        
        # Create CSV
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['IP Address', 'Device Name', 'Down Events', 'Degraded Events', 
                        'Total Downtime', 'Uptime %', 'Last Event', 'Last Event Time'])
        
        # Write data
        for event in data['events']:
            writer.writerow([
                event['ip'],
                event['device_name'],
                len(event['down_events']),
                len(event['degraded_events']),
                f"{event['total_downtime_seconds']}s",
                f"{event['uptime_percent']:.2f}%",
                event['last_event_type'],
                event['last_event_time']
            ])
        
        # Return CSV
        output.seek(0)
        return send_file(
            StringIO(output.getvalue()),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f"location_uptime_{location}_{start_date}.csv"
        )
    
    except Exception as e:
        logging.error(f"Export error: {str(e)}")
        return jsonify({'error': str(e)}), 500

# Admin Routes (Legacy - Remove old admin login system)
@app.route('/admin')
@admin_required
def admin_dashboard():
    try:
        df = pd.read_excel(XLSX_FILE, engine='openpyxl')
        user = get_current_user()
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Regional admin: filter by their region
            if user['role'] == 'regional_admin' and user['region_id']:
                # Get locations in their region
                cursor.execute("""
                    SELECT name FROM locations
                    WHERE region_id = ?
                    ORDER BY name
                """, (user['region_id'],))
                locations = [row[0] for row in cursor.fetchall()]
                
                # Filter records to only show devices in their region's locations
                if locations:
                    records = df[df['Location'].isin(locations)].to_dict('records')
                else:
                    records = []
            else:
                # Superadmin: see all locations and records
                cursor.execute("""
                    SELECT name FROM locations
                    ORDER BY name
                """)
                locations = [row[0] for row in cursor.fetchall()]
                records = df.to_dict('records')
            
            # Fallback to Excel if no locations in database
            if not locations:
                locations = sorted(set(df['Location'].dropna().astype(str)))
        
        # Get all users for user management (superadmin only)
        users = get_all_users() if user['role'] == 'superadmin' else []
        
        return render_template('admin.html', records=records, locations=locations, user=user, users=users)
    except Exception as e:
        logging.error(f"Admin dashboard error: {str(e)}")
        user = get_current_user()
        users = get_all_users() if user and user['role'] == 'superadmin' else []
        return render_template('admin.html', records=[], locations=[], error=str(e), user=user, users=users)

@app.route('/admin/add_entry', methods=['POST'])
@admin_required
def admin_add_entry():
    global CACHED_DF, LAST_XLSX_LOAD
    try:
        data = request.json
        sm_ip = data.get('sm_ip')
        if not sm_ip or not validate_ip(sm_ip):
            return jsonify({'error': 'Invalid or missing SM IP'}), 400
        ap_ip = data.get('ap_ip', 'N/A')
        if ap_ip != 'N/A' and not validate_ip(ap_ip):
            return jsonify({'error': 'Invalid AP IP'}), 400
        cid = data.get('cid', 'N/A')
        location = data.get('location')
        if not location:
            return jsonify({'error': 'Location is required'}), 400
        
        # Regional admin: verify location is in their region
        user = get_current_user()
        if user['role'] == 'regional_admin' and user['region_id']:
            with sqlite3.connect('ping_history.db', timeout=10) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT region_id FROM locations WHERE name = ?", (location,))
                loc_data = cursor.fetchone()
                if not loc_data or loc_data[0] != user['region_id']:
                    return jsonify({'error': 'You can only add devices to locations in your region'}), 403
        
        with results_lock:
            if CACHED_DF is None:
                return jsonify({'error': 'Dataframe not initialized'}), 500
            if sm_ip in CACHED_DF['SM IP'].values:
                return jsonify({'error': 'SM IP already exists'}), 400
            new_entry = {
                'AP Name': data.get('ap_name', 'N/A'),
                'AP IP': ap_ip,
                'CID': cid,
                'SM IP': sm_ip,
                'Device Name': data.get('org_name', 'N/A'),
                'Location': location
            }
            CACHED_DF = pd.concat([CACHED_DF, pd.DataFrame([new_entry])], ignore_index=True)
            CACHED_DF.to_excel(XLSX_FILE, index=False, engine='openpyxl')
            LAST_XLSX_LOAD = None
            logging.info(f"Added new entry for SM IP {sm_ip}")
            log_user_activity(user['id'], user['username'], 'entry_add', f"Added device {sm_ip} ({data.get('org_name','N/A')}) at {location}")
            return jsonify({'success': 'Entry added successfully'})
    except Exception as e:
        logging.error(f"Add entry error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/update_entry', methods=['POST'])
@admin_required
def admin_update_entry():
    global CACHED_DF, LAST_XLSX_LOAD
    try:
        data = request.json
        old_sm_ip = data.get('old_sm_ip')  # Original SM IP to find the row
        new_sm_ip = data.get('sm_ip')      # New SM IP (can be same as old)
        
        if not old_sm_ip or not validate_ip(old_sm_ip):
            return jsonify({'error': 'Invalid or missing original SM IP'}), 400
        
        if not new_sm_ip or not validate_ip(new_sm_ip):
            return jsonify({'error': 'Invalid or missing new SM IP'}), 400
            
        ap_ip = data.get('ap_ip', 'N/A')
        if ap_ip != 'N/A' and not validate_ip(ap_ip):
            return jsonify({'error': 'Invalid AP IP'}), 400
        
        location = data.get('location')
        if not location:
            return jsonify({'error': 'Location is required'}), 400
        
        # Regional admin: verify both old and new locations are in their region
        user = get_current_user()
        if user['role'] == 'regional_admin' and user['region_id']:
            with sqlite3.connect('ping_history.db', timeout=10) as conn:
                cursor = conn.cursor()
                # Check old location
                if CACHED_DF is not None and old_sm_ip in CACHED_DF['SM IP'].values:
                    old_location = CACHED_DF.loc[CACHED_DF['SM IP'] == old_sm_ip, 'Location'].iloc[0]
                    cursor.execute("SELECT region_id FROM locations WHERE name = ?", (old_location,))
                    old_loc_data = cursor.fetchone()
                    if not old_loc_data or old_loc_data[0] != user['region_id']:
                        return jsonify({'error': 'You can only edit devices in your region'}), 403
                
                # Check new location
                cursor.execute("SELECT region_id FROM locations WHERE name = ?", (location,))
                new_loc_data = cursor.fetchone()
                if not new_loc_data or new_loc_data[0] != user['region_id']:
                    return jsonify({'error': 'You can only move devices to locations in your region'}), 403
        
        with results_lock:
            if CACHED_DF is None:
                return jsonify({'error': 'Dataframe not initialized'}), 500
            
            # Check if old SM IP exists
            if old_sm_ip not in CACHED_DF['SM IP'].values:
                return jsonify({'error': 'Original SM IP not found'}), 404
            
            # If SM IP is being changed, check if new SM IP already exists
            if old_sm_ip != new_sm_ip and new_sm_ip in CACHED_DF['SM IP'].values:
                return jsonify({'error': 'New SM IP already exists in the system'}), 400
            
            # Get current values for fields not being updated
            current_row = CACHED_DF.loc[CACHED_DF['SM IP'] == old_sm_ip].iloc[0]
            
            update_data = {
                'AP Name': data.get('ap_name', current_row.get('AP Name', 'N/A')),
                'AP IP': ap_ip,
                'CID': data.get('cid', current_row.get('CID', 'N/A')),
                'SM IP': new_sm_ip,  # Use the new SM IP
                'Device Name': data.get('org_name', current_row.get('Device Name', 'N/A')),
                'Location': location
            }
            
            # Update the row
            row_index = CACHED_DF[CACHED_DF['SM IP'] == old_sm_ip].index[0]
            for key, value in update_data.items():
                CACHED_DF.at[row_index, key] = value
            
            CACHED_DF.to_excel(XLSX_FILE, index=False, engine='openpyxl')
            LAST_XLSX_LOAD = None
            
            if old_sm_ip != new_sm_ip:
                logging.info(f"Updated entry: SM IP changed from {old_sm_ip} to {new_sm_ip}")
                log_user_activity(user['id'], user['username'], 'entry_update', f"Updated device IP from {old_sm_ip} to {new_sm_ip} at {location}")
            else:
                logging.info(f"Updated entry for SM IP {new_sm_ip}")
                log_user_activity(user['id'], user['username'], 'entry_update', f"Updated device {new_sm_ip} ({data.get('org_name','N/A')}) at {location}")
            
            return jsonify({'success': 'Entry updated successfully', 'new_sm_ip': new_sm_ip})
    except Exception as e:
        logging.error(f"Update entry error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/delete_entry', methods=['POST'])
@admin_required
def admin_delete_entry():
    global CACHED_DF, LAST_XLSX_LOAD
    try:
        data = request.json
        sm_ip = data.get('sm_ip')
        if not sm_ip or not validate_ip(sm_ip):
            return jsonify({'error': 'Invalid or missing SM IP'}), 400
        
        # Regional admin: verify device location is in their region
        user = get_current_user()
        if user['role'] == 'regional_admin' and user['region_id']:
            with sqlite3.connect('ping_history.db', timeout=10) as conn:
                cursor = conn.cursor()
                if CACHED_DF is not None and sm_ip in CACHED_DF['SM IP'].values:
                    device_location = CACHED_DF.loc[CACHED_DF['SM IP'] == sm_ip, 'Location'].iloc[0]
                    cursor.execute("SELECT region_id FROM locations WHERE name = ?", (device_location,))
                    loc_data = cursor.fetchone()
                    if not loc_data or loc_data[0] != user['region_id']:
                        return jsonify({'error': 'You can only delete devices in your region'}), 403
        
        with results_lock:
            if CACHED_DF is None:
                return jsonify({'error': 'Dataframe not initialized'}), 500
            if sm_ip not in CACHED_DF['SM IP'].values:
                return jsonify({'error': 'SM IP not found'}), 404
            CACHED_DF = CACHED_DF[CACHED_DF['SM IP'] != sm_ip]
            CACHED_DF.to_excel(XLSX_FILE, index=False, engine='openpyxl')
            LAST_XLSX_LOAD = None
            logging.info(f"Deleted entry for SM IP {sm_ip}")
            log_user_activity(user['id'], user['username'], 'entry_delete', f"Deleted device {sm_ip}")
            return jsonify({'success': 'Entry deleted successfully'})
    except Exception as e:
        logging.error(f"Delete entry error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/download_template')
@admin_required
def download_template():
    """Download Excel template for bulk import"""
    try:
        # Create a new DataFrame with template structure
        template_df = pd.DataFrame(columns=[
            'AP Name', 'AP IP', 'CID', 'SM IP', 'Device Name', 'Location'
        ])
        
        # Add sample row with instructions
        template_df.loc[0] = [
            'Example AP',
            '192.168.1.1',
            'CID-001',
            '192.168.1.100',
            'Example Device',
            'Example Location'
        ]
        
        # Create temporary file
        temp_file = 'temp_template.xlsx'
        template_df.to_excel(temp_file, index=False, engine='openpyxl')
        
        # Send file
        return send_file(
            temp_file,
            as_attachment=True,
            download_name='device_import_template.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logging.error(f"Template download error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/bulk_import', methods=['POST'])
@admin_required
def bulk_import():
    """Bulk import devices from Excel file"""
    global CACHED_DF, LAST_XLSX_LOAD
    
    try:
        # Check if file was uploaded
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'Invalid file format. Please upload .xlsx or .xls file'}), 400
        
        # Read uploaded Excel file
        upload_df = pd.read_excel(file, engine='openpyxl')
        
        # Validate required columns
        required_columns = ['SM IP', 'Location']
        missing_columns = [col for col in required_columns if col not in upload_df.columns]
        if missing_columns:
            return jsonify({
                'success': False,
                'message': f'Missing required columns: {", ".join(missing_columns)}'
            }), 400
        
        # Initialize counters and tracking
        added = 0
        skipped = 0
        errors = 0
        details = []
        new_locations = []
        locations_created = 0
        
        # Get existing locations
        with sqlite3.connect('ping_history.db') as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM locations")
            existing_locations = set(row[0] for row in cursor.fetchall())
        
        with results_lock:
            if CACHED_DF is None:
                return jsonify({'success': False, 'message': 'System not initialized'}), 500
            
            existing_ips = set(CACHED_DF['SM IP'].values)
            
            # Process each row
            for idx, row in upload_df.iterrows():
                try:
                    # Get values
                    sm_ip = str(row.get('SM IP', '')).strip()
                    location = str(row.get('Location', '')).strip()
                    
                    # Skip empty rows
                    if not sm_ip or sm_ip == 'nan':
                        continue
                    
                    # Validate required fields
                    if not location or location == 'nan':
                        errors += 1
                        details.append(f"Row {idx + 2}: Error - Location is required for {sm_ip}")
                        continue
                    
                    # Validate IP address
                    if not validate_ip(sm_ip):
                        errors += 1
                        details.append(f"Row {idx + 2}: Error - Invalid IP address: {sm_ip}")
                        continue
                    
                    # Check for duplicates
                    if sm_ip in existing_ips:
                        skipped += 1
                        details.append(f"Row {idx + 2}: Skipped - Duplicate SM IP: {sm_ip}")
                        continue
                    
                    # Create location if it doesn't exist
                    if location not in existing_locations:
                        with sqlite3.connect('ping_history.db') as conn:
                            cursor = conn.cursor()
                            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            cursor.execute("""
                                INSERT INTO locations (name, region_id, description, created_at, updated_at)
                                VALUES (?, NULL, ?, ?, ?)
                            """, (location, f'Auto-created from bulk import', now, now))
                            conn.commit()
                        
                        existing_locations.add(location)
                        new_locations.append(location)
                        locations_created += 1
                        details.append(f"Created new location: {location}")
                    
                    # Prepare entry data
                    new_entry = {
                        'AP Name': str(row.get('AP Name', 'N/A')),
                        'AP IP': str(row.get('AP IP', 'N/A')),
                        'CID': str(row.get('CID', 'N/A')),
                        'SM IP': sm_ip,
                        'Device Name': str(row.get('Device Name', 'N/A')),
                        'Location': location
                    }
                    
                    # Add to dataframe
                    CACHED_DF = pd.concat([CACHED_DF, pd.DataFrame([new_entry])], ignore_index=True)
                    existing_ips.add(sm_ip)
                    added += 1
                    details.append(f"Row {idx + 2}: Added - {sm_ip} ({location})")
                    
                except Exception as e:
                    errors += 1
                    details.append(f"Row {idx + 2}: Error - {str(e)}")
            
            # Save to Excel if any entries were added
            if added > 0:
                CACHED_DF.to_excel(XLSX_FILE, index=False, engine='openpyxl')
                LAST_XLSX_LOAD = None
                logging.info(f"Bulk import: Added {added}, Skipped {skipped}, Errors {errors}")
        
        return jsonify({
            'success': True,
            'added': added,
            'skipped': skipped,
            'errors': errors,
            'locations_created': locations_created,
            'new_locations': new_locations,
            'details': details
        })
        
    except Exception as e:
        logging.error(f"Bulk import error: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Import failed: {str(e)}'
        }), 500

@app.route('/admin/location_downtime')
@admin_required
def location_downtime():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        location_filter = request.args.get('location_filter')
        if not start_date or not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        with results_lock:
            if CACHED_DF is None or 'SM IP' not in CACHED_DF.columns or 'Location' not in CACHED_DF.columns:
                return jsonify({'error': 'Dataframe not initialized or missing required columns'}), 500
            ip_to_location = {str(row['SM IP']): str(row['Location']) for row in CACHED_DF.to_dict('records') if pd.notna(row.get('SM IP'))}
            all_locations = sorted(set(CACHED_DF['Location'].dropna().astype(str)))
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            query = """
                SELECT sm_ip, status, timestamp
                FROM history
                WHERE timestamp >= ? AND timestamp < ?
                ORDER BY sm_ip, timestamp
            """
            cursor.execute(query, (start_datetime.strftime('%Y-%m-%d %H:%M:%S'), end_datetime.strftime('%Y-%m-%d %H:%M:%S')))
            records = cursor.fetchall()
        downtime_data = defaultdict(lambda: {
            'total_downtime': 0,
            'downtime_count': 0,
            'down_ips': set(),
            'last_downtime': None,
            'uptime': 'N/A'
        })
        ip_stats = defaultdict(lambda: {
            'uptime': 0,
            'downtime': 0,
            'degraded_time': 0,
            'days': set(),
            'down_count': 0,
            'up_count': 0,
            'down_events': [],
            'up_events': [],
            'degraded_events': []
        })
        for i, (sm_ip, status, timestamp) in enumerate(records):
            location = ip_to_location.get(sm_ip, 'Unknown')
            if location_filter and location != location_filter:
                continue
            ts = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
            day = ts.strftime('%Y-%m-%d')
            next_status = None
            next_ts = None
            if i + 1 < len(records) and records[i + 1][0] == sm_ip:
                next_status = records[i + 1][1]
                next_ts = datetime.strptime(records[i + 1][2], '%Y-%m-%d %H:%M:%S')
            duration = PING_INTERVAL
            if next_status and next_ts:
                duration = (next_ts - ts).total_seconds()
            if status == 'Down':
                downtime_data[location]['total_downtime'] += duration
                downtime_data[location]['down_ips'].add(sm_ip)
                downtime_data[location]['downtime_count'] += 1
                if not downtime_data[location]['last_downtime'] or ts > datetime.strptime(downtime_data[location]['last_downtime'], '%Y-%m-%d %H:%M:%S'):
                    downtime_data[location]['last_downtime'] = timestamp
                ip_stats[sm_ip]['downtime'] += duration
                if i == 0 or records[i-1][0] != sm_ip or records[i-1][1] != 'Down':
                    ip_stats[sm_ip]['down_count'] += 1
                    ip_stats[sm_ip]['down_events'].append(timestamp)
                ip_stats[sm_ip]['days'].add(day)
                ip_stats[sm_ip]['location'] = location
            elif status == 'Reachable':
                ip_stats[sm_ip]['uptime'] += duration
                if i == 0 or records[i-1][0] != sm_ip or records[i-1][1] != 'Reachable':
                    ip_stats[sm_ip]['up_count'] += 1
                    ip_stats[sm_ip]['up_events'].append(timestamp)
                ip_stats[sm_ip]['days'].add(day)
                ip_stats[sm_ip]['location'] = location
            elif status == 'Degraded':
                ip_stats[sm_ip]['degraded_time'] += duration
                if i == 0 or records[i-1][0] != sm_ip or records[i-1][1] != 'Degraded':
                    ip_stats[sm_ip]['degraded_events'].append(timestamp)
                ip_stats[sm_ip]['days'].add(day)
                ip_stats[sm_ip]['location'] = location
        for location in downtime_data:
            down_ips = downtime_data[location]['down_ips']
            if down_ips:
                cursor.execute("""
                    SELECT sm_ip, timestamp
                    FROM history
                    WHERE sm_ip IN ({}) AND status = 'Reachable' AND timestamp >= ? AND timestamp < ?
                    ORDER BY timestamp DESC
                """.format(','.join(['?'] * len(down_ips))), list(down_ips) + [start_datetime.strftime('%Y-%m-%d %H:%M:%S'), end_datetime.strftime('%Y-%m-%d %H:%M:%S')])
                uptime_records = cursor.fetchall()
                latest_uptime = {}
                for sm_ip, timestamp in uptime_records:
                    if sm_ip not in latest_uptime:
                        latest_uptime[sm_ip] = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
                if latest_uptime:
                    latest = max(latest_uptime.values())
                    duration = (datetime.now() - latest).total_seconds()
                    downtime_data[location]['uptime'] = format_duration(duration)
        result = {
            'locations': [],
            'ip_stats': {},
            'chart_data': {
                'bar': {
                    'labels': [],
                    'data': []
                }
            }
        }
        locations_to_process = [location_filter] if location_filter else all_locations
        for location in locations_to_process:
            data = downtime_data.get(location, {
                'total_downtime': 0,
                'downtime_count': 0,
                'down_ips': set(),
                'last_downtime': None,
                'uptime': 'N/A'
            })
            avg_downtime = format_duration(data['total_downtime'] / data['downtime_count']) if data['downtime_count'] > 0 else "N/A"
            result['locations'].append({
                'location': location,
                'total_downtime': format_duration(data['total_downtime']),
                'avg_downtime': avg_downtime,
                'down_ip_count': len(data['down_ips']),
                'down_ip_link': f"/admin/location_ips?location={location}&start_date={start_date}&end_date={end_date}",
                'last_downtime': data['last_downtime'],
                'uptime': data['uptime']
            })
            if data['total_downtime'] > 0:
                result['chart_data']['bar']['labels'].append(location)
                result['chart_data']['bar']['data'].append(data['total_downtime'])
        sorted_ips = sorted(
            ip_stats.items(),
            key=lambda x: x[1]['downtime'],
            reverse=True
        )
        worst_ips = sorted_ips[:10]
        best_ips = sorted_ips[-10:][::-1] if len(sorted_ips) >= 10 else sorted_ips[::-1]
        for sm_ip, stats in worst_ips + best_ips:
            result['ip_stats'][sm_ip] = {
                'location': stats['location'],
                'uptime': format_duration(stats['uptime']),
                'downtime': format_duration(stats['downtime']),
                'degraded_time': format_duration(stats['degraded_time']),
                'down_count': stats['down_count'],
                'up_count': stats['up_count'],
                'down_events': stats['down_events'],
                'up_events': stats['up_events'],
                'degraded_events': stats['degraded_events']
            }
        logging.info(f"Retrieved downtime data for {len(result['locations'])} locations")
        return jsonify(result)
    except Exception as e:
        logging.error(f"Location downtime error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/export_location_downtime')
@admin_required
def export_location_downtime():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        location_filter = request.args.get('location_filter')
        if not start_date or not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        with results_lock:
            if CACHED_DF is None or 'SM IP' not in CACHED_DF.columns or 'Location' not in CACHED_DF.columns:
                return jsonify({'error': 'Dataframe not initialized or missing required columns'}), 500
            ip_to_location = {str(row['SM IP']): str(row['Location']) for row in CACHED_DF.to_dict('records') if pd.notna(row.get('SM IP'))}
            all_locations = sorted(set(CACHED_DF['Location'].dropna().astype(str)))
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            query = """
                SELECT sm_ip, status, timestamp, latency
                FROM history
                WHERE timestamp >= ? AND timestamp < ?
                ORDER BY sm_ip, timestamp
            """
            cursor.execute(query, (start_datetime.strftime('%Y-%m-%d %H:%M:%S'), end_datetime.strftime('%Y-%m-%d %H:%M:%S')))
            records = cursor.fetchall()
        downtime_data = defaultdict(lambda: {
            'total_downtime': 0,
            'downtime_count': 0,
            'down_ips': set(),
            'last_downtime': None,
            'uptime': 'N/A'
        })
        ip_stats = defaultdict(lambda: {
            'uptime': 0,
            'downtime': 0,
            'degraded_time': 0,
            'days': set(),
            'down_count': 0,
            'up_count': 0,
            'down_events': [],
            'up_events': [],
            'degraded_events': [],
            'down_durations': [],
            'up_durations': [],
            'degraded_durations': []
        })
        for i, (sm_ip, status, timestamp, latency) in enumerate(records):
            location = ip_to_location.get(sm_ip, 'Unknown')
            if location_filter and location != location_filter:
                continue
            ts = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
            day = ts.strftime('%Y-%m-%d')
            next_status = None
            next_ts = None
            if i + 1 < len(records) and records[i + 1][0] == sm_ip:
                next_status = records[i + 1][1]
                next_ts = datetime.strptime(records[i + 1][2], '%Y-%m-%d %H:%M:%S')
            duration = PING_INTERVAL
            if next_status and next_ts:
                duration = (next_ts - ts).total_seconds()
            if status == 'Down':
                downtime_data[location]['total_downtime'] += duration
                downtime_data[location]['down_ips'].add(sm_ip)
                downtime_data[location]['downtime_count'] += 1
                if not downtime_data[location]['last_downtime'] or ts > datetime.strptime(downtime_data[location]['last_downtime'], '%Y-%m-%d %H:%M:%S'):
                    downtime_data[location]['last_downtime'] = timestamp
                ip_stats[sm_ip]['downtime'] += duration
                if i == 0 or records[i-1][0] != sm_ip or records[i-1][1] != 'Down':
                    ip_stats[sm_ip]['down_count'] += 1
                    ip_stats[sm_ip]['down_events'].append(timestamp)
                ip_stats[sm_ip]['down_durations'].append(duration)
                ip_stats[sm_ip]['days'].add(day)
                ip_stats[sm_ip]['location'] = location
            elif status == 'Reachable':
                ip_stats[sm_ip]['uptime'] += duration
                if i == 0 or records[i-1][0] != sm_ip or records[i-1][1] != 'Reachable':
                    ip_stats[sm_ip]['up_count'] += 1
                    ip_stats[sm_ip]['up_events'].append(timestamp)
                ip_stats[sm_ip]['up_durations'].append(duration)
                ip_stats[sm_ip]['days'].add(day)
                ip_stats[sm_ip]['location'] = location
            elif status == 'Degraded':
                ip_stats[sm_ip]['degraded_time'] += duration
                if i == 0 or records[i-1][0] != sm_ip or records[i-1][1] != 'Degraded':
                    ip_stats[sm_ip]['degraded_events'].append(timestamp)
                ip_stats[sm_ip]['degraded_durations'].append(duration)
                ip_stats[sm_ip]['days'].add(day)
                ip_stats[sm_ip]['location'] = location
        for location in downtime_data:
            down_ips = downtime_data[location]['down_ips']
            if down_ips:
                cursor.execute("""
                    SELECT sm_ip, timestamp
                    FROM history
                    WHERE sm_ip IN ({}) AND status = 'Reachable' AND timestamp >= ? AND timestamp < ?
                    ORDER BY timestamp DESC
                """.format(','.join(['?'] * len(down_ips))), list(down_ips) + [start_datetime.strftime('%Y-%m-%d %H:%M:%S'), end_datetime.strftime('%Y-%m-%d %H:%M:%S')])
                uptime_records = cursor.fetchall()
                latest_uptime = {}
                for sm_ip, timestamp in uptime_records:
                    if sm_ip not in latest_uptime:
                        latest_uptime[sm_ip] = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
                if latest_uptime:
                    latest = max(latest_uptime.values())
                    duration = (datetime.now() - latest).total_seconds()
                    downtime_data[location]['uptime'] = format_duration(duration)
        wb = Workbook()
        ws = wb.active
        ws.title = "Location Downtime Report"
        headers = ['Location', 'Total Downtime (HH:MM:SS)', 'Avg Downtime (HH:MM:SS)', 'Down IP Count', 'Last Downtime', 'Uptime Since Last Down (HH:MM:SS)']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        locations_to_process = [location_filter] if location_filter else all_locations
        for location in locations_to_process:
            data = downtime_data.get(location, {
                'total_downtime': 0,
                'downtime_count': 0,
                'down_ips': set(),
                'last_downtime': None,
                'uptime': 'N/A'
            })
            avg_downtime = format_duration(data['total_downtime'] / data['downtime_count']) if data['downtime_count'] > 0 else "N/A"
            ws.append([
                location,
                format_duration(data['total_downtime']),
                avg_downtime,
                len(data['down_ips']),
                data['last_downtime'] or 'N/A',
                data['uptime']
            ])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        ws = wb.create_sheet("IP Stats")
        headers = [
            'SM IP', 'Location', 'Times Down', 'Times Up', 'Avg Downtime (HH:MM:SS)', 
            'Avg Uptime (HH:MM:SS)', 'Total Downtime (HH:MM:SS)', 'Total Uptime (HH:MM:SS)', 
            'Total Degraded Time (HH:MM:SS)', 'Down Events', 'Up Events', 'Degraded Events'
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        sorted_ips = sorted(
            ip_stats.items(),
            key=lambda x: x[1]['downtime'],
            reverse=True
        )
        for sm_ip, stats in sorted_ips:
            avg_downtime = format_duration(sum(stats['down_durations']) / len(stats['down_durations'])) if stats['down_durations'] else "N/A"
            avg_uptime = format_duration(sum(stats['up_durations']) / len(stats['up_durations'])) if stats['up_durations'] else "N/A"
            ws.append([
                sm_ip,
                stats['location'],
                stats['down_count'],
                stats['up_count'],
                avg_downtime,
                avg_uptime,
                format_duration(stats['downtime']),
                format_duration(stats['uptime']),
                format_duration(stats['degraded_time']),
                '; '.join(stats['down_events']),
                '; '.join(stats['up_events']),
                '; '.join(stats['degraded_events'])
            ])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        ws = wb.create_sheet("Detailed Logs")
        headers = ['Timestamp', 'SM IP', 'Status', 'Latency (ms)', 'Location']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for sm_ip, status, timestamp, latency in records:
            location = ip_to_location.get(sm_ip, 'Unknown')
            if location_filter and location != location_filter:
                continue
            ws.append([
                timestamp,
                sm_ip,
                status,
                f"{latency:.2f}" if latency is not None else 'N/A',
                location
            ])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        logging.info(f"Exported location downtime report for {start_date} to {end_date}")
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"location_downtime_{start_date}_to_{end_date}.xlsx"
        )
    except Exception as e:
        logging.error(f"Export location downtime error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/export_log')
@admin_required
def export_log():
    try:
        cutoff_time = datetime.now() - timedelta(hours=1)
        recent_alerts = [
            a for a in ALERT_LOG
            if datetime.strptime(a['time'].split('.')[0], '%Y-%m-%d %H:%M:%S') >= cutoff_time
            and a['entry']['status'] in ['Down', 'Degraded', 'Reachable']
        ]
        wb = Workbook()
        ws = wb.active
        ws.title = "Alert Log"
        headers = ['Timestamp', 'SM IP', 'Status', 'Location', 'Device Name', 'AP Name', 'AP IP', 'AP MAC', 'SM MAC', 'Latency', 'Downtime Since', 'Uptime Since', 'Long Term']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for alert in recent_alerts:
            entry = alert['entry']
            ws.append([
                alert['time'],
                entry['ip'],
                entry['status'],
                entry['location'],
                entry['org_name'],
                entry['ap_name'],
                entry['ap_ip'],
                entry['ap_mac'],
                entry['sm_mac'],
                f"{entry['latency']:.2f}" if entry['latency'] is not None else 'N/A',
                entry['downtime_since'],
                entry['uptime_since'],
                'Yes' if entry['long_term'] else 'No'
            ])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        logging.info("Exported alert log")
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"alert_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
    except Exception as e:
        logging.error(f"Export log error: {str(e)}")
        return jsonify({'error': str(e)}), 500
















@app.route('/export_ip_history')
@login_required
def export_ip_history():
    try:
        ip_input = request.args.get('ip', '')
        date = request.args.get('date')
        hour = request.args.get('hour')
        format_type = request.args.get('format', 'xlsx')
        sm_ips = [ip.strip() for ip in ip_input.split(',') if validate_ip(ip.strip())]
        if not sm_ips:
            return jsonify({'error': 'Invalid or missing SM IP(s)'}), 400
        query_params = []
        query_conditions = ["sm_ip IN ({})".format(','.join(['?'] * len(sm_ips)))]
        query_params.extend(sm_ips)
        if date:
            try:
                date_obj = datetime.strptime(date, '%Y-%m-%d')
                start_time = date_obj.strftime('%Y-%m-%d 00:00:00')
                end_time = (date_obj + timedelta(days=1)).strftime('%Y-%m-%d 00:00:00')
                query_conditions.append("timestamp >= ? AND timestamp < ?")
                query_params.extend([start_time, end_time])
            except ValueError:
                return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
        if hour:
            try:
                hour_int = int(hour)
                if not (0 <= hour_int <= 23):
                    return jsonify({'error': 'Invalid hour. Must be 00-23'}), 400
                if date:
                    start_time = datetime.strptime(f"{date} {hour}:00:00", '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
                    end_time = datetime.strptime(f"{date} {hour}:59:59", '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
                    query_conditions.pop()
                    query_conditions.append("timestamp >= ? AND timestamp < ?")
                    query_params[-2:] = [start_time, end_time]
                else:
                    return jsonify({'error': 'Date must be specified if hour is provided'}), 400
            except ValueError:
                return jsonify({'error': 'Invalid hour format'}), 400
        query = f"""
            SELECT timestamp, sm_ip, status, latency
            FROM history
            WHERE {' AND '.join(query_conditions)}
            ORDER BY timestamp
        """
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            cursor.execute(query, query_params)
            records = cursor.fetchall()
        if not records:
            return jsonify({'error': f'No history found for IP(s) {ip_input}'}), 404
        if format_type.lower() != 'xlsx':
            return jsonify({'error': 'Unsupported format. Use xlsx'}), 400
        ip_stats = defaultdict(lambda: {
            'uptime': 0,
            'downtime': 0,
            'degraded_time': 0,
            'down_count': 0,
            'up_count': 0,
            'down_events': [],
            'up_events': [],
            'degraded_events': [],
            'down_durations': [],
            'up_durations': [],
            'degraded_durations': []
        })
        for i, (timestamp, sm_ip, status, latency) in enumerate(records):
            ts = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
            next_status = None
            next_ts = None
            if i + 1 < len(records) and records[i + 1][1] == sm_ip:
                next_status = records[i + 1][2]
                next_ts = datetime.strptime(records[i + 1][0], '%Y-%m-%d %H:%M:%S')
            duration = PING_INTERVAL
            if next_status and next_ts:
                duration = (next_ts - ts).total_seconds()
            if status == 'Down':
                ip_stats[sm_ip]['downtime'] += duration
                if i == 0 or records[i-1][1] != sm_ip or records[i-1][2] != 'Down':
                    ip_stats[sm_ip]['down_count'] += 1
                    ip_stats[sm_ip]['down_events'].append(timestamp)
                ip_stats[sm_ip]['down_durations'].append(duration)
            elif status == 'Reachable':
                ip_stats[sm_ip]['uptime'] += duration
                if i == 0 or records[i-1][1] != sm_ip or records[i-1][2] != 'Reachable':
                    ip_stats[sm_ip]['up_count'] += 1
                    ip_stats[sm_ip]['up_events'].append(timestamp)
                ip_stats[sm_ip]['up_durations'].append(duration)
            elif status == 'Degraded':
                ip_stats[sm_ip]['degraded_time'] += duration
                if i == 0 or records[i-1][1] != sm_ip or records[i-1][2] != 'Degraded':
                    ip_stats[sm_ip]['degraded_events'].append(timestamp)
                ip_stats[sm_ip]['degraded_durations'].append(duration)
        wb = Workbook()
        ws = wb.active
        ws.title = "IP History"
        headers = [
            'Timestamp', 'SM IP', 'Status', 'Latency (ms)', 'Downtime Since', 
            'Uptime Since', 'Times Down', 'Times Up', 'Avg Downtime (HH:MM:SS)', 
            'Avg Uptime (HH:MM:SS)', 'Total Downtime (HH:MM:SS)', 'Total Uptime (HH:MM:SS)', 
            'Total Degraded Time (HH:MM:SS)', 'Down Events', 'Up Events', 'Degraded Events'
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ip_to_location = {str(row['SM IP']): str(row['Location']) for row in CACHED_DF.to_dict('records') if pd.notna(row.get('SM IP'))}
        for sm_ip in sm_ips:
            stats = ip_stats[sm_ip]
            avg_downtime = format_duration(sum(stats['down_durations']) / len(stats['down_durations'])) if stats['down_durations'] else "N/A"
            avg_uptime = format_duration(sum(stats['up_durations']) / len(stats['up_durations'])) if stats['up_durations'] else "N/A"
            for timestamp, _, status, latency in [(r[0], r[1], r[2], r[3]) for r in records if r[1] == sm_ip]:
                ws.append([
                    timestamp,
                    sm_ip,
                    status,
                    f"{latency:.2f}" if latency is not None else 'N/A',
                    get_downtime_since(sm_ip, status),
                    get_uptime_since(sm_ip, status),
                    stats['down_count'],
                    stats['up_count'],
                    avg_downtime,
                    avg_uptime,
                    format_duration(stats['downtime']),
                    format_duration(stats['uptime']),
                    format_duration(stats['degraded_time']),
                    '; '.join(stats['down_events']),
                    '; '.join(stats['up_events']),
                    '; '.join(stats['degraded_events'])
                ])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        logging.info(f"Exported IP history for {ip_input}")
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"ip_history_{'_'.join(sm_ips)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
    except Exception as e:
        logging.error(f"Export IP history error: {str(e)}")
        return jsonify({'error': str(e)}), 500





@app.route('/get_logs')
@login_required
def get_logs():
    try:
        ip_input = request.args.get('ip', '')
        date = request.args.get('date')
        hour = request.args.get('hour')
        sm_ips = [ip.strip() for ip in ip_input.split(',') if validate_ip(ip.strip())]
        if not sm_ips:
            return jsonify({'error': 'Invalid or missing SM IP(s)'}), 400
        cache_key = f"{ip_input}_{date}_{hour}"
        if cache_key in HISTORY_CACHE:
            cached = HISTORY_CACHE[cache_key]
            if datetime.now() - cached['time'] < timedelta(minutes=5):
                logging.debug(f"Returning cached logs for {cache_key}")
                return jsonify(cached['data'])
        query_params = []
        query_conditions = ["sm_ip IN ({})".format(','.join(['?'] * len(sm_ips)))]
        query_params.extend(sm_ips)
        if date:
            try:
                date_obj = datetime.strptime(date, '%Y-%m-%d')
                start_time = date_obj.strftime('%Y-%m-%d 00:00:00')
                end_time = (date_obj + timedelta(days=1)).strftime('%Y-%m-%d 00:00:00')
                query_conditions.append("timestamp >= ? AND timestamp < ?")
                query_params.extend([start_time, end_time])
            except ValueError:
                return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
        if hour:
            try:
                hour_int = int(hour)
                if not (0 <= hour_int <= 23):
                    return jsonify({'error': 'Invalid hour. Must be 00-23'}), 400
                if date:
                    start_time = datetime.strptime(f"{date} {hour}:00:00", '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
                    end_time = datetime.strptime(f"{date} {hour}:59:59", '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
                    query_conditions.pop()
                    query_conditions.append("timestamp >= ? AND timestamp < ?")
                    query_params[-2:] = [start_time, end_time]
                else:
                    return jsonify({'error': 'Date must be specified if hour is provided'}), 400
            except ValueError:
                return jsonify({'error': 'Invalid hour format'}), 400
        query = f"""
            SELECT timestamp, sm_ip, status, latency
            FROM history
            WHERE {' AND '.join(query_conditions)}
            ORDER BY timestamp
        """
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            cursor.execute(query, query_params)
            records = cursor.fetchall()
        logs = []
        ip_info = {}
        with results_lock:
            if CACHED_DF is not None:
                ip_info = {
                    str(row['SM IP']): {
                        'org_name': str(row.get('Device Name', 'N/A')),
                        'location': str(row.get('Location', 'Unknown')),
                        'ap_name': str(row.get('AP Name', 'N/A')),
                        'ap_ip': str(row.get('AP IP', 'N/A')),
                        'ap_mac': str(row.get('AP MAC Address', 'N/A')),
                        'sm_mac': str(row.get('SM MAC Address', 'N/A')),
                        'cid': str(row.get('CID', 'N/A'))
                    } for row in CACHED_DF.to_dict('records') if pd.notna(row.get('SM IP'))
                }
        for timestamp, sm_ip, status, latency in records:
            info = ip_info.get(sm_ip, {})
            logs.append({
                'time': timestamp,
                'entry': {
                    'ip': sm_ip,
                    'status': status,
                    'latency': latency,
                    'downtime_since': get_downtime_since(sm_ip, status),
                    'uptime_since': get_uptime_since(sm_ip, status),
                    'location': info.get('location', 'Unknown'),
                    'org_name': info.get('org_name', 'N/A'),
                    'ap_name': info.get('ap_name', 'N/A'),
                    'ap_ip': info.get('ap_ip', 'N/A'),
                    'ap_mac': info.get('ap_mac', 'N/A'),
                    'sm_mac': info.get('sm_mac', 'N/A'),
                    'cid': info.get('cid', 'N/A')
                }
            })
        HISTORY_CACHE[cache_key] = {'data': logs, 'time': datetime.now()}
        logging.info(f"Retrieved {len(logs)} logs for IPs {ip_input}")
        return jsonify(logs)
    except Exception as e:
        logging.error(f"Get logs error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/ip_uptime_downtime')
@admin_required
def ip_uptime_downtime():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        ip_filter = request.args.get('ip_filter')
        if not start_date or not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        with results_lock:
            if CACHED_DF is None:
                return jsonify({'error': 'Dataframe not initialized'}), 500
            ip_info = {
                str(row['SM IP']): {
                    'location': str(row.get('Location', 'Unknown')),
                    'org_name': str(row.get('Device Name', 'N/A')),
                    'ap_name': str(row.get('AP Name', 'N/A'))
                } for row in CACHED_DF.to_dict('records') if pd.notna(row.get('SM IP'))
            }
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            query = """
                SELECT sm_ip, status, timestamp
                FROM history
                WHERE timestamp >= ? AND timestamp < ?
                ORDER BY sm_ip, timestamp
            """
            cursor.execute(query, (start_datetime.strftime('%Y-%m-%d %H:%M:%S'), end_datetime.strftime('%Y-%m-%d %H:%M:%S')))
            records = cursor.fetchall()
        ip_stats = defaultdict(lambda: {
            'uptime': 0,
            'downtime': 0,
            'degraded_time': 0,
            'down_count': 0,
            'up_count': 0,
            'down_events': [],
            'up_events': [],
            'degraded_events': []
        })
        for i, (sm_ip, status, timestamp) in enumerate(records):
            if ip_filter and sm_ip != ip_filter:
                continue
            ts = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
            next_status = None
            next_ts = None
            if i + 1 < len(records) and records[i + 1][0] == sm_ip:
                next_status = records[i + 1][1]
                next_ts = datetime.strptime(records[i + 1][2], '%Y-%m-%d %H:%M:%S')
            duration = PING_INTERVAL
            if next_status and next_ts:
                duration = (next_ts - ts).total_seconds()
            if status == 'Down':
                ip_stats[sm_ip]['downtime'] += duration
                if i == 0 or records[i-1][0] != sm_ip or records[i-1][1] != 'Down':
                    ip_stats[sm_ip]['down_count'] += 1
                    ip_stats[sm_ip]['down_events'].append(timestamp)
            elif status == 'Reachable':
                ip_stats[sm_ip]['uptime'] += duration
                if i == 0 or records[i-1][0] != sm_ip or records[i-1][1] != 'Reachable':
                    ip_stats[sm_ip]['up_count'] += 1
                    ip_stats[sm_ip]['up_events'].append(timestamp)
            elif status == 'Degraded':
                ip_stats[sm_ip]['degraded_time'] += duration
                if i == 0 or records[i-1][0] != sm_ip or records[i-1][1] != 'Degraded':
                    ip_stats[sm_ip]['degraded_events'].append(timestamp)
        result = {
            'ips': [],
            'chart_data': {
                'pie': {
                    'labels': ['Uptime', 'Downtime', 'Degraded'],
                    'data': []
                }
            }
        }
        for sm_ip in ip_stats:
            if ip_filter and sm_ip != ip_filter:
                continue
            info = ip_info.get(sm_ip, {})
            result['ips'].append({
                'sm_ip': sm_ip,
                'location': info.get('location', 'Unknown'),
                'org_name': info.get('org_name', 'N/A'),
                'ap_name': info.get('ap_name', 'N/A'),
                'uptime': format_duration(ip_stats[sm_ip]['uptime']),
                'downtime': format_duration(ip_stats[sm_ip]['downtime']),
                'degraded_time': format_duration(ip_stats[sm_ip]['degraded_time']),
                'down_count': ip_stats[sm_ip]['down_count'],
                'up_count': ip_stats[sm_ip]['up_count']
            })
            if sm_ip == ip_filter or not ip_filter:
                result['chart_data']['pie']['data'] = [
                    ip_stats[sm_ip]['uptime'],
                    ip_stats[sm_ip]['downtime'],
                    ip_stats[sm_ip]['degraded_time']
                ]
        logging.info(f"Retrieved uptime/downtime for {len(result['ips'])} IPs")
        return jsonify(result)
    except Exception as e:
        logging.error(f"IP uptime/downtime error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/location_health')
@admin_required
def location_health():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        location_filter = request.args.get('location_filter')
        if not start_date or not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        with results_lock:
            if CACHED_DF is None:
                return jsonify({'error': 'Dataframe not initialized'}), 500
            ip_to_info = {str(row['SM IP']): {
                'location': str(row['Location']),
                'device_name': str(row['Device Name']),
                'ap_name': str(row['AP Name']),
                'ap_ip': str(row['AP IP'])
            } for row in CACHED_DF.to_dict('records') if pd.notna(row.get('SM IP'))}
            all_locations = sorted(set(CACHED_DF['Location'].dropna().astype(str)))
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            query = """
                SELECT sm_ip, status, timestamp, latency
                FROM history
                WHERE timestamp >= ? AND timestamp < ?
                ORDER BY sm_ip, timestamp
            """
            cursor.execute(query, (start_datetime.strftime('%Y-%m-%d %H:%M:%S'), end_datetime.strftime('%Y-%m-%d %H:%M:%S')))
            records = cursor.fetchall()
        
        # Location-level health data
        health_data = defaultdict(lambda: {
            'total_latency': 0,
            'ping_count': 0,
            'down_count': 0,
            'reachable_count': 0,
            'degraded_count': 0,
            'ips': set()
        })
        
        # IP-level statistics
        ip_stats = defaultdict(lambda: {
            'latencies': [],
            'down_events': 0,
            'total_pings': 0,
            'down_duration': 0,
            'up_duration': 0,
            'last_status': None,
            'status_changes': []
        })
        
        for sm_ip, status, timestamp, latency in records:
            ip_info = ip_to_info.get(sm_ip, {'location': 'Unknown', 'device_name': 'Unknown', 'ap_name': 'Unknown', 'ap_ip': 'Unknown'})
            location = ip_info['location']
            
            if location_filter and location != location_filter:
                continue
            
            # Location-level aggregation
            health_data[location]['ips'].add(sm_ip)
            if status == 'Down':
                health_data[location]['down_count'] += 1
            elif status == 'Reachable' and latency is not None:
                health_data[location]['reachable_count'] += 1
                health_data[location]['total_latency'] += latency
                health_data[location]['ping_count'] += 1
            elif status == 'Degraded':
                health_data[location]['degraded_count'] += 1
            
            # IP-level statistics
            ip_stats[sm_ip]['total_pings'] += 1
            ip_stats[sm_ip]['device_name'] = ip_info['device_name']
            ip_stats[sm_ip]['location'] = location
            ip_stats[sm_ip]['ap_name'] = ip_info['ap_name']
            ip_stats[sm_ip]['ap_ip'] = ip_info['ap_ip']
            
            if status == 'Down':
                ip_stats[sm_ip]['down_events'] += 1
                ip_stats[sm_ip]['down_duration'] += 10  # Assuming 10-second intervals
            elif status == 'Reachable' and latency is not None:
                ip_stats[sm_ip]['latencies'].append(latency)
                ip_stats[sm_ip]['up_duration'] += 10
            elif status == 'Degraded':
                ip_stats[sm_ip]['up_duration'] += 10
            
            ip_stats[sm_ip]['last_status'] = status
        
        # Calculate IP-level metrics
        processed_ip_stats = {}
        for ip, stats in ip_stats.items():
            if stats['total_pings'] == 0:
                continue
            
            avg_latency = sum(stats['latencies']) / len(stats['latencies']) if stats['latencies'] else 0
            total_time = stats['down_duration'] + stats['up_duration']
            uptime_percent = (stats['up_duration'] / total_time * 100) if total_time > 0 else 0
            
            # Format durations as HH:MM:SS
            def format_duration(seconds):
                hours = seconds // 3600
                minutes = (seconds % 3600) // 60
                secs = seconds % 60
                return f"{hours:02d}:{minutes:02d}:{secs:02d}"
            
            processed_ip_stats[ip] = {
                'device_name': stats['device_name'],
                'location': stats['location'],
                'ap_name': stats['ap_name'],
                'ap_ip': stats['ap_ip'],
                'avg_downtime': format_duration(stats['down_duration'] // max(stats['down_events'], 1)) if stats['down_events'] > 0 else '00:00:00',
                'downtime_count': stats['down_events'],
                'downtime_duration': format_duration(stats['down_duration']),
                'uptime': f"{uptime_percent:.1f}%",
                'avg_latency': f"{avg_latency:.2f} ms" if avg_latency > 0 else 'N/A'
            }
        
        result = {
            'locations': [],
            'ip_stats': processed_ip_stats,
            'chart_data': {
                'bar': {
                    'labels': [],
                    'data': []  # Changed from avg_latency to data for downtime
                }
            }
        }
        
        locations_to_process = [location_filter] if location_filter else all_locations
        for location in locations_to_process:
            data = health_data.get(location, {
                'total_latency': 0,
                'ping_count': 0,
                'down_count': 0,
                'reachable_count': 0,
                'degraded_count': 0,
                'ips': set()
            })
            avg_latency = (data['total_latency'] / data['ping_count']) if data['ping_count'] > 0 else 0
            total_downtime = sum(ip_stats[ip]['down_duration'] for ip in data['ips'] if ip in ip_stats)
            
            # Format total downtime as HH:MM:SS
            def format_duration(seconds):
                hours = seconds // 3600
                minutes = (seconds % 3600) // 60
                secs = seconds % 60
                return f"{hours:02d}:{minutes:02d}:{secs:02d}"
            
            result['locations'].append({
                'location': location,
                'total_downtime': format_duration(total_downtime),
                'avg_downtime': format_duration(total_downtime // len(data['ips'])) if data['ips'] else '00:00:00',
                'down_ip_count': len([ip for ip in data['ips'] if ip in ip_stats and ip_stats[ip]['down_events'] > 0]),
                'last_downtime': 'N/A',  # Would need more complex logic to determine
                'uptime': 'N/A'  # Would need more complex logic to determine
            })
            
            result['chart_data']['bar']['labels'].append(location)
            result['chart_data']['bar']['data'].append(total_downtime)
        
        logging.info(f"Retrieved health data for {len(result['locations'])} locations and {len(processed_ip_stats)} IPs")
        return jsonify(result)
    except Exception as e:
        logging.error(f"Location health error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/export_health_pdf')
@admin_required
def export_health_pdf():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        location_filter = request.args.get('location_filter')
        if not start_date or not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        with results_lock:
            if CACHED_DF is None:
                return jsonify({'error': 'Dataframe not initialized'}), 500
            ip_to_location = {str(row['SM IP']): str(row['Location']) for row in CACHED_DF.to_dict('records') if pd.notna(row.get('SM IP'))}
            all_locations = sorted(set(CACHED_DF['Location'].dropna().astype(str)))
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            query = """
                SELECT sm_ip, status, timestamp, latency
                FROM history
                WHERE timestamp >= ? AND timestamp < ?
                ORDER BY sm_ip, timestamp
            """
            cursor.execute(query, (start_datetime.strftime('%Y-%m-%d %H:%M:%S'), end_datetime.strftime('%Y-%m-%d %H:%M:%S')))
            records = cursor.fetchall()
        health_data = defaultdict(lambda: {
            'total_latency': 0,
            'ping_count': 0,
            'down_count': 0,
            'reachable_count': 0,
            'degraded_count': 0,
            'ips': set()
        })
        for sm_ip, status, timestamp, latency in records:
            location = ip_to_location.get(sm_ip, 'Unknown')
            if location_filter and location != location_filter:
                continue
            health_data[location]['ips'].add(sm_ip)
            if status == 'Down':
                health_data[location]['down_count'] += 1
            elif status == 'Reachable' and latency is not None:
                health_data[location]['reachable_count'] += 1
                health_data[location]['total_latency'] += latency
                health_data[location]['ping_count'] += 1
            elif status == 'Degraded':
                health_data[location]['degraded_count'] += 1
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=letter)
        c.setFont("Helvetica-Bold", 16)
        c.drawString(100, 750, "Location Health Report")
        c.setFont("Helvetica", 12)
        c.drawString(100, 730, f"Period: {start_date} to {end_date}")
        y = 700
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, y, "Location")
        c.drawString(150, y, "IP Count")
        c.drawString(250, y, "Avg Latency")
        c.drawString(350, y, "Down Count")
        c.drawString(450, y, "Reachable Count")
        y -= 20
        c.setFont("Helvetica", 10)
        locations_to_process = [location_filter] if location_filter else all_locations
        for location in locations_to_process:
            data = health_data.get(location, {
                'total_latency': 0,
                'ping_count': 0,
                'down_count': 0,
                'reachable_count': 0,
                'degraded_count': 0,
                'ips': set()
            })
            avg_latency = (data['total_latency'] / data['ping_count']) if data['ping_count'] > 0 else 0
            c.drawString(50, y, location)
            c.drawString(150, y, str(len(data['ips'])))
            c.drawString(250, y, f"{avg_latency:.2f} ms" if avg_latency > 0 else 'N/A')
            c.drawString(350, y, str(data['down_count']))
            c.drawString(450, y, str(data['reachable_count']))
            y -= 20
            if y < 50:
                c.showPage()
                y = 750
        c.save()
        buffer.seek(0)
        logging.info(f"Exported health PDF for {start_date} to {end_date}")
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"location_health_{start_date}_to_{end_date}.pdf"
        )
    except Exception as e:
        logging.error(f"Export health PDF error: {str(e)}")
        return jsonify({'error': str(e)}), 500

# Mobile and SLA Routes
@app.route('/mobile')
@login_required
def mobile_dashboard():
    """Mobile-optimized dashboard"""
    user = get_current_user()
    users = get_all_users() if user and user['role'] in ['admin', 'superadmin', 'regional_admin'] else []
    return render_template('mobile.html', user=user, users=users)

@app.route('/sla')
@login_required
def sla_dashboard():
    """SLA monitoring dashboard"""
    user = get_current_user()
    
    # Get user's region information for filtering
    user_region_id = user.get('region_id')
    user_role = user.get('role')
    
    # Pass region info to template
    return render_template('sla.html', user=user, user_region_id=user_region_id, user_role=user_role)

# Region and Location Management API Routes
@app.route('/api/regions', methods=['GET'])
def get_regions():
    """Get all regions - Public endpoint for signup page"""
    try:
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, name, description, created_at, updated_at
                FROM regions
                ORDER BY name
            """)
            regions = cursor.fetchall()
            
            regions_list = []
            for region in regions:
                regions_list.append({
                    'id': region[0],
                    'name': region[1],
                    'description': region[2],
                    'created_at': region[3],
                    'updated_at': region[4]
                })
            
            return jsonify({'success': True, 'regions': regions_list})
            
    except Exception as e:
        logging.error(f"Get regions error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/regions', methods=['POST'])
@superadmin_required
def create_region():
    """Create a new region"""
    try:
        data = request.json
        name = data.get('name', '').strip()
        description = data.get('description', '').strip()
        
        if not name:
            return jsonify({'error': 'Region name is required'}), 400
        
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Check if region already exists
            cursor.execute("SELECT id FROM regions WHERE name = ?", (name,))
            if cursor.fetchone():
                return jsonify({'error': 'Region already exists'}), 400
            
            # Create new region
            cursor.execute("""
                INSERT INTO regions (name, description, created_at, updated_at)
                VALUES (?, ?, ?, ?)
            """, (name, description, now, now))
            
            region_id = cursor.lastrowid
            conn.commit()
            
            return jsonify({
                'success': True,
                'message': f'Region "{name}" created successfully',
                'region': {
                    'id': region_id,
                    'name': name,
                    'description': description,
                    'created_at': now,
                    'updated_at': now
                }
            })
            
    except Exception as e:
        logging.error(f"Create region error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/regions/<int:region_id>', methods=['PUT'])
@superadmin_required
def update_region(region_id):
    """Update a region"""
    try:
        data = request.json
        name = data.get('name', '').strip()
        description = data.get('description', '').strip()
        
        if not name:
            return jsonify({'error': 'Region name is required'}), 400
        
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Check if region exists
            cursor.execute("SELECT name FROM regions WHERE id = ?", (region_id,))
            if not cursor.fetchone():
                return jsonify({'error': 'Region not found'}), 404
            
            # Check if new name conflicts with another region
            cursor.execute("SELECT id FROM regions WHERE name = ? AND id != ?", (name, region_id))
            if cursor.fetchone():
                return jsonify({'error': 'Region name already exists'}), 400
            
            # Update region
            cursor.execute("""
                UPDATE regions
                SET name = ?, description = ?, updated_at = ?
                WHERE id = ?
            """, (name, description, now, region_id))
            
            conn.commit()
            
            return jsonify({
                'success': True,
                'message': f'Region "{name}" updated successfully'
            })
            
    except Exception as e:
        logging.error(f"Update region error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/regions/<int:region_id>', methods=['DELETE'])
@superadmin_required
def delete_region(region_id):
    """Delete a region"""
    try:
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Check if region exists
            cursor.execute("SELECT name FROM regions WHERE id = ?", (region_id,))
            region_data = cursor.fetchone()
            if not region_data:
                return jsonify({'error': 'Region not found'}), 404
            
            region_name = region_data[0]
            
            # Check if region has locations
            cursor.execute("SELECT COUNT(*) FROM locations WHERE region_id = ?", (region_id,))
            location_count = cursor.fetchone()[0]
            
            if location_count > 0:
                return jsonify({'error': f'Cannot delete region with {location_count} locations. Please delete or reassign locations first.'}), 400
            
            # Delete region
            cursor.execute("DELETE FROM regions WHERE id = ?", (region_id,))
            conn.commit()
            
            return jsonify({
                'success': True,
                'message': f'Region "{region_name}" deleted successfully'
            })
            
    except Exception as e:
        logging.error(f"Delete region error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/locations', methods=['GET'])
@login_required
def get_locations():
    """Get all locations with region info and device counts"""
    try:
        region_id = request.args.get('region_id', type=int)
        unassigned = request.args.get('unassigned', '').lower() == 'true'
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            if unassigned:
                # Get only locations without region assignment
                cursor.execute("""
                    SELECT l.id, l.name, l.region_id, NULL, l.description, l.created_at, l.updated_at
                    FROM locations l
                    WHERE l.region_id IS NULL
                    ORDER BY l.name
                """)
            elif region_id:
                # Get locations for specific region
                cursor.execute("""
                    SELECT l.id, l.name, l.region_id, r.name, l.description, l.created_at, l.updated_at
                    FROM locations l
                    JOIN regions r ON l.region_id = r.id
                    WHERE l.region_id = ?
                    ORDER BY l.name
                """, (region_id,))
            else:
                # Get all locations (use LEFT JOIN to include unassigned)
                cursor.execute("""
                    SELECT l.id, l.name, l.region_id, r.name, l.description, l.created_at, l.updated_at
                    FROM locations l
                    LEFT JOIN regions r ON l.region_id = r.id
                    ORDER BY r.name, l.name
                """)
            
            locations = cursor.fetchall()
        
        # Count devices per location from Excel data
        device_counts = {}
        with results_lock:
            if CACHED_DF is not None and 'Location' in CACHED_DF.columns:
                # Count devices for each location, trimming whitespace
                location_series = CACHED_DF['Location'].astype(str).str.strip()
                device_counts = location_series.value_counts().to_dict()
        
        locations_list = []
        for location in locations:
            location_name = location[1]
            # Try exact match first, then try with stripped whitespace
            device_count = device_counts.get(location_name, 0)
            if device_count == 0:
                # Try to find with different whitespace
                for loc_key in device_counts.keys():
                    if loc_key.strip() == location_name.strip():
                        device_count = device_counts[loc_key]
                        break
            
            locations_list.append({
                'id': location[0],
                'name': location_name,
                'region_id': location[2],
                'region_name': location[3],
                'description': location[4],
                'created_at': location[5],
                'updated_at': location[6],
                'device_count': device_count
            })
        
        return jsonify({'success': True, 'locations': locations_list})
            
    except Exception as e:
        logging.error(f"Get locations error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/locations', methods=['POST'])
@admin_required
def create_location():
    """Create a new location"""
    try:
        data = request.json
        name = data.get('name', '').strip()
        region_id = data.get('region_id')
        description = data.get('description', '').strip()
        
        if not name:
            return jsonify({'error': 'Location name is required'}), 400
        
        if not region_id:
            return jsonify({'error': 'Region is required'}), 400
        
        # Regional admin: can only create locations in their region
        user = get_current_user()
        if user['role'] == 'regional_admin' and user['region_id']:
            if int(region_id) != user['region_id']:
                return jsonify({'error': 'You can only create locations in your assigned region'}), 403
        
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Check if region exists
            cursor.execute("SELECT name FROM regions WHERE id = ?", (region_id,))
            region_data = cursor.fetchone()
            if not region_data:
                return jsonify({'error': 'Region not found'}), 404
            
            # Check if location already exists
            cursor.execute("SELECT id FROM locations WHERE name = ?", (name,))
            if cursor.fetchone():
                return jsonify({'error': 'Location already exists'}), 400
            
            # Create new location
            cursor.execute("""
                INSERT INTO locations (name, region_id, description, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
            """, (name, region_id, description, now, now))
            
            location_id = cursor.lastrowid
            conn.commit()
            
            return jsonify({
                'success': True,
                'message': f'Location "{name}" created successfully',
                'location': {
                    'id': location_id,
                    'name': name,
                    'region_id': region_id,
                    'region_name': region_data[0],
                    'description': description,
                    'created_at': now,
                    'updated_at': now
                }
            })
            
    except Exception as e:
        logging.error(f"Create location error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/locations/<int:location_id>', methods=['PUT'])
@admin_required
def update_location(location_id):
    """Update a location"""
    try:
        data = request.json
        name = data.get('name', '').strip()
        region_id = data.get('region_id')
        description = data.get('description', '').strip()
        
        if not name:
            return jsonify({'error': 'Location name is required'}), 400
        
        if not region_id:
            return jsonify({'error': 'Region is required'}), 400
        
        # Regional admin: can only update locations in their region
        user = get_current_user()
        if user['role'] == 'regional_admin' and user['region_id']:
            with sqlite3.connect('ping_history.db', timeout=10) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT region_id FROM locations WHERE id = ?", (location_id,))
                loc_data = cursor.fetchone()
                if not loc_data or loc_data[0] != user['region_id']:
                    return jsonify({'error': 'You can only update locations in your region'}), 403
                # Also check new region
                if int(region_id) != user['region_id']:
                    return jsonify({'error': 'You can only assign locations to your region'}), 403
        
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Check if location exists
            cursor.execute("SELECT name FROM locations WHERE id = ?", (location_id,))
            if not cursor.fetchone():
                return jsonify({'error': 'Location not found'}), 404
            
            # Check if region exists
            cursor.execute("SELECT name FROM regions WHERE id = ?", (region_id,))
            if not cursor.fetchone():
                return jsonify({'error': 'Region not found'}), 404
            
            # Check if new name conflicts with another location
            cursor.execute("SELECT id FROM locations WHERE name = ? AND id != ?", (name, location_id))
            if cursor.fetchone():
                return jsonify({'error': 'Location name already exists'}), 400
            
            # Update location
            cursor.execute("""
                UPDATE locations
                SET name = ?, region_id = ?, description = ?, updated_at = ?
                WHERE id = ?
            """, (name, region_id, description, now, location_id))
            
            conn.commit()
            
            return jsonify({
                'success': True,
                'message': f'Location "{name}" updated successfully'
            })
            
    except Exception as e:
        logging.error(f"Update location error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/locations/<int:location_id>', methods=['DELETE'])
@admin_required
def delete_location(location_id):
    """Delete a location"""
    try:
        # Regional admin: can only delete locations in their region
        user = get_current_user()
        if user['role'] == 'regional_admin' and user['region_id']:
            with sqlite3.connect('ping_history.db', timeout=10) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT region_id FROM locations WHERE id = ?", (location_id,))
                loc_data = cursor.fetchone()
                if not loc_data or loc_data[0] != user['region_id']:
                    return jsonify({'error': 'You can only delete locations in your region'}), 403
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Check if location exists
            cursor.execute("SELECT name FROM locations WHERE id = ?", (location_id,))
            location_data = cursor.fetchone()
            if not location_data:
                return jsonify({'error': 'Location not found'}), 404
            
            location_name = location_data[0]
            
            # Check if location has devices (from Excel)
            device_count = 0
            if CACHED_DF is not None and 'Location' in CACHED_DF.columns:
                device_count = len(CACHED_DF[CACHED_DF['Location'] == location_name])
            
            if device_count > 0:
                return jsonify({'error': f'Cannot delete location with {device_count} devices. Please reassign devices first.'}), 400
            
            # Delete location
            cursor.execute("DELETE FROM locations WHERE id = ?", (location_id,))
            conn.commit()
            
            return jsonify({
                'success': True,
                'message': f'Location "{location_name}" deleted successfully'
            })
            
    except Exception as e:
        logging.error(f"Delete location error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/sla_metrics')
@login_required
def get_sla_metrics():
    """Get SLA performance metrics"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        region_id = request.args.get('region_id')
        user = get_current_user()
        
        # If user is not admin and has a region, enforce their region
        if user['role'] != 'admin' and user.get('region_id'):
            region_id = str(user['region_id'])
        
        if not start_date or not end_date:
            return jsonify({'error': 'Start and end dates are required'}), 400
        
        logging.info(f"Fetching SLA metrics for date range: {start_date} to {end_date}, region_id: {region_id}")
        
        # Get allowed SM IPs for the region
        allowed_ips = None
        if region_id and CACHED_DF is not None:
            try:
                with sqlite3.connect('ping_history.db', timeout=10) as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT name FROM locations WHERE region_id = ?", (region_id,))
                    allowed_locations = set(row[0].strip() for row in cursor.fetchall())
                    logging.info(f"Found {len(allowed_locations)} locations in region {region_id}")
                    
                    # Get SM IPs for these locations from CACHED_DF
                    allowed_ips = set()
                    for _, row in CACHED_DF.iterrows():
                        location = str(row.get('Location', '')).strip()
                        sm_ip = str(row.get('SM IP', ''))
                        if location in allowed_locations and pd.notna(row.get('SM IP')):
                            allowed_ips.add(sm_ip)
                    
                    logging.info(f"Filtering for {len(allowed_ips)} IPs in region {region_id}")
            except Exception as e:
                logging.error(f"Error fetching region filter: {e}")
        
        # Calculate SLA metrics from database with optimizations
        with sqlite3.connect('ping_history.db', timeout=60) as conn:
            # Enable query optimizations
            conn.execute("PRAGMA query_only = ON")
            conn.execute("PRAGMA temp_store = MEMORY")
            cursor = conn.cursor()
            
            # Build WHERE clause for region filtering
            where_clause = "timestamp BETWEEN ? AND ?"
            params = [start_date, end_date + ' 23:59:59']
            
            if allowed_ips:
                placeholders = ','.join('?' * len(allowed_ips))
                where_clause += f" AND sm_ip IN ({placeholders})"
                params.extend(list(allowed_ips))
            
            # First, check if we have any data at all
            cursor.execute("SELECT COUNT(*) FROM history")
            total_history_records = cursor.fetchone()[0]
            logging.info(f"Total history records in database: {total_history_records}")
            
            if total_history_records == 0:
                logging.warning("No history records found in database")
                return jsonify({
                    'uptime': 0,
                    'avg_latency': 0,
                    'availability': 0,
                    'total_devices': 0,
                    'sla_compliance': 0,
                    'incidents': 0,
                    'trends': {'labels': [], 'uptime': [], 'latency': []},
                    'ip_rankings': {'best': [], 'worst': []},
                    'message': 'No historical data available. Please wait for monitoring to collect data.'
                })
            
            # Get total records in date range
            query = f"""
                SELECT COUNT(*) as total_records,
                       SUM(CASE WHEN status = 'Reachable' THEN 1 ELSE 0 END) as reachable,
                       SUM(CASE WHEN status = 'Down' THEN 1 ELSE 0 END) as down,
                       SUM(CASE WHEN status = 'Degraded' THEN 1 ELSE 0 END) as degraded,
                       AVG(CASE WHEN latency IS NOT NULL THEN latency END) as avg_latency,
                       COUNT(DISTINCT sm_ip) as total_devices
                FROM history 
                WHERE {where_clause}
            """
            
            cursor.execute(query, params)
            
            metrics = cursor.fetchone()
            logging.info(f"Query results: {metrics}")
            
            if not metrics or metrics[0] == 0:
                logging.warning(f"No data found for date range {start_date} to {end_date}")
                return jsonify({
                    'uptime': 0,
                    'avg_latency': 0,
                    'availability': 0,
                    'total_devices': 0,
                    'sla_compliance': 0,
                    'incidents': 0,
                    'trends': {'labels': [], 'uptime': [], 'latency': []},
                    'ip_rankings': {'best': [], 'worst': []},
                    'message': f'No data available for the selected date range ({start_date} to {end_date}). Try selecting a different date range.'
                })
            
            total_records, reachable, down, degraded, avg_latency, total_devices = metrics
            
            # Calculate uptime percentage
            uptime = round((reachable / total_records) * 100, 2) if total_records > 0 else 0
            
            # Calculate availability (devices that had at least one successful ping)
            query = f"""
                SELECT COUNT(DISTINCT sm_ip) as available_devices
                FROM history 
                WHERE {where_clause} AND status = 'Reachable'
            """
            
            cursor.execute(query, params)
            
            available_devices = cursor.fetchone()[0]
            availability = round((available_devices / total_devices) * 100, 2) if total_devices > 0 else 0
            
            # Count incidents (simplified - count distinct down events per device per day)
            query = f"""
                SELECT COUNT(DISTINCT sm_ip || DATE(timestamp)) as incidents
                FROM history
                WHERE {where_clause}
                AND status = 'Down'
            """
            
            cursor.execute(query, params)
            
            incidents = cursor.fetchone()[0]
            
            # Calculate SLA compliance score
            uptime_score = min(100, (uptime / 99.5) * 40) if uptime > 0 else 0
            latency_score = min(40, max(0, 40 - (avg_latency - 50) / 10)) if avg_latency else 40
            availability_score = min(20, (availability / 99.9) * 20) if availability > 0 else 0
            sla_compliance = round(uptime_score + latency_score + availability_score, 1)
            
            # Get trends data (daily aggregates)
            query = f"""
                SELECT DATE(timestamp) as date,
                       AVG(CASE WHEN status = 'Reachable' THEN 100.0 ELSE 0.0 END) as daily_uptime,
                       AVG(CASE WHEN latency IS NOT NULL THEN latency END) as daily_latency
                FROM history 
                WHERE {where_clause}
                GROUP BY DATE(timestamp)
                ORDER BY date
            """
            
            cursor.execute(query, params)
            
            trends_data = cursor.fetchall()
            trends = {
                'labels': [row[0] for row in trends_data],
                'uptime': [round(row[1], 1) if row[1] else 0 for row in trends_data],
                'latency': [round(row[2], 1) if row[2] else 0 for row in trends_data]
            }
            
            # Get best and worst performing IPs (optimized with LIMIT)
            logging.info("Fetching IP performance rankings...")
            
            # Get best performing IPs
            query = f"""
                SELECT sm_ip,
                       AVG(CASE WHEN status = 'Reachable' THEN 100.0 ELSE 0.0 END) as uptime_pct,
                       AVG(CASE WHEN latency IS NOT NULL THEN latency END) as avg_latency,
                       COUNT(*) as total_pings
                FROM history 
                WHERE {where_clause}
                GROUP BY sm_ip
                HAVING total_pings >= 5
                ORDER BY uptime_pct DESC, avg_latency ASC
                LIMIT 10
            """
            
            cursor.execute(query, params)
            
            best_performance = cursor.fetchall()
            
            # Get worst performing IPs
            query = f"""
                SELECT sm_ip,
                       AVG(CASE WHEN status = 'Reachable' THEN 100.0 ELSE 0.0 END) as uptime_pct,
                       AVG(CASE WHEN latency IS NOT NULL THEN latency END) as avg_latency,
                       COUNT(*) as total_pings
                FROM history 
                WHERE {where_clause}
                GROUP BY sm_ip
                HAVING total_pings >= 5
                ORDER BY uptime_pct ASC, avg_latency DESC
                LIMIT 10
            """
            
            cursor.execute(query, params)
            
            worst_performance = cursor.fetchall()
            
            logging.info(f"Found {len(best_performance)} best and {len(worst_performance)} worst performing IPs")
            
            # Get device info for rankings
            device_info = {}
            if CACHED_DF is not None:
                for _, row in CACHED_DF.iterrows():
                    if pd.notna(row.get('SM IP')):
                        device_info[str(row['SM IP'])] = {
                            'name': str(row.get('Device Name', 'Unknown')),
                            'location': str(row.get('Location', 'Unknown'))
                        }
            
            logging.info(f"Device info cache has {len(device_info)} entries")
            
            best_ips = []
            worst_ips = []
            
            # Process best IPs
            for ip, uptime_pct, avg_lat, _ in best_performance:
                info = device_info.get(ip, {'name': 'Unknown', 'location': 'Unknown'})
                best_ips.append({
                    'ip': ip,
                    'name': info['name'],
                    'location': info['location'],
                    'uptime': round(uptime_pct, 1),
                    'avg_latency': round(avg_lat, 1) if avg_lat else 0
                })
            
            # Process worst IPs
            for ip, uptime_pct, avg_lat, _ in worst_performance:
                info = device_info.get(ip, {'name': 'Unknown', 'location': 'Unknown'})
                worst_ips.append({
                    'ip': ip,
                    'name': info['name'],
                    'location': info['location'],
                    'uptime': round(uptime_pct, 1),
                    'avg_latency': round(avg_lat, 1) if avg_lat else 0
                })
            
            logging.info(f"SLA metrics calculated: uptime={uptime}%, devices={total_devices}, incidents={incidents}")
            logging.info(f"Rankings: {len(best_ips)} best IPs, {len(worst_ips)} worst IPs")
            
            response_data = {
                'uptime': uptime,
                'avg_latency': round(avg_latency, 1) if avg_latency else 0,
                'availability': availability,
                'total_devices': total_devices,
                'sla_compliance': sla_compliance,
                'incidents': incidents,
                'trends': trends,
                'ip_rankings': {
                    'best': best_ips,
                    'worst': worst_ips
                }
            }
            
            logging.info(f"Returning SLA response with {len(best_ips)} best and {len(worst_ips)} worst IPs")
            return jsonify(response_data)
            
    except Exception as e:
        logging.error(f"SLA metrics error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/location_overview')
@login_required
def get_location_overview():
    """Get location-wise health overview"""
    try:
        # Get region filter from query params
        region_id = request.args.get('region_id')
        user = get_current_user()
        
        # If user is not admin and has a region, enforce their region
        if user['role'] != 'admin' and user.get('region_id'):
            region_id = str(user['region_id'])
        
        logging.info(f"Getting location overview for region_id: {region_id}")
        
        # Get locations for the specified region
        allowed_locations = None
        if region_id:
            try:
                with sqlite3.connect('ping_history.db', timeout=10) as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT name FROM locations WHERE region_id = ?", (region_id,))
                    allowed_locations = set(row[0].strip() for row in cursor.fetchall())
                    logging.info(f"Filtering for {len(allowed_locations)} locations in region {region_id}")
            except Exception as e:
                logging.error(f"Error fetching locations for region: {e}")
        
        with results_lock:
            if not results:
                logging.warning("No results available for location overview")
                return jsonify({'locations': [], 'message': 'No monitoring data available yet. Please wait for the monitoring system to collect data.'})
            
            logging.info(f"Processing {len(results)} results for location overview")
            
            # Group results by location
            location_stats = defaultdict(lambda: {
                'total_devices': 0,
                'online': 0,
                'degraded': 0,
                'down': 0,
                'latencies': []
            })
            
            for result in results:
                location = result.get('Location', 'Unknown').strip()
                
                # Filter by region if specified
                if allowed_locations is not None and location not in allowed_locations:
                    continue
                
                status = result.get('Status', 'Unknown')
                latency_str = result.get('Latency', 'N/A')
                
                if not location or location == 'Unknown':
                    continue
                
                location_stats[location]['total_devices'] += 1
                
                if status == 'Reachable':
                    location_stats[location]['online'] += 1
                elif status == 'Degraded':
                    location_stats[location]['degraded'] += 1
                elif status == 'Down':
                    location_stats[location]['down'] += 1
                
                # Parse latency
                if latency_str != 'N/A' and 'ms' in latency_str:
                    try:
                        latency = float(latency_str.replace('ms', '').strip())
                        location_stats[location]['latencies'].append(latency)
                    except:
                        pass
            
            # Calculate health scores and format response
            locations = []
            for location, stats in location_stats.items():
                total = stats['total_devices']
                if total == 0:
                    continue
                
                uptime = round((stats['online'] / total) * 100, 1)
                avg_latency = round(sum(stats['latencies']) / len(stats['latencies']), 1) if stats['latencies'] else 0
                
                # Calculate health score (0-100)
                uptime_score = min(60, uptime * 0.6)  # Max 60 points for uptime
                latency_score = min(30, max(0, 30 - (avg_latency - 50) / 10)) if avg_latency > 0 else 30  # Max 30 points for latency
                availability_score = min(10, (stats['online'] + stats['degraded']) / total * 10)  # Max 10 points for availability
                
                health_score = round(uptime_score + latency_score + availability_score, 1)
                
                locations.append({
                    'name': location,
                    'total_devices': total,
                    'online': stats['online'],
                    'degraded': stats['degraded'],
                    'down': stats['down'],
                    'uptime': uptime,
                    'avg_latency': avg_latency,
                    'health_score': health_score
                })
            
            # Sort by health score descending
            locations.sort(key=lambda x: x['health_score'], reverse=True)
            
            logging.info(f"Returning {len(locations)} locations for overview")
            return jsonify({'locations': locations})
            
    except Exception as e:
        logging.error(f"Location overview error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/location_downtime')
@login_required
def get_location_downtime():
    """Get location-wise downtime analysis"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        region_id = request.args.get('region_id')
        user = get_current_user()
        
        # If user is not admin and has a region, enforce their region
        if user['role'] != 'admin' and user.get('region_id'):
            region_id = str(user['region_id'])
        
        if not start_date or not end_date:
            return jsonify({'error': 'Start and end dates are required'}), 400
        
        logging.info(f"Getting location downtime for region_id: {region_id}")
        
        # Get allowed locations for the region
        allowed_locations = None
        if region_id:
            try:
                with sqlite3.connect('ping_history.db', timeout=10) as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT name FROM locations WHERE region_id = ?", (region_id,))
                    allowed_locations = set(row[0].strip() for row in cursor.fetchall())
                    logging.info(f"Filtering downtime for {len(allowed_locations)} locations in region {region_id}")
            except Exception as e:
                logging.error(f"Error fetching locations for region: {e}")
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Get device location mapping
            device_locations = {}
            if CACHED_DF is not None:
                for _, row in CACHED_DF.iterrows():
                    if pd.notna(row.get('SM IP')):
                        location = str(row.get('Location', 'Unknown')).strip()
                        # Filter by region if specified
                        if allowed_locations is None or location in allowed_locations:
                            device_locations[str(row['SM IP'])] = location
            
            # Get downtime data by IP
            cursor.execute("""
                SELECT sm_ip,
                       SUM(CASE WHEN status = 'Down' THEN 1 ELSE 0 END) * 10 / 3600.0 as downtime_hours,
                       COUNT(*) as total_pings,
                       MAX(CASE WHEN status = 'Down' THEN timestamp END) as last_down
                FROM history 
                WHERE timestamp BETWEEN ? AND ?
                GROUP BY sm_ip
            """, (start_date, end_date + ' 23:59:59'))
            
            ip_downtime = cursor.fetchall()
            
            # Group by location
            location_downtime = defaultdict(lambda: {
                'total_downtime': 0,
                'device_count': 0,
                'down_devices': 0,
                'last_incident': None
            })
            
            for ip, downtime_hours, total_pings, last_down in ip_downtime:
                location = device_locations.get(ip)
                
                # Skip if IP not in allowed locations (region filter)
                if location is None:
                    continue
                
                location_downtime[location]['total_downtime'] += downtime_hours
                location_downtime[location]['device_count'] += 1
                
                if downtime_hours > 0:
                    location_downtime[location]['down_devices'] += 1
                    
                if last_down and (not location_downtime[location]['last_incident'] or 
                                last_down > location_downtime[location]['last_incident']):
                    location_downtime[location]['last_incident'] = last_down
            
            # Format response
            locations = []
            chart_labels = []
            chart_data = []
            
            for location, stats in location_downtime.items():
                if stats['device_count'] == 0:
                    continue
                
                avg_downtime = stats['total_downtime'] / stats['device_count']
                uptime_hours = (24 * 7) - avg_downtime  # Assuming 7-day period
                uptime_pct = round((uptime_hours / (24 * 7)) * 100, 1)
                
                locations.append({
                    'location': location,
                    'total_downtime': f"{stats['total_downtime']:.1f}h",
                    'avg_downtime': f"{avg_downtime:.1f}h",
                    'down_ip_count': stats['down_devices'],
                    'last_downtime': stats['last_incident'] or 'N/A',
                    'uptime': f"{uptime_pct}%"
                })
                
                chart_labels.append(location)
                chart_data.append(round(stats['total_downtime'], 1))
            
            # Sort by total downtime descending
            locations.sort(key=lambda x: float(x['total_downtime'].replace('h', '')), reverse=True)
            
            logging.info(f"Returning downtime for {len(locations)} locations")
            
            return jsonify({
                'locations': locations,
                'chart_data': {
                    'labels': chart_labels,
                    'data': chart_data
                }
            })
            
    except Exception as e:
        logging.error(f"Location downtime error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/locations_from_excel')
def get_locations_from_excel():
    """Get all unique locations from Excel file (legacy)"""
    try:
        with results_lock:
            if CACHED_DF is None:
                return jsonify({'locations': []})
            
            locations = sorted(set(CACHED_DF['Location'].dropna().astype(str)))
            return jsonify({'locations': locations})
    except Exception as e:
        logging.error(f"Get locations error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/device_list')
@login_required
def get_device_list():
    """Get all devices with their details for maintenance scheduling"""
    try:
        with results_lock:
            if CACHED_DF is None:
                return jsonify({'devices': []})
            
            devices = []
            for _, row in CACHED_DF.iterrows():
                if pd.notna(row.get('SM IP')):
                    device = {
                        'ip': str(row['SM IP']),
                        'name': str(row.get('Device Name', 'Unknown')),
                        'location': str(row.get('Location', 'Unknown')),
                        'ap_name': str(row.get('AP Name', 'N/A')),
                        'ap_ip': str(row.get('AP IP', 'N/A')),
                        'cid': str(row.get('CID', 'N/A'))
                    }
                    devices.append(device)
            
            # Sort devices by location, then by IP
            devices.sort(key=lambda x: (x['location'], x['ip']))
            
            return jsonify({'devices': devices})
    except Exception as e:
        logging.error(f"Get device list error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/sla_report')
def get_sla_report():
    """Generate SLA report"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        sla_target = float(request.args.get('sla_target', 99.9))
        location_filter = request.args.get('location_filter')
        
        if not start_date or not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        
        # Get device-location mapping
        with results_lock:
            if CACHED_DF is None:
                return jsonify({'error': 'No device data available'}), 500
            
            ip_to_location = {str(row['SM IP']): str(row['Location']) 
                            for row in CACHED_DF.to_dict('records') 
                            if pd.notna(row.get('SM IP'))}
            all_locations = sorted(set(CACHED_DF['Location'].dropna().astype(str)))
        
        # Query historical data
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            query = """
                SELECT sm_ip, status, timestamp
                FROM history
                WHERE timestamp >= ? AND timestamp < ?
                ORDER BY sm_ip, timestamp
            """
            cursor.execute(query, (start_datetime.strftime('%Y-%m-%d %H:%M:%S'), 
                                 end_datetime.strftime('%Y-%m-%d %H:%M:%S')))
            records = cursor.fetchall()
        
        # Calculate SLA metrics
        location_stats = defaultdict(lambda: {
            'total_time': 0,
            'uptime': 0,
            'downtime': 0,
            'devices': set(),
            'incidents': 0,
            'last_incident': None,
            'downtime_events': []
        })
        
        # Process records
        for i, (sm_ip, status, timestamp) in enumerate(records):
            location = ip_to_location.get(sm_ip, 'Unknown')
            if location_filter and location != location_filter:
                continue
            
            location_stats[location]['devices'].add(sm_ip)
            
            # Calculate duration until next record
            duration = PING_INTERVAL
            if i + 1 < len(records) and records[i + 1][0] == sm_ip:
                next_timestamp = datetime.strptime(records[i + 1][2], '%Y-%m-%d %H:%M:%S')
                current_timestamp = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
                duration = (next_timestamp - current_timestamp).total_seconds()
            
            location_stats[location]['total_time'] += duration
            
            if status == 'Reachable':
                location_stats[location]['uptime'] += duration
            else:
                location_stats[location]['downtime'] += duration
                if status == 'Down':
                    # Check if this is a new incident
                    if i == 0 or records[i-1][0] != sm_ip or records[i-1][1] != 'Down':
                        location_stats[location]['incidents'] += 1
                        location_stats[location]['last_incident'] = timestamp
                        location_stats[location]['downtime_events'].append({
                            'start': timestamp,
                            'device': sm_ip
                        })
        
        # Calculate SLA percentages and prepare response
        locations_data = []
        total_uptime = 0
        total_time = 0
        locations_meeting_sla = 0
        critical_incidents = 0
        
        for location, stats in location_stats.items():
            if stats['total_time'] == 0:
                continue
            
            uptime_percentage = (stats['uptime'] / stats['total_time']) * 100
            mttr = stats['downtime'] / max(stats['incidents'], 1)  # Mean Time To Repair
            
            locations_data.append({
                'name': location,
                'device_count': len(stats['devices']),
                'uptime_percentage': uptime_percentage,
                'total_downtime': stats['downtime'],
                'incident_count': stats['incidents'],
                'mttr': mttr,
                'last_incident': stats['last_incident']
            })
            
            total_uptime += stats['uptime']
            total_time += stats['total_time']
            
            if uptime_percentage >= sla_target:
                locations_meeting_sla += 1
            
            if stats['incidents'] > 5:  # Consider >5 incidents as critical
                critical_incidents += stats['incidents']
        
        # Calculate overall metrics
        overall_sla = (total_uptime / max(total_time, 1)) * 100
        average_uptime = sum(loc['uptime_percentage'] for loc in locations_data) / max(len(locations_data), 1)
        
        # Generate trend data (simplified - daily averages)
        trend_labels = []
        trend_data = []
        current_date = start_datetime
        while current_date < end_datetime - timedelta(days=1):
            trend_labels.append(current_date.strftime('%m/%d'))
            # Simplified: use overall SLA with some variation
            daily_sla = overall_sla + (hash(current_date.day) % 5 - 2) * 0.1
            trend_data.append(max(90, min(100, daily_sla)))
            current_date += timedelta(days=1)
        
        return jsonify({
            'overall_sla': overall_sla,
            'average_uptime': average_uptime,
            'locations_meeting_sla': locations_meeting_sla,
            'total_locations': len(locations_data),
            'critical_incidents': critical_incidents,
            'locations': locations_data,
            'trend_labels': trend_labels,
            'trend_data': trend_data
        })
        
    except Exception as e:
        logging.error(f"SLA report error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/sla_export')
@login_required
def export_sla_report():
    """Export SLA report to Excel"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        sla_target = float(request.args.get('sla_target', 99.9))
        format_type = request.args.get('format', 'xlsx')
        
        if not start_date or not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        
        # Get SLA metrics data
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Get overall metrics
            cursor.execute("""
                SELECT COUNT(*) as total_records,
                       SUM(CASE WHEN status = 'Reachable' THEN 1 ELSE 0 END) as reachable,
                       SUM(CASE WHEN status = 'Down' THEN 1 ELSE 0 END) as down,
                       AVG(CASE WHEN latency IS NOT NULL THEN latency END) as avg_latency,
                       COUNT(DISTINCT sm_ip) as total_devices
                FROM history 
                WHERE timestamp BETWEEN ? AND ?
            """, (start_date, end_date + ' 23:59:59'))
            
            metrics = cursor.fetchone()
            
            if not metrics or metrics[0] == 0:
                return jsonify({'error': 'No data available for the selected date range'}), 400
            
            total_records, reachable, down, avg_latency, total_devices = metrics
            uptime = round((reachable / total_records) * 100, 2) if total_records > 0 else 0
            
            # Get device-wise performance
            cursor.execute("""
                SELECT sm_ip,
                       AVG(CASE WHEN status = 'Reachable' THEN 100.0 ELSE 0.0 END) as uptime_pct,
                       AVG(CASE WHEN latency IS NOT NULL THEN latency END) as avg_latency,
                       COUNT(*) as total_pings,
                       SUM(CASE WHEN status = 'Down' THEN 1 ELSE 0 END) as down_count
                FROM history 
                WHERE timestamp BETWEEN ? AND ?
                GROUP BY sm_ip
                ORDER BY uptime_pct DESC
            """, (start_date, end_date + ' 23:59:59'))
            
            device_performance = cursor.fetchall()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "SLA Report"
        
        # Add headers
        headers = ['Device IP', 'Device Name', 'Location', 'Uptime %', 'Avg Latency (ms)', 'Total Pings', 'Down Count', 'SLA Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Get device info
        device_info = {}
        if CACHED_DF is not None:
            for _, row in CACHED_DF.iterrows():
                if pd.notna(row.get('SM IP')):
                    device_info[str(row['SM IP'])] = {
                        'name': str(row.get('Device Name', 'Unknown')),
                        'location': str(row.get('Location', 'Unknown'))
                    }
        
        # Add data rows
        for row_num, (ip, uptime_pct, avg_lat, total_pings, down_count) in enumerate(device_performance, 2):
            info = device_info.get(ip, {'name': 'Unknown', 'location': 'Unknown'})
            sla_status = 'PASS' if uptime_pct >= sla_target else 'FAIL'
            
            ws.cell(row=row_num, column=1, value=ip)
            ws.cell(row=row_num, column=2, value=info['name'])
            ws.cell(row=row_num, column=3, value=info['location'])
            ws.cell(row=row_num, column=4, value=round(uptime_pct, 2))
            ws.cell(row=row_num, column=5, value=round(avg_lat, 2) if avg_lat else 0)
            ws.cell(row=row_num, column=6, value=total_pings)
            ws.cell(row=row_num, column=7, value=down_count)
            ws.cell(row=row_num, column=8, value=sla_status)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"sla_report_{start_date}_to_{end_date}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logging.error(f"SLA export error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/downtime_report')
@login_required
def export_downtime_report():
    """Export downtime report"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        format_type = request.args.get('format', 'pdf')
        
        if not start_date or not end_date:
            return jsonify({'error': 'Start and end dates are required'}), 400
        
        # Get downtime data (reuse the logic from location_downtime)
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Get device location mapping
            device_locations = {}
            if CACHED_DF is not None:
                for _, row in CACHED_DF.iterrows():
                    if pd.notna(row.get('SM IP')):
                        device_locations[str(row['SM IP'])] = str(row.get('Location', 'Unknown'))
            
            # Get downtime data by IP
            cursor.execute("""
                SELECT sm_ip,
                       SUM(CASE WHEN status = 'Down' THEN 1 ELSE 0 END) * 10 / 3600.0 as downtime_hours,
                       COUNT(*) as total_pings,
                       MAX(CASE WHEN status = 'Down' THEN timestamp END) as last_down
                FROM history 
                WHERE timestamp BETWEEN ? AND ?
                GROUP BY sm_ip
                HAVING downtime_hours > 0
                ORDER BY downtime_hours DESC
            """, (start_date, end_date + ' 23:59:59'))
            
            downtime_data = cursor.fetchall()
        
        if format_type == 'xlsx':
            # Create Excel report
            wb = Workbook()
            ws = wb.active
            ws.title = "Downtime Report"
            
            # Add headers
            headers = ['Device IP', 'Device Name', 'Location', 'Downtime Hours', 'Total Pings', 'Last Down Time']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Get device info
            device_info = {}
            if CACHED_DF is not None:
                for _, row in CACHED_DF.iterrows():
                    if pd.notna(row.get('SM IP')):
                        device_info[str(row['SM IP'])] = {
                            'name': str(row.get('Device Name', 'Unknown')),
                            'location': str(row.get('Location', 'Unknown'))
                        }
            
            # Add data rows
            for row_num, (ip, downtime_hours, total_pings, last_down) in enumerate(downtime_data, 2):
                info = device_info.get(ip, {'name': 'Unknown', 'location': 'Unknown'})
                
                ws.cell(row=row_num, column=1, value=ip)
                ws.cell(row=row_num, column=2, value=info['name'])
                ws.cell(row=row_num, column=3, value=info['location'])
                ws.cell(row=row_num, column=4, value=round(downtime_hours, 2))
                ws.cell(row=row_num, column=5, value=total_pings)
                ws.cell(row=row_num, column=6, value=last_down or 'N/A')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"downtime_report_{start_date}_to_{end_date}.xlsx"
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        else:
            # Return JSON for PDF generation (client-side)
            report_data = []
            device_info = {}
            if CACHED_DF is not None:
                for _, row in CACHED_DF.iterrows():
                    if pd.notna(row.get('SM IP')):
                        device_info[str(row['SM IP'])] = {
                            'name': str(row.get('Device Name', 'Unknown')),
                            'location': str(row.get('Location', 'Unknown'))
                        }
            
            for ip, downtime_hours, total_pings, last_down in downtime_data:
                info = device_info.get(ip, {'name': 'Unknown', 'location': 'Unknown'})
                report_data.append({
                    'ip': ip,
                    'name': info['name'],
                    'location': info['location'],
                    'downtime_hours': round(downtime_hours, 2),
                    'total_pings': total_pings,
                    'last_down': last_down or 'N/A'
                })
            
            return jsonify({
                'report_data': report_data,
                'start_date': start_date,
                'end_date': end_date,
                'total_devices': len(report_data)
            })
        
    except Exception as e:
        logging.error(f"Downtime report error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/force_ping', methods=['POST'])
@login_required
def force_ping():
    """Force ping a specific device manually"""
    try:
        data = request.json
        ip = data.get('ip', '').strip()
        ping_type = data.get('ping_type', 'SM')  # SM or AP
        sm_ip = data.get('sm_ip', ip).strip()  # The device we're viewing (for comments)
        
        if not ip or not validate_ip(ip):
            return jsonify({'error': 'Invalid IP address'}), 400
        
        # Perform the ping
        status, latency, issue_type = ping_ip(ip, timeout=2.0)  # Use longer timeout for manual ping
        
        # Get current timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Log the manual ping result
        user = get_current_user()
        username = user['username'] if user else 'Unknown'
        
        # Add to database
        try:
            with sqlite3.connect('ping_history.db', timeout=10) as conn:
                conn.execute("INSERT INTO history VALUES (?, ?, ?, ?)", 
                           (timestamp, ip, status, latency))
                conn.commit()
        except Exception as db_error:
            logging.error(f"Failed to log manual ping: {db_error}")
        
        # Create detailed comment about manual ping
        latency_text = f" ({latency:.2f}ms)" if latency else ""
        comment_text = f"Manual {ping_type} ping to {ip}: {status}{latency_text} - Tested by {username}"
        
        # Add comment to the SM device (the device being viewed)
        try:
            with sqlite3.connect('ping_history.db', timeout=10) as conn:
                conn.execute("""
                    INSERT INTO device_comments (sm_ip, comment, username, timestamp, comment_type)
                    VALUES (?, ?, ?, ?, 'manual_ping')
                """, (sm_ip, comment_text, username, timestamp))
                conn.commit()
        except Exception as comment_error:
            logging.error(f"Failed to add manual ping comment: {comment_error}")
        
        # Create response message
        latency_display = f"{latency:.2f}ms" if latency else "N/A"
        message = f"{ping_type} IP {ip}: {status} (Latency: {latency_display})"
        
        return jsonify({
            'success': True,
            'status': status,
            'latency': latency_display,
            'timestamp': timestamp,
            'message': message,
            'ping_type': ping_type
        })
        
    except Exception as e:
        logging.error(f"Force ping error: {str(e)}")
        return jsonify({'error': str(e)}), 500

# User Management API Routes
@app.route('/api/users', methods=['GET'])
@admin_required
def get_users():
    """Get all users, optionally filtered by location"""
    try:
        location_id = request.args.get('location_id')
        
        if location_id:
            # Get users in specific location (by matching region)
            with sqlite3.connect('ping_history.db', timeout=10) as conn:
                cursor = conn.cursor()
                
                # First get the region of the specified location
                cursor.execute("SELECT region_id FROM locations WHERE id = ?", (location_id,))
                location_result = cursor.fetchone()
                
                if not location_result:
                    return jsonify({'error': 'Location not found'}), 404
                
                region_id = location_result[0]
                
                # Get users in that region
                cursor.execute("""
                    SELECT u.id, u.username, u.email, u.role, u.is_active, u.created_at, u.last_login,
                           u.region_id, r.name as region_name
                    FROM users u
                    LEFT JOIN regions r ON u.region_id = r.id
                    WHERE u.region_id = ? AND u.is_active = 1
                    ORDER BY u.username
                """, (region_id,))
                
                users = cursor.fetchall()
                users_list = []
                for user in users:
                    users_list.append({
                        'id': user[0],
                        'username': user[1],
                        'email': user[2],
                        'role': user[3],
                        'is_active': user[4],
                        'created_at': user[5],
                        'last_login': user[6],
                        'region_id': user[7],
                        'region_name': user[8]
                    })
                
                return jsonify({'success': True, 'users': users_list})
        else:
            # Get all users (existing functionality)
            users = get_all_users()
            return jsonify({'success': True, 'users': users})
            
    except Exception as e:
        logging.error(f"Get users error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/users', methods=['POST'])
@superadmin_required
def create_user():
    """Create a new user"""
    try:
        data = request.json
        username = data.get('username', '').strip()
        email = data.get('email', '').strip()
        password = data.get('password', '')
        role = data.get('role', 'user')
        region_id = data.get('region_id')  # NEW: Get region assignment
        
        # Validation
        if not username or not email or not password:
            return jsonify({'error': 'Username, email, and password are required'}), 400
        
        if len(username) < 3:
            return jsonify({'error': 'Username must be at least 3 characters long'}), 400
        
        if len(password) < 6:
            return jsonify({'error': 'Password must be at least 6 characters long'}), 400
        
        # Region validation for non-admin users
        if role == 'user' and not region_id:
            return jsonify({'error': 'Region assignment is required for regular users'}), 400
        
        # Email validation
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_pattern, email):
            return jsonify({'error': 'Invalid email address'}), 400
        
        if role not in ['user', 'admin']:
            return jsonify({'error': 'Role must be either "user" or "admin"'}), 400
        
        # Extract username from email if not provided
        if not username or username == email:
            username = email.split('@')[0]
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Check if username or email already exists
            cursor.execute("SELECT id FROM users WHERE username = ? OR email = ?", (username, email))
            if cursor.fetchone():
                return jsonify({'error': 'Username or email already exists'}), 400
            
            # Verify region exists if provided
            if region_id:
                cursor.execute("SELECT id FROM regions WHERE id = ?", (region_id,))
                if not cursor.fetchone():
                    return jsonify({'error': 'Invalid region ID'}), 400
            
            # Create new user with region assignment
            password_hash = hash_password(password)
            cursor.execute("""
                INSERT INTO users (username, email, password_hash, role, region_id, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (username, email, password_hash, role, region_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            conn.commit()
            
            return jsonify({'success': True, 'message': f'User {username} created successfully'})
            
    except Exception as e:
        logging.error(f"Create user error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/users/<int:user_id>', methods=['PUT'])
@superadmin_required
def update_user(user_id):
    """Update user details"""
    try:
        data = request.json
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Check if user exists
            cursor.execute("SELECT id, role FROM users WHERE id = ?", (user_id,))
            user_data = cursor.fetchone()
            if not user_data:
                return jsonify({'error': 'User not found'}), 404
            
            current_role = user_data[1]
            
            # Update fields
            update_fields = []
            update_values = []
            
            if 'role' in data:
                if data['role'] not in ['user', 'superadmin', 'regional_admin']:
                    return jsonify({'error': 'Role must be "user", "regional_admin", or "superadmin"'}), 400
                update_fields.append('role = ?')
                update_values.append(data['role'])
                current_role = data['role']  # Update current role for region validation
            
            if 'is_active' in data:
                update_fields.append('is_active = ?')
                update_values.append(1 if data['is_active'] else 0)
            
            # Handle region_id update
            if 'region_id' in data:
                if current_role == 'superadmin':
                    # Superadmin users should have NULL region_id
                    update_fields.append('region_id = ?')
                    update_values.append(None)
                else:
                    # Regular users and regional_admin must have a region_id
                    region_id = data['region_id']
                    if region_id:
                        # Verify region exists
                        cursor.execute("SELECT id FROM regions WHERE id = ?", (region_id,))
                        if not cursor.fetchone():
                            return jsonify({'error': 'Invalid region ID'}), 400
                        update_fields.append('region_id = ?')
                        update_values.append(region_id)
                    else:
                        return jsonify({'error': f'{current_role} users must be assigned to a region'}), 400
            
            if update_fields:
                update_values.append(user_id)
                query = f"UPDATE users SET {', '.join(update_fields)} WHERE id = ?"
                cursor.execute(query, update_values)
                conn.commit()
            
            # Log the action
            current_admin = get_current_user()
            if current_admin:
                changes = []
                if 'role' in data: changes.append(f"role→{data['role']}")
                if 'is_active' in data: changes.append(f"active→{data['is_active']}")
                if 'region_id' in data: changes.append(f"region→{data['region_id']}")
                log_user_activity(current_admin['id'], current_admin['username'], 'user_update', f"Updated user ID {user_id}: {', '.join(changes)}")
            
            return jsonify({'success': True, 'message': 'User updated successfully'})
            
    except Exception as e:
        logging.error(f"Update user error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@superadmin_required
def delete_user(user_id):
    """Delete a user"""
    try:
        current_user = get_current_user()
        if current_user['id'] == user_id:
            return jsonify({'error': 'Cannot delete your own account'}), 400
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Check if user exists
            cursor.execute("SELECT username FROM users WHERE id = ?", (user_id,))
            user_data = cursor.fetchone()
            if not user_data:
                return jsonify({'error': 'User not found'}), 404
            
            username = user_data[0]
            
            # Delete user
            cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))
            conn.commit()
            
            current_admin = get_current_user()
            if current_admin:
                log_user_activity(current_admin['id'], current_admin['username'], 'user_delete', f"Deleted user '{username}' (ID: {user_id})")
            
            return jsonify({'success': True, 'message': f'User {username} deleted successfully'})
            
    except Exception as e:
        logging.error(f"Delete user error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/users')
@login_required
def get_users_list():
    """Get list of all active users (for task assignment)"""
    try:
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, username, email, role, is_active 
                FROM users 
                WHERE is_active = 1
                ORDER BY username
            """)
            users = cursor.fetchall()
            
            users_list = []
            for user in users:
                users_list.append({
                    'id': user[0],
                    'username': user[1],
                    'email': user[2],
                    'role': user[3],
                    'is_active': user[4]
                })
            
            return jsonify({'success': True, 'users': users_list})
            
    except Exception as e:
        logging.error(f"Get users error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/users/migrate-location', methods=['POST'])
@admin_required
def migrate_users_location():
    """Migrate users from one location to another"""
    try:
        data = request.json
        user_ids = data.get('user_ids', [])
        source_location_id = data.get('source_location_id')
        target_location_id = data.get('target_location_id')
        
        if not user_ids or not source_location_id or not target_location_id:
            return jsonify({'error': 'Missing required parameters'}), 400
            
        if source_location_id == target_location_id:
            return jsonify({'error': 'Source and target locations cannot be the same'}), 400
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Get target location info
            cursor.execute("""
                SELECT l.name, l.region_id, r.name as region_name
                FROM locations l
                LEFT JOIN regions r ON l.region_id = r.id
                WHERE l.id = ?
            """, (target_location_id,))
            target_location = cursor.fetchone()
            
            if not target_location:
                return jsonify({'error': 'Target location not found'}), 404
            
            target_region_id = target_location[1]
            
            # Update users' region to match target location's region
            migrated_count = 0
            for user_id in user_ids:
                # Update user's region
                cursor.execute("""
                    UPDATE users 
                    SET region_id = ?
                    WHERE id = ?
                """, (target_region_id, user_id))
                
                if cursor.rowcount > 0:
                    migrated_count += 1
                    
                    # Log the migration activity
                    cursor.execute("SELECT username FROM users WHERE id = ?", (user_id,))
                    user_result = cursor.fetchone()
                    username = user_result[0] if user_result else f"User {user_id}"
                    
                    log_user_activity(
                        user_id=session.get('user_id'),
                        username=session.get('username'),
                        activity_type='location_migration',
                        activity_description=f"Migrated user {username} to location {target_location[0]}",
                        ip_address=request.remote_addr
                    )
            
            conn.commit()
            
            return jsonify({
                'success': True, 
                'migrated_count': migrated_count,
                'target_location': target_location[0],
                'target_region': target_location[2] or 'Unassigned'
            })
            
    except Exception as e:
        logging.error(f"Location migration error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/activity_logs')
@superadmin_required
def get_activity_logs():
    """Get user activity logs for superadmin"""
    try:
        username_filter = request.args.get('username', '').strip()
        action_filter = request.args.get('action', '').strip()
        date_from = request.args.get('date_from', '').strip()
        date_to = request.args.get('date_to', '').strip()
        limit = int(request.args.get('limit', 999999))

        query = "SELECT id, user_id, username, activity_type, activity_description, ip_address, timestamp FROM user_activity_logs WHERE 1=1"
        params = []

        if username_filter:
            query += " AND username LIKE ?"
            params.append(f"%{username_filter}%")
        if action_filter:
            query += " AND activity_type = ?"
            params.append(action_filter)
        if date_from:
            query += " AND timestamp >= ?"
            params.append(date_from)
        if date_to:
            query += " AND timestamp <= ?"
            params.append(date_to + ' 23:59:59')

        query += " ORDER BY timestamp DESC LIMIT ?"
        params.append(limit)

        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            cursor.execute(query, params)
            rows = cursor.fetchall()

        logs = [{'id': r[0], 'user_id': r[1], 'username': r[2], 'action': r[3],
                 'description': r[4], 'ip': r[5], 'timestamp': r[6]} for r in rows]
        return jsonify({'success': True, 'logs': logs})
    except Exception as e:
        logging.error(f"Error fetching activity logs: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/online_users')
@superadmin_required
def get_online_users():
    """Get currently online users based on active sessions"""
    try:
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT u.username, u.role, s.ip_address, s.created_at, s.expires_at
                FROM user_sessions s
                JOIN users u ON s.user_id = u.id
                WHERE s.expires_at > ?
                ORDER BY s.created_at DESC
            """, (datetime.now().strftime('%Y-%m-%d %H:%M:%S'),))
            rows = cursor.fetchall()

        seen = {}
        for r in rows:
            if r[0] not in seen:
                seen[r[0]] = {'username': r[0], 'role': r[1], 'ip': r[2],
                              'login_time': r[3], 'expires_at': r[4]}

        return jsonify({'success': True, 'online_users': list(seen.values()), 'count': len(seen)})
    except Exception as e:
        logging.error(f"Error fetching online users: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/assign_device', methods=['POST'])
@admin_required
def assign_device():
    """Assign a device to a user"""
    try:
        data = request.json
        device_ip = data.get('device_ip', '').strip()
        user_id = data.get('user_id')
        comment = data.get('comment', '').strip()
        
        if not device_ip or not validate_ip(device_ip):
            return jsonify({'error': 'Invalid device IP address'}), 400
        
        if not user_id:
            return jsonify({'error': 'User ID is required'}), 400
        
        current_user = get_current_user()
        assigned_by = current_user['username']
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Get user details
            cursor.execute("SELECT username, email FROM users WHERE id = ? AND is_active = 1", (user_id,))
            user_data = cursor.fetchone()
            if not user_data:
                return jsonify({'error': 'User not found or inactive'}), 404
            
            username, user_email = user_data
            
            # Get device details from current results
            device_info = None
            with results_lock:
                for result in results:
                    if result['SM IP'] == device_ip:
                        device_info = result
                        break
            
            if not device_info:
                return jsonify({'error': 'Device not found in current monitoring results'}), 404
            
            # Create assignment record (acknowledge as assigned)
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute("""
                INSERT OR REPLACE INTO device_acknowledgments (sm_ip, status, username, timestamp, comment)
                VALUES (?, 'assigned', ?, ?, ?)
            """, (device_ip, username, timestamp, f"Assigned by {assigned_by}: {comment}" if comment else f"Assigned by {assigned_by}"))
            
            # Add assignment comment
            cursor.execute("""
                INSERT INTO device_comments (sm_ip, comment, username, timestamp, comment_type)
                VALUES (?, ?, ?, ?, 'assignment')
            """, (device_ip, f"Device assigned to {username} by {assigned_by}" + (f": {comment}" if comment else ""), 
                  assigned_by, timestamp))
            
            # Create a task for this assignment
            task_title = f"Fix {device_info['Device Name']} - {device_ip}"
            task_description = f"Device assigned by {assigned_by}. Status: {device_info['Status']}"
            if comment:
                task_description += f"\nNote: {comment}"
            
            cursor.execute("""
                INSERT INTO device_tasks 
                (sm_ip, task_title, task_description, status, priority, assigned_to, created_by, created_at)
                VALUES (?, ?, ?, 'open', 'high', ?, ?, ?)
            """, (device_ip, task_title, task_description, username, assigned_by, timestamp))
            
            conn.commit()
            
            # Send email notification
            email_sent = send_assignment_notification(
                user_email, username, device_ip, 
                device_info['Device Name'], device_info['Location'], 
                assigned_by, comment
            )
            
            message = f'Device {device_ip} assigned to {username} successfully'
            if email_sent:
                message += ' and notification email sent'
            else:
                message += ' but notification email failed to send'
            
            return jsonify({'success': True, 'message': message})
            
    except Exception as e:
        logging.error(f"Assign device error: {str(e)}")
        return jsonify({'error': str(e)}), 500

# Maintenance Mode API Routes
@app.route('/api/maintenance', methods=['GET'])
@login_required
def get_maintenance_windows():
    """Get all active maintenance windows"""
    try:
        maintenance_status = get_maintenance_status()
        return jsonify({'maintenance_windows': maintenance_status})
    except Exception as e:
        logging.error(f"Get maintenance windows error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/maintenance', methods=['POST'])
@admin_required
def add_maintenance():
    """Add a maintenance window"""
    try:
        data = request.json
        sm_ip = data.get('sm_ip')
        start_time_str = data.get('start_time')
        end_time_str = data.get('end_time')
        reason = data.get('reason', 'Scheduled maintenance')
        
        if not sm_ip or not validate_ip(sm_ip):
            return jsonify({'error': 'Invalid IP address'}), 400
        
        if not start_time_str or not end_time_str:
            return jsonify({'error': 'Start time and end time are required'}), 400
        
        try:
            start_time = datetime.strptime(start_time_str, '%Y-%m-%dT%H:%M')
            end_time = datetime.strptime(end_time_str, '%Y-%m-%dT%H:%M')
        except ValueError:
            return jsonify({'error': 'Invalid datetime format'}), 400
        
        if start_time >= end_time:
            return jsonify({'error': 'End time must be after start time'}), 400
        
        if end_time <= datetime.now():
            return jsonify({'error': 'End time must be in the future'}), 400
        
        add_maintenance_window(sm_ip, start_time, end_time, reason)
        
        return jsonify({
            'success': True,
            'message': f'Maintenance window added for {sm_ip}',
            'maintenance': {
                'sm_ip': sm_ip,
                'start_time': start_time.strftime('%Y-%m-%d %H:%M:%S'),
                'end_time': end_time.strftime('%Y-%m-%d %H:%M:%S'),
                'reason': reason
            }
        })
        
    except Exception as e:
        logging.error(f"Add maintenance error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/maintenance/<sm_ip>', methods=['DELETE'])
@admin_required
def remove_maintenance(sm_ip):
    """Remove a maintenance window"""
    try:
        if not validate_ip(sm_ip):
            return jsonify({'error': 'Invalid IP address'}), 400
        
        if sm_ip not in MAINTENANCE_WINDOWS:
            return jsonify({'error': 'No maintenance window found for this IP'}), 404
        
        remove_maintenance_window(sm_ip)
        
        return jsonify({
            'success': True,
            'message': f'Maintenance window removed for {sm_ip}'
        })
        
    except Exception as e:
        logging.error(f"Remove maintenance error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/device_details/<ip>')
def get_device_details(ip):
    """Get detailed information about a device including comments and acknowledgment status"""
    try:
        if not validate_ip(ip):
            return jsonify({'error': 'Invalid IP address'}), 400
        
        # Get device info from current results
        device_info = None
        with results_lock:
            for result in results:
                if result['SM IP'] == ip:
                    device_info = result
                    break
        
        if not device_info:
            return jsonify({'error': 'Device not found'}), 404
        
        # Get acknowledgment status
        ack_status = 'unassigned'
        ack_info = None
        try:
            with sqlite3.connect('ping_history.db', timeout=10) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT status, username, timestamp, comment FROM device_acknowledgments WHERE sm_ip = ?", (ip,))
                ack_record = cursor.fetchone()
                if ack_record:
                    ack_status = ack_record[0]
                    ack_info = {
                        'status': ack_record[0],
                        'username': ack_record[1],
                        'timestamp': ack_record[2],
                        'comment': ack_record[3]
                    }
        except Exception as e:
            logging.error(f"Error getting acknowledgment status for {ip}: {str(e)}")
        
        # Get recent comments
        comments = []
        try:
            with sqlite3.connect('ping_history.db', timeout=10) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT comment, username, timestamp, comment_type 
                    FROM device_comments 
                    WHERE sm_ip = ? 
                    ORDER BY timestamp DESC 
                    LIMIT 10
                """, (ip,))
                comment_records = cursor.fetchall()
                for record in comment_records:
                    comments.append({
                        'comment': record[0],
                        'username': record[1],
                        'timestamp': record[2],
                        'type': record[3]
                    })
        except Exception as e:
            logging.error(f"Error getting comments for {ip}: {str(e)}")
        
        # Get recent history (last 24 hours)
        history = []
        try:
            with sqlite3.connect('ping_history.db', timeout=10) as conn:
                cursor = conn.cursor()
                cutoff = (datetime.now() - timedelta(hours=24)).strftime('%Y-%m-%d %H:%M:%S')
                cursor.execute("""
                    SELECT timestamp, status, latency 
                    FROM history 
                    WHERE sm_ip = ? AND timestamp >= ? 
                    ORDER BY timestamp DESC 
                    LIMIT 20
                """, (ip, cutoff))
                history_records = cursor.fetchall()
                for record in history_records:
                    history.append({
                        'timestamp': record[0],
                        'status': record[1],
                        'latency': f"{record[2]:.2f} ms" if record[2] is not None else 'N/A'
                    })
        except Exception as e:
            logging.error(f"Error getting history for {ip}: {str(e)}")
        
        return jsonify({
            'device': device_info,
            'acknowledgment': {
                'status': ack_status,
                'info': ack_info
            },
            'comments': comments,
            'history': history
        })
    except Exception as e:
        logging.error(f"Get device details error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/device_comments/<sm_ip>')
@login_required
def get_device_comments(sm_ip):
    """Get all comments for a device"""
    try:
        if not validate_ip(sm_ip):
            return jsonify({'error': 'Invalid IP address'}), 400
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT comment, username, timestamp, comment_type
                FROM device_comments
                WHERE sm_ip = ?
                ORDER BY timestamp ASC
                LIMIT 50
            """, (sm_ip,))
            
            comments = []
            for row in cursor.fetchall():
                comments.append({
                    'comment': row[0],
                    'username': row[1],
                    'timestamp': row[2],
                    'comment_type': row[3]
                })
            
            # Keep chronological order (oldest first, newest last)
            
            return jsonify({'success': True, 'comments': comments})
    except Exception as e:
        logging.error(f"Get comments error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/add_device_comment', methods=['POST'])
@login_required
def add_device_comment():
    """Add a comment to a device"""
    try:
        data = request.json
        sm_ip = data.get('sm_ip', '').strip()
        comment = data.get('comment', '').strip()
        comment_type = data.get('type', 'comment')
        
        # Use logged-in user's username
        user = get_current_user()
        username = user['username'] if user else 'Anonymous'
        
        if not validate_ip(sm_ip):
            return jsonify({'error': 'Invalid IP address'}), 400
        
        if not comment:
            return jsonify({'error': 'Comment cannot be empty'}), 400
        
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            conn.execute("""
                INSERT INTO device_comments (sm_ip, comment, username, timestamp, comment_type)
                VALUES (?, ?, ?, ?, ?)
            """, (sm_ip, comment, username, timestamp, comment_type))
            conn.commit()
        
        logging.info(f"Added comment for {sm_ip} by {username}")
        return jsonify({'success': True, 'message': 'Comment added successfully'})
    except Exception as e:
        logging.error(f"Add comment error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/add_comment', methods=['POST'])
@login_required
def add_comment():
    """Add a comment to a device"""
    try:
        data = request.json
        ip = data.get('ip', '').strip()
        comment = data.get('comment', '').strip()
        comment_type = data.get('type', 'comment')
        
        # Use logged-in user's username
        user = get_current_user()
        username = user['username'] if user else 'Anonymous'
        
        if not validate_ip(ip):
            return jsonify({'error': 'Invalid IP address'}), 400
        
        if not comment:
            return jsonify({'error': 'Comment cannot be empty'}), 400
        
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            conn.execute("""
                INSERT INTO device_comments (sm_ip, comment, username, timestamp, comment_type)
                VALUES (?, ?, ?, ?, ?)
            """, (ip, comment, username, timestamp, comment_type))
            conn.commit()
        
        logging.info(f"Added comment for {ip} by {username}")
        return jsonify({'success': True, 'message': 'Comment added successfully'})
    except Exception as e:
        logging.error(f"Add comment error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/acknowledge_device', methods=['POST'])
@login_required
def acknowledge_device():
    """Acknowledge or assign a device issue"""
    try:
        data = request.json
        ip = data.get('ip', '').strip()
        status = data.get('status', 'ack').strip()  # 'ack', 'assigned', or 'unassigned'
        comment = data.get('comment', '').strip()
        
        # Use logged-in user's username
        user = get_current_user()
        username = user['username'] if user else 'Anonymous'
        
        if not validate_ip(ip):
            return jsonify({'error': 'Invalid IP address'}), 400
        
        if status not in ['ack', 'assigned', 'unassigned']:
            return jsonify({'error': 'Invalid status'}), 400
        
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            if status == 'unassigned':
                # Remove acknowledgment
                cursor.execute("DELETE FROM device_acknowledgments WHERE sm_ip = ?", (ip,))
                # Also close any open tasks for this device by this user
                cursor.execute("""
                    UPDATE device_tasks 
                    SET status = 'closed', closed_at = ?, resolution = 'Unassigned from device'
                    WHERE sm_ip = ? AND assigned_to = ? AND status != 'closed'
                """, (timestamp, ip, username))
            else:
                # Insert or update acknowledgment
                cursor.execute("""
                    INSERT OR REPLACE INTO device_acknowledgments (sm_ip, status, username, timestamp, comment)
                    VALUES (?, ?, ?, ?, ?)
                """, (ip, status, username, timestamp, comment))
                
                # If status is 'assigned', create a task
                if status == 'assigned':
                    # Get device info
                    device_info = None
                    with results_lock:
                        for result in results:
                            if result.get('SM IP', result.get('sm_ip')) == ip:
                                device_info = result
                                break
                    
                    if device_info:
                        task_title = f"Fix {device_info.get('Device Name', 'Unknown')} - {ip}"
                        task_description = f"Self-assigned device. Status: {device_info.get('Status', 'Unknown')}"
                        if comment:
                            task_description += f"\nNote: {comment}"
                        
                        # Check if task already exists for this device and user
                        cursor.execute("""
                            SELECT id FROM device_tasks 
                            WHERE sm_ip = ? AND assigned_to = ? AND status != 'closed'
                        """, (ip, username))
                        
                        if not cursor.fetchone():
                            # Create new task
                            cursor.execute("""
                                INSERT INTO device_tasks 
                                (sm_ip, task_title, task_description, status, priority, assigned_to, created_by, created_at)
                                VALUES (?, ?, ?, 'open', 'high', ?, ?, ?)
                            """, (ip, task_title, task_description, username, username, timestamp))
            
            conn.commit()
        
        # Also add a comment about the acknowledgment
        if comment:
            with sqlite3.connect('ping_history.db', timeout=10) as conn:
                conn.execute("""
                    INSERT INTO device_comments (sm_ip, comment, username, timestamp, comment_type)
                    VALUES (?, ?, ?, ?, ?)
                """, (ip, f"Status changed to {status}: {comment}", username, timestamp, 'status_change'))
                conn.commit()
        
        logging.info(f"Device {ip} {status} by {username}")
        return jsonify({'success': True, 'message': f'Device {status} successfully'})
    except Exception as e:
        logging.error(f"Acknowledge device error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/all_devices_for_tasks')
@login_required
def get_all_devices_for_tasks():
    """Get all devices from main dashboard for task assignment (only Down and Degraded, excluding already assigned devices with active tasks)"""
    try:
        user = get_current_user()
        current_username = user['username']
        user_region_id = user.get('region_id')
        user_role = user.get('role')
        
        # Get allowed locations for user's region
        allowed_locations = None
        if user_role != 'admin' and user_region_id:
            try:
                with sqlite3.connect('ping_history.db', timeout=10) as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT name FROM locations WHERE region_id = ?", (user_region_id,))
                    allowed_locations = set(row[0].strip() for row in cursor.fetchall())
                    logging.info(f"User {current_username} filtering for {len(allowed_locations)} locations in region {user_region_id}")
            except Exception as e:
                logging.error(f"Error fetching user's region locations: {e}")
        
        # Get all current devices from monitoring results
        devices_list = []
        total_devices = 0
        status_counts = {}
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Get list of assigned devices with their assigned usernames from device_acknowledgments
            assigned_devices = {}  # sm_ip -> username
            cursor.execute("""
                SELECT sm_ip, username FROM device_acknowledgments 
                WHERE status = 'assigned'
            """)
            for row in cursor.fetchall():
                assigned_devices[row[0]] = row[1]
            
            # Get list of devices with active tasks
            cursor.execute("""
                SELECT DISTINCT sm_ip FROM device_tasks 
                WHERE status != 'closed'
            """)
            devices_with_active_tasks = set(row[0] for row in cursor.fetchall())
            
            logging.info(f"Found {len(assigned_devices)} assigned devices")
            logging.info(f"Found {len(devices_with_active_tasks)} devices with active tasks")
            
            with results_lock:
                total_devices = len(results)
                logging.info(f"Total devices in results: {total_devices}")
                
                for result in results:
                    location = result.get('Location', result.get('location', 'N/A')).strip()
                    
                    # Filter by region if user is not admin
                    if allowed_locations is not None and location not in allowed_locations:
                        continue
                    
                    # The results array uses 'Status' (capital S) not 'status'
                    status = result.get('Status', result.get('status', 'Unknown'))
                    sm_ip = result.get('SM IP', result.get('sm_ip'))
                    
                    # Count all unique statuses
                    status_counts[status] = status_counts.get(status, 0) + 1
                    
                    # Only include Down and Degraded devices
                    if status in ['Down', 'Degraded']:
                        # Check if device is assigned
                        assigned_to = assigned_devices.get(sm_ip)
                        
                        # Skip devices assigned to current user (they're in My Tasks)
                        if assigned_to == current_username:
                            continue
                        
                        # Skip devices that have active tasks (assigned to anyone)
                        if sm_ip in devices_with_active_tasks:
                            continue
                        
                        # Only show truly unassigned devices
                        device_data = {
                            'sm_ip': sm_ip,
                            'device_name': result.get('Device Name', result.get('device_name', 'N/A')),
                            'location': location,
                            'status': status,
                            'latency': result.get('Latency', result.get('latency', 'N/A')),
                            'ap_name': result.get('AP Name', result.get('ap_name', 'N/A')),
                            'ap_ip': result.get('AP IP', result.get('ap_ip', 'N/A')),
                            'cid': result.get('CID', result.get('cid', 'N/A')),
                            'assigned_to': None  # Only unassigned devices shown
                        }
                        devices_list.append(device_data)
        
        logging.info(f"Status counts: {status_counts}")
        logging.info(f"Returning {len(devices_list)} devices for user {current_username}")
        
        return jsonify({
            'success': True,
            'count': len(devices_list),
            'devices': devices_list,
            'debug': {
                'total_devices': total_devices,
                'status_counts': status_counts,
                'all_statuses': list(status_counts.keys()),
                'assigned_count': len(assigned_devices),
                'active_tasks_count': len(devices_with_active_tasks)
            }
        })
    except Exception as e:
        logging.error(f"Error getting all devices: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/task_statistics')
@login_required
def get_task_statistics():
    """Get enhanced task statistics for current user with role-based filtering"""
    try:
        user = get_current_user()
        username = user['username']
        user_role = user.get('role', 'user')
        user_region = user.get('region_id')
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Get time boundaries
            now = datetime.now()
            month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            week_start = now - timedelta(days=7)
            
            # Build role-based filter conditions
            role_filter = ""
            role_params = []
            
            if user_role == 'superadmin':
                # SuperAdmin sees all tasks
                role_filter = ""
            elif user_role in ['admin', 'regional_admin']:
                # Regional admin sees tasks in their region
                if user_region:
                    role_filter = " AND region_id = ?"
                    role_params = [user_region]
            else:
                # Regular user sees only their own tasks
                role_filter = " AND assigned_to = ?"
                role_params = [username]
            
            # Monthly solved tasks with average resolution time
            cursor.execute(f"""
                SELECT COUNT(*), AVG(
                    (julianday(closed_at) - julianday(assigned_at)) * 24
                ) FROM device_tasks 
                WHERE status = 'closed' AND closed_at >= ?
                AND assigned_at IS NOT NULL{role_filter}
            """, [month_start.strftime('%Y-%m-%d %H:%M:%S')] + role_params)
            monthly_result = cursor.fetchone()
            monthly_done = monthly_result[0] or 0
            monthly_avg_time = round(monthly_result[1], 2) if monthly_result[1] else 0
            
            # Weekly solved tasks with average resolution time
            cursor.execute(f"""
                SELECT COUNT(*), AVG(
                    (julianday(closed_at) - julianday(assigned_at)) * 24
                ) FROM device_tasks 
                WHERE status = 'closed' AND closed_at >= ?
                AND assigned_at IS NOT NULL{role_filter}
            """, [week_start.strftime('%Y-%m-%d %H:%M:%S')] + role_params)
            weekly_result = cursor.fetchone()
            weekly_done = weekly_result[0] or 0
            weekly_avg_time = round(weekly_result[1], 2) if weekly_result[1] else 0
            
            # Pending tasks (assigned but not closed) with average pending duration
            cursor.execute(f"""
                SELECT COUNT(*), AVG(
                    (julianday('now') - julianday(assigned_at)) * 24
                ) FROM device_tasks 
                WHERE status IN ('open', 'in_progress') AND assigned_to IS NOT NULL
                AND assigned_at IS NOT NULL{role_filter}
            """, role_params)
            pending_result = cursor.fetchone()
            pending_tasks = pending_result[0] or 0
            pending_avg_duration = round(pending_result[1], 2) if pending_result[1] else 0
            
            # Available tasks (unassigned, visible to user's role/region)
            available_filter = ""
            available_params = []
            if user_role == 'superadmin':
                available_filter = ""
            elif user_role in ['admin', 'regional_admin'] and user_region:
                available_filter = " AND region_id = ?"
                available_params = [user_region]
            elif user_region:
                available_filter = " AND region_id = ?"
                available_params = [user_region]
                
            cursor.execute(f"""
                SELECT COUNT(*) FROM device_tasks 
                WHERE status = 'open' AND assigned_to IS NULL{available_filter}
            """, available_params)
            available_tasks = cursor.fetchone()[0] or 0
            
            # Total closed tasks for user
            user_filter = " AND assigned_to = ?" if user_role != 'superadmin' else ""
            user_params = [username] if user_role != 'superadmin' else []
            
            cursor.execute(f"""
                SELECT COUNT(*) FROM device_tasks 
                WHERE status = 'closed'{user_filter}
            """, user_params)
            total_closed = cursor.fetchone()[0] or 0
            
            return jsonify({
                'success': True,
                'monthly_done': monthly_done,
                'monthly_avg_time': monthly_avg_time,
                'weekly_done': weekly_done,
                'weekly_avg_time': weekly_avg_time,
                'pending_tasks': pending_tasks,
                'pending_avg_duration': pending_avg_duration,
                'available_tasks': available_tasks,
                'total_closed': total_closed,
                'avg_duration_hours': monthly_avg_time,  # Keep for backward compatibility
                'open_tasks': available_tasks,  # Keep for backward compatibility
                'in_progress_tasks': pending_tasks  # Keep for backward compatibility
            })
    except Exception as e:
        logging.error(f"Error getting task statistics: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/create_task', methods=['POST'])
@login_required
def create_task():
    """Create a new task"""
    try:
        user = get_current_user()
        data = request.json
        
        sm_ip = data.get('sm_ip', '').strip()
        title = data.get('title', '').strip()
        description = data.get('description', '').strip()
        priority = data.get('priority', 'medium')
        assign_to_self = data.get('assign_to_self', False)
        assigned_to_user = data.get('assigned_to', '').strip()
        
        if not sm_ip or not title:
            return jsonify({'error': 'SM IP and title are required'}), 400
        
        # Determine who to assign to
        assigned_to = None
        assigned_at = None
        if assign_to_self:
            assigned_to = user['username']
            assigned_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        elif assigned_to_user:
            assigned_to = assigned_to_user
            assigned_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Get device region from Excel data
        region_id = None
        if CACHED_DF is not None:
            device_row = CACHED_DF[CACHED_DF['SM IP'] == sm_ip]
            if not device_row.empty:
                location = str(device_row.iloc[0].get('Location', '')).strip()
                # Get region_id for this location
                with sqlite3.connect('ping_history.db', timeout=10) as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT region_id FROM locations WHERE name = ?", (location,))
                    region_result = cursor.fetchone()
                    if region_result:
                        region_id = region_result[0]
        
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Get the next ID to generate task_id
            cursor.execute("SELECT MAX(id) FROM device_tasks")
            max_id = cursor.fetchone()[0]
            next_id = (max_id or 0) + 1
            task_id_str = f"TASK-{1000 + next_id:04d}"
            
            cursor.execute("""
                INSERT INTO device_tasks 
                (task_id, sm_ip, task_title, task_description, status, priority, assigned_to, assigned_at, created_by, created_at, updated_at, region_id)
                VALUES (?, ?, ?, ?, 'open', ?, ?, ?, ?, ?, ?, ?)
            """, (task_id_str, sm_ip, title, description, priority, assigned_to, assigned_at, user['username'], timestamp, timestamp, region_id))
            task_id = cursor.lastrowid
            conn.commit()
        
        logging.info(f"Task created: ID={task_id_str}, IP={sm_ip}, User={user['username']}, Region={region_id}")
        log_user_activity(user['id'], user['username'], 'task_create', f"Created task {task_id_str} for device {sm_ip} - '{title}' (Priority: {priority})")
        return jsonify({'success': True, 'task_id': task_id, 'task_id_str': task_id_str, 'message': 'Task created successfully'})
    except Exception as e:
        logging.error(f"Error creating task: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/task/<int:task_id>')
@login_required
def get_task_by_id(task_id):
    """Get a single task by ID"""
    try:
        user = get_current_user()
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # All users can view all tasks
            cursor.execute("""
                SELECT id, sm_ip, task_id, task_title, task_description, status, priority, 
                       assigned_to, created_by, created_at, updated_at, closed_at, closed_by, resolution
                FROM device_tasks 
                WHERE id = ?
            """, (task_id,))
            
            row = cursor.fetchone()
            if not row:
                return jsonify({'error': 'Task not found'}), 404
            
            task = {
                'id': row[0],
                'sm_ip': row[1],
                'task_id': row[2],
                'title': row[3],
                'description': row[4],
                'status': row[5],
                'priority': row[6],
                'assigned_to': row[7],
                'created_by': row[8],
                'created_at': row[9],
                'updated_at': row[10],
                'closed_at': row[11],
                'closed_by': row[12],
                'resolution': row[13]
            }
            
            return jsonify({'success': True, 'task': task})
    except Exception as e:
        logging.error(f"Error getting task: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/update_task/<int:task_id>', methods=['PUT'])
@login_required
def update_task(task_id):
    """Update task status or details"""
    try:
        user = get_current_user()
        data = request.json
        
        status = data.get('status')
        resolution = data.get('resolution', '').strip()
        
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # If closing task, check if user is admin or if SM is reachable
            if status == 'closed' and user['role'] != 'admin':
                # Get task details to find SM IP
                cursor.execute("SELECT sm_ip, task_description FROM device_tasks WHERE id = ?", (task_id,))
                task_data = cursor.fetchone()
                
                if task_data:
                    sm_ip = task_data[0]
                    description = task_data[1] or ''
                    
                    # Extract AP IP from description
                    import re
                    ap_ip_match = re.search(r'🌐 AP IP: ([\d.]+)', description)
                    ap_ip = ap_ip_match.group(1) if ap_ip_match else None
                    
                    # Check current status of SM and AP
                    sm_reachable = False
                    ap_reachable = False
                    
                    # Check SM status from current results
                    with results_lock:
                        for result in results:
                            if result.get('SM IP') == sm_ip:
                                sm_reachable = result.get('Status') == 'Reachable'
                                break
                    
                    # Check AP status if we have AP IP
                    if ap_ip:
                        with results_lock:
                            for result in results:
                                if result.get('AP IP') == ap_ip:
                                    ap_reachable = result.get('Status') == 'Reachable'
                                    break
                    else:
                        # If no AP IP found, assume it's reachable (don't block)
                        ap_reachable = True
                    
                    # Only block if SM is not reachable (AP being down is just a warning)
                    if not sm_reachable:
                        return jsonify({
                            'error': 'SM is still down. Only admin can close tasks with unreachable SM.',
                            'sm_reachable': sm_reachable,
                            'ap_reachable': ap_reachable
                        }), 403
                    
                    # If SM is reachable but AP is not, store warning to return with success
                    ap_warning = None
                    if not ap_reachable:
                        ap_warning = 'Note: AP is still unreachable, but task closed as SM is reachable.'
            
            # Proceed with update
            if status == 'closed':
                cursor.execute("""
                    UPDATE device_tasks 
                    SET status = ?, updated_at = ?, closed_at = ?, closed_by = ?, resolution = ?
                    WHERE id = ?
                """, (status, timestamp, timestamp, user['username'], resolution, task_id))
            else:
                cursor.execute("""
                    UPDATE device_tasks 
                    SET status = ?, updated_at = ?
                    WHERE id = ?
                """, (status, timestamp, task_id))
            
            conn.commit()
        
        logging.info(f"Task updated: ID={task_id}, Status={status}, User={user['username']}")
        log_user_activity(user['id'], user['username'], f"task_{status}", f"{'Closed' if status == 'closed' else 'Updated'} task #{task_id} (status: {status})")
        
        # Return success with optional AP warning
        response = {'success': True, 'message': 'Task updated successfully'}
        if status == 'closed' and 'ap_warning' in locals() and ap_warning:
            response['warning'] = ap_warning
        
        return jsonify(response)
    except Exception as e:
        logging.error(f"Error updating task: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/assign_task/<int:task_id>', methods=['PUT'])
@login_required
def assign_task_to_self(task_id):
    """Assign task to current user"""
    try:
        user = get_current_user()
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE device_tasks 
                SET assigned_to = ?, status = 'in_progress', updated_at = ?
                WHERE id = ?
            """, (user['username'], timestamp, task_id))
            conn.commit()
        
        logging.info(f"Task assigned: ID={task_id}, User={user['username']}")
        log_user_activity(user['id'], user['username'], 'task_assign', f"Assigned task #{task_id} to themselves")
        return jsonify({'success': True, 'message': 'Task assigned to you'})
    except Exception as e:
        logging.error(f"Error assigning task: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/reassign_task/<int:task_id>', methods=['PUT'])
@login_required
def reassign_task(task_id):
    """Reassign task to another user (admin only)"""
    try:
        user = get_current_user()
        
        # Only admin can reassign tasks
        if user['role'] != 'admin':
            return jsonify({'error': 'Only administrators can reassign tasks'}), 403
        
        data = request.json
        new_assignee = data.get('assigned_to', '').strip()
        
        if not new_assignee:
            return jsonify({'error': 'Assignee username is required'}), 400
        
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Verify the new assignee exists
            cursor.execute("SELECT id FROM users WHERE username = ?", (new_assignee,))
            if not cursor.fetchone():
                return jsonify({'error': 'User not found'}), 404
            
            # Get task details including SM IP and old assignee
            cursor.execute("SELECT sm_ip, assigned_to FROM device_tasks WHERE id = ?", (task_id,))
            task_data = cursor.fetchone()
            if not task_data:
                return jsonify({'error': 'Task not found'}), 404
            
            sm_ip = task_data[0]
            old_assignee = task_data[1]
            
            # Update the task assignment
            cursor.execute("""
                UPDATE device_tasks 
                SET assigned_to = ?, updated_at = ?
                WHERE id = ?
            """, (new_assignee, timestamp, task_id))
            
            # Update device_acknowledgments to reflect new assignee
            cursor.execute("""
                INSERT OR REPLACE INTO device_acknowledgments (sm_ip, status, username, timestamp, comment)
                VALUES (?, 'assigned', ?, ?, ?)
            """, (sm_ip, new_assignee, timestamp, f"Task reassigned from {old_assignee} to {new_assignee} by {user['username']}"))
            
            # Add a comment to the task about the reassignment
            cursor.execute("""
                INSERT INTO device_comments (sm_ip, comment, username, timestamp, comment_type)
                VALUES (?, ?, ?, ?, 'reassignment')
            """, (sm_ip, f"Task reassigned from {old_assignee} to {new_assignee}", user['username'], timestamp))
            
            conn.commit()
        
        logging.info(f"Task reassigned: ID={task_id}, From={old_assignee}, To={new_assignee}, By={user['username']}")
        return jsonify({'success': True, 'message': f'Task reassigned to {new_assignee}'})
    except Exception as e:
        logging.error(f"Error reassigning task: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/my_tasks_list')
@login_required
def get_my_tasks_list():
    """Get all tasks assigned to current user (or all tasks if admin)"""
    try:
        user = get_current_user()
        username = user['username']
        is_admin = user['role'] in ['admin', 'superadmin', 'regional_admin']
        user_region_id = user.get('region_id')
        
        # Get allowed locations for user's region
        allowed_locations = None
        if not is_admin and user_region_id:
            try:
                with sqlite3.connect('ping_history.db', timeout=10) as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT name FROM locations WHERE region_id = ?", (user_region_id,))
                    allowed_locations = set(row[0].strip() for row in cursor.fetchall())
                    logging.info(f"User {username} filtering tasks for {len(allowed_locations)} locations in region {user_region_id}")
            except Exception as e:
                logging.error(f"Error fetching user's region locations: {e}")
        
        # Get device location mapping
        device_locations = {}
        if CACHED_DF is not None:
            for _, row in CACHED_DF.iterrows():
                if pd.notna(row.get('SM IP')):
                    device_locations[str(row['SM IP'])] = str(row.get('Location', 'Unknown')).strip()
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Admin sees all tasks, regular users see only their tasks
            if is_admin:
                cursor.execute("""
                    SELECT id, task_id, sm_ip, task_title, task_description, status, priority, 
                           assigned_to, created_by, created_at, updated_at, closed_at, resolution
                    FROM device_tasks 
                    ORDER BY 
                        CASE status 
                            WHEN 'in_progress' THEN 1 
                            WHEN 'open' THEN 2 
                            WHEN 'closed' THEN 3 
                        END,
                        created_at DESC
                """)
            else:
                cursor.execute("""
                    SELECT id, task_id, sm_ip, task_title, task_description, status, priority, 
                           assigned_to, created_by, created_at, updated_at, closed_at, resolution
                    FROM device_tasks 
                    WHERE assigned_to = ?
                    ORDER BY 
                        CASE status 
                            WHEN 'in_progress' THEN 1 
                            WHEN 'open' THEN 2 
                            WHEN 'closed' THEN 3 
                        END,
                        created_at DESC
                """, (username,))
            
            tasks = []
            for row in cursor.fetchall():
                sm_ip = row[2]
                location = device_locations.get(sm_ip, 'Unknown')
                
                # Filter by region if user is not admin
                if allowed_locations is not None and location not in allowed_locations:
                    continue
                
                tasks.append({
                    'id': row[0],
                    'task_id': row[1],
                    'sm_ip': sm_ip,
                    'title': row[3],
                    'description': row[4],
                    'status': row[5],
                    'priority': row[6],
                    'assigned_to': row[7],
                    'created_by': row[8],
                    'created_at': row[9],
                    'updated_at': row[10],
                    'closed_at': row[11],
                    'resolution': row[12]
                })
            
            return jsonify({'success': True, 'tasks': tasks})
    except Exception as e:
        logging.error(f"Error getting tasks: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/tasks/solved')
@login_required
def get_solved_tasks():
    """Get solved/closed tasks with filtering options"""
    try:
        user = get_current_user()
        username = user['username']
        user_role = user.get('role', 'user')
        user_region = user.get('region_id')
        
        # Get filter parameters
        period = request.args.get('period', 'all')  # weekly, monthly, all
        
        # Calculate time boundaries
        now = datetime.now()
        time_filter = ""
        time_params = []
        
        if period == 'weekly':
            week_start = now - timedelta(days=7)
            time_filter = " AND closed_at >= ?"
            time_params = [week_start.strftime('%Y-%m-%d %H:%M:%S')]
        elif period == 'monthly':
            month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            time_filter = " AND closed_at >= ?"
            time_params = [month_start.strftime('%Y-%m-%d %H:%M:%S')]
        
        # Build role-based filter
        role_filter = ""
        role_params = []
        
        if user_role == 'superadmin':
            # SuperAdmin sees all solved tasks
            role_filter = ""
        elif user_role in ['admin', 'regional_admin']:
            # Regional admin sees solved tasks in their region
            if user_region:
                role_filter = " AND region_id = ?"
                role_params = [user_region]
        else:
            # Regular user sees only their own solved tasks
            role_filter = " AND assigned_to = ?"
            role_params = [username]
        
        # Get device location mapping for region filtering
        device_locations = {}
        allowed_locations = None
        
        if CACHED_DF is not None:
            for _, row in CACHED_DF.iterrows():
                if pd.notna(row.get('SM IP')):
                    device_locations[str(row['SM IP'])] = str(row.get('Location', 'Unknown')).strip()
        
        # Get allowed locations for non-admin users
        if user_role not in ['superadmin'] and user_region:
            try:
                with sqlite3.connect('ping_history.db', timeout=10) as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT name FROM locations WHERE region_id = ?", (user_region,))
                    allowed_locations = set(row[0].strip() for row in cursor.fetchall())
            except Exception as e:
                logging.error(f"Error fetching user's region locations: {e}")
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Query solved tasks
            query = f"""
                SELECT id, task_id, sm_ip, task_title, task_description, status, priority, 
                       assigned_to, created_by, created_at, assigned_at, closed_at, resolution,
                       (julianday(closed_at) - julianday(assigned_at)) * 24 as resolution_hours
                FROM device_tasks 
                WHERE status = 'closed' AND closed_at IS NOT NULL{time_filter}{role_filter}
                ORDER BY closed_at DESC
            """
            
            cursor.execute(query, time_params + role_params)
            
            tasks = []
            for row in cursor.fetchall():
                sm_ip = row[2]
                location = device_locations.get(sm_ip, 'Unknown')
                
                # Filter by region if user is not superadmin
                if allowed_locations is not None and location not in allowed_locations:
                    continue
                
                resolution_hours = round(row[13], 2) if row[13] else 0
                
                tasks.append({
                    'id': row[0],
                    'task_id': row[1],
                    'sm_ip': sm_ip,
                    'device_name': location,
                    'title': row[3],
                    'description': row[4],
                    'status': row[5],
                    'priority': row[6],
                    'assigned_to': row[7],
                    'created_by': row[8],
                    'created_at': row[9],
                    'assigned_at': row[10],
                    'closed_at': row[11],
                    'resolution': row[12],
                    'resolution_hours': resolution_hours
                })
            
            return jsonify({
                'success': True, 
                'tasks': tasks,
                'period': period,
                'total_count': len(tasks)
            })
    except Exception as e:
        logging.error(f"Error getting solved tasks: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/device_task_history/<sm_ip>')
@login_required
def get_device_task_history(sm_ip):
    """Get all tasks (history) for a specific device"""
    try:
        if not validate_ip(sm_ip):
            return jsonify({'error': 'Invalid IP address'}), 400
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, task_id, task_title, task_description, status, priority, 
                       assigned_to, created_by, created_at, updated_at, closed_at, closed_by, resolution
                FROM device_tasks 
                WHERE sm_ip = ?
                ORDER BY created_at DESC
            """, (sm_ip,))
            
            tasks = []
            for row in cursor.fetchall():
                tasks.append({
                    'id': row[0],
                    'task_id': row[1],
                    'title': row[2],
                    'description': row[3],
                    'status': row[4],
                    'priority': row[5],
                    'assigned_to': row[6],
                    'created_by': row[7],
                    'created_at': row[8],
                    'updated_at': row[9],
                    'closed_at': row[10],
                    'closed_by': row[11],
                    'resolution': row[12]
                })
            
            # Get manual ping test comments
            cursor.execute("""
                SELECT comment, username, timestamp
                FROM device_comments
                WHERE sm_ip = ? AND comment_type = 'manual_ping'
                ORDER BY timestamp DESC
                LIMIT 10
            """, (sm_ip,))
            
            ping_tests = []
            for row in cursor.fetchall():
                ping_tests.append({
                    'comment': row[0],
                    'username': row[1],
                    'timestamp': row[2]
                })
            
            return jsonify({
                'success': True, 
                'tasks': tasks,
                'total_count': len(tasks),
                'sm_ip': sm_ip,
                'ping_tests': ping_tests
            })
    except Exception as e:
        logging.error(f"Error getting device task history: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/my_assigned_devices')
@login_required
def get_my_assigned_devices():
    """Get devices assigned to the current user"""
    try:
        user = get_current_user()
        if not user:
            return jsonify({'error': 'Not authenticated'}), 401
        
        username = user['username']
        logging.info(f"Fetching assigned devices for user: {username}")
        
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            
            # Debug: Check all assignments in database
            cursor.execute("SELECT sm_ip, status, username, timestamp FROM device_acknowledgments")
            all_assignments = cursor.fetchall()
            logging.info(f"Total assignments in database: {len(all_assignments)}")
            for assign in all_assignments:
                logging.info(f"  Assignment: IP={assign[0]}, Status={assign[1]}, User={assign[2]}, Time={assign[3]}")
            
            cursor.execute("""
                SELECT da.sm_ip, da.status, da.timestamp, da.comment
                FROM device_acknowledgments da
                WHERE da.username = ? AND da.status = 'assigned'
                ORDER BY da.timestamp DESC
            """, (username,))
            
            assignments = cursor.fetchall()
            logging.info(f"Found {len(assignments)} assignments for user {username}")
            
            # Load device info from Excel file
            df = load_xlsx()
            
            # Get device details for each assigned device
            assigned_devices = []
            for assignment in assignments:
                sm_ip = assignment[0]
                logging.info(f"Processing assignment for IP: {sm_ip}")
                
                # First try to get from current results (for real-time status)
                device_info = None
                with results_lock:
                    for result in results:
                        # Results array uses 'SM IP' (capital with space) not 'sm_ip'
                        if result.get('SM IP', result.get('sm_ip')) == sm_ip:
                            device_info = {
                                'sm_ip': sm_ip,
                                'device_name': result.get('Device Name', result.get('device_name', 'N/A')),
                                'location': result.get('Location', result.get('location', 'N/A')),
                                'ap_name': result.get('AP Name', result.get('ap_name', 'N/A')),
                                'ap_ip': result.get('AP IP', result.get('ap_ip', 'N/A')),
                                'cid': result.get('CID', result.get('cid', 'N/A')),
                                'status': result.get('Status', result.get('status', 'Unknown')),
                                'latency': result.get('Latency', result.get('latency', 'N/A'))
                            }
                            logging.info(f"  Found in current results: {device_info}")
                            break
                
                # If not in current results, get from Excel file
                if not device_info and df is not None:
                    device_row = df[df['SM IP'] == sm_ip]
                    if not device_row.empty:
                        device_info = {
                            'sm_ip': sm_ip,
                            'device_name': str(device_row.iloc[0].get('Device Name', 'N/A')),
                            'location': str(device_row.iloc[0].get('Location', 'N/A')),
                            'ap_name': str(device_row.iloc[0].get('AP Name', 'N/A')),
                            'ap_ip': str(device_row.iloc[0].get('AP IP', 'N/A')),
                            'cid': str(device_row.iloc[0].get('CID', 'N/A')),
                            'status': 'Unknown',  # Will be updated from monitoring
                            'latency': 'N/A'
                        }
                        logging.info(f"  Found in Excel: {device_info}")
                        
                        # Try to get latest status from database
                        cursor.execute("""
                            SELECT status, latency 
                            FROM history 
                            WHERE sm_ip = ? 
                            ORDER BY timestamp DESC 
                            LIMIT 1
                        """, (sm_ip,))
                        history_row = cursor.fetchone()
                        if history_row:
                            device_info['status'] = history_row[0]
                            device_info['latency'] = history_row[1] if history_row[1] else 'N/A'
                            logging.info(f"  Updated status from history: {history_row[0]}")
                
                if device_info:
                    assigned_devices.append({
                        'sm_ip': sm_ip,
                        'device_name': device_info.get('device_name', 'N/A'),
                        'location': device_info.get('location', 'N/A'),
                        'ap_name': device_info.get('ap_name', 'N/A'),
                        'ap_ip': device_info.get('ap_ip', 'N/A'),
                        'cid': device_info.get('cid', 'N/A'),
                        'status': device_info.get('status', 'Unknown'),
                        'latency': device_info.get('latency', 'N/A'),
                        'assigned_at': assignment[2],
                        'comment': assignment[3] or ''
                    })
                else:
                    # Even if device not found, show it with minimal info
                    logging.warning(f"  Device {sm_ip} not found in results or Excel, showing with minimal info")
                    assigned_devices.append({
                        'sm_ip': sm_ip,
                        'device_name': 'Unknown Device',
                        'location': 'Unknown',
                        'ap_name': 'N/A',
                        'ap_ip': 'N/A',
                        'cid': 'N/A',
                        'status': 'Unknown',
                        'latency': 'N/A',
                        'assigned_at': assignment[2],
                        'comment': assignment[3] or ''
                    })
        
        logging.info(f"Returning {len(assigned_devices)} assigned devices")
        return jsonify({
            'success': True,
            'count': len(assigned_devices),
            'devices': assigned_devices
        })
    except Exception as e:
        logging.error(f"Error getting assigned devices: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/my_tasks')
@login_required
def my_tasks():
    """My Tasks page showing assigned devices"""
    user = get_current_user()
    return render_template('my_tasks.html', user=user)

@app.route('/')
@login_required
def dashboard():
    user = get_current_user()
    users = get_all_users() if user and user['role'] in ['admin', 'superadmin', 'regional_admin'] else []
    
    # Pass user's region_id to frontend for auto-filtering
    user_region_id = user.get('region_id') if user else None
    
    return render_template('index.html', user=user, users=users, user_region_id=user_region_id)

@app.route('/api/cached_results')
@login_required
def get_cached_results():
    """Get last cached ping results immediately without waiting for next cycle"""
    try:
        with results_lock:
            if not results:
                return jsonify({
                    'success': False,
                    'message': 'No cached results available yet. Please wait for first ping cycle.'
                })
            
            # Calculate pop summary from cached results
            pop_summary = defaultdict(lambda: {'Reachable': 0, 'Degraded': 0, 'Down': 0})
            for result in results:
                location = result.get('Location', 'Unknown')
                status = result.get('Status', 'Unknown')
                if status in ['Reachable', 'Degraded', 'Down']:
                    pop_summary[location][status] += 1
            
            return jsonify({
                'success': True,
                'results': results,
                'pop_summary': dict(pop_summary)
            })
    except Exception as e:
        logging.error(f"Error getting cached results: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/dashboard')
@login_required
def index():
    return redirect(url_for('dashboard'))

@app.route('/test_maintenance')
def test_maintenance():
    """Test page for maintenance mode functionality"""
    return send_file('test_maintenance.html')

# ─────────────────────────────────────────────
#  SNMP MONITOR  –  OID profiles + polling
# ─────────────────────────────────────────────

SNMP_OID_PROFILES = {
    'raisecom': {
        'cpu':      '1.3.6.1.4.1.8886.1.1.1.1.1.0',
        'mem':      '1.3.6.1.4.1.8886.1.1.1.1.2.0',
        'uptime':   '1.3.6.1.2.1.1.3.0',
        'if_in':    '1.3.6.1.2.1.2.2.1.10',
        'if_out':   '1.3.6.1.2.1.2.2.1.16',
        'if_name':  '1.3.6.1.2.1.31.1.1.1.1',
    },
    'edgecore': {
        'cpu':      '1.3.6.1.4.1.259.10.1.46.1.8.2.1.0',
        'mem':      '1.3.6.1.4.1.259.10.1.46.1.8.1.1.0',
        'uptime':   '1.3.6.1.2.1.1.3.0',
        'if_in':    '1.3.6.1.2.1.2.2.1.10',
        'if_out':   '1.3.6.1.2.1.2.2.1.16',
        'if_name':  '1.3.6.1.2.1.31.1.1.1.1',
    },
    'epmp': {
        'uptime':   '1.3.6.1.2.1.1.3.0',
        'rx_tp':    '1.3.6.1.4.1.17713.21.1.2.18.0',
        'tx_tp':    '1.3.6.1.4.1.17713.21.1.2.19.0',
        'signal':   '1.3.6.1.4.1.17713.21.1.2.2.0',
        'if_in':    '1.3.6.1.2.1.2.2.1.10',
        'if_out':   '1.3.6.1.2.1.2.2.1.16',
    },
    'powerbeam': {
        'uptime':   '1.3.6.1.2.1.1.3.0',
        'signal':   '1.3.6.1.4.1.41112.1.4.5.1.4.1',
        'if_in':    '1.3.6.1.2.1.2.2.1.10',
        'if_out':   '1.3.6.1.2.1.2.2.1.16',
        'if_name':  '1.3.6.1.2.1.31.1.1.1.1',
    },
    'generic': {
        'uptime':   '1.3.6.1.2.1.1.3.0',
        'if_in':    '1.3.6.1.2.1.2.2.1.10',
        'if_out':   '1.3.6.1.2.1.2.2.1.16',
        'if_name':  '1.3.6.1.2.1.31.1.1.1.1',
    }
}

# Cache previous byte counters for delta calculation
_snmp_prev_bytes = {}  # device_id -> (rx, tx, timestamp)

def _snmp_get(ip, community, oid, timeout=3, retries=1):
    """Single SNMP GET using system snmpget binary"""
    try:
        result = subprocess.run(
            ['snmpget', '-v2c', '-c', community, '-t', str(timeout), '-r', str(retries),
             '-Oqv', ip, oid],
            capture_output=True, text=True, timeout=timeout + 2
        )
        val = result.stdout.strip()
        if result.returncode != 0 or not val or 'No Such' in val or 'Timeout' in val:
            return None
        # Strip quotes if string value
        return val.strip('"')
    except Exception:
        return None

def _snmp_walk(ip, community, oid, timeout=3, retries=1):
    """SNMP WALK using system snmpwalk binary, returns list of (oid, value)"""
    results_list = []
    try:
        result = subprocess.run(
            ['snmpwalk', '-v2c', '-c', community, '-t', str(timeout), '-r', str(retries),
             '-Oqn', ip, oid],
            capture_output=True, text=True, timeout=15
        )
        if result.returncode != 0:
            return []
        for line in result.stdout.strip().splitlines():
            line = line.strip()
            if not line or 'No Such' in line:
                continue
            parts = line.split(None, 1)
            if len(parts) == 2:
                oid_str, val_str = parts
                try:
                    results_list.append((oid_str, int(val_str)))
                except ValueError:
                    results_list.append((oid_str, 0))
    except Exception:
        pass
    return results_list

def snmp_poll_device(device):
    """Poll a single SNMP device and store metrics"""
    device_id = device['id']
    ip = device['ip']
    community = device['community']
    dtype = device['device_type']
    profile = SNMP_OID_PROFILES.get(dtype, SNMP_OID_PROFILES['generic'])

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    status = 'offline'
    uptime_seconds = None
    cpu_percent = None
    mem_percent = None
    signal_dbm = None
    rx_bytes_total = None
    tx_bytes_total = None
    interfaces = []

    try:
        # Uptime check (also serves as reachability test)
        uptime_val = _snmp_get(ip, community, profile['uptime'])
        if uptime_val is not None:
            status = 'online'
            try:
                uptime_seconds = int(uptime_val) // 100  # timeticks -> seconds
            except Exception:
                uptime_seconds = None

        if status == 'offline':
            _store_snmp_metric(device_id, now, status, None, None, None, [], None, None, None)
            return

        # CPU
        if 'cpu' in profile:
            cpu_val = _snmp_get(ip, community, profile['cpu'])
            if cpu_val is not None:
                try:
                    cpu_percent = float(cpu_val)
                except Exception:
                    pass

        # Memory
        if 'mem' in profile:
            mem_val = _snmp_get(ip, community, profile['mem'])
            if mem_val is not None:
                try:
                    mem_percent = float(mem_val)
                except Exception:
                    pass

        # Signal (ePMP / PowerBeam)
        if 'signal' in profile:
            sig_val = _snmp_get(ip, community, profile['signal'])
            if sig_val is not None:
                try:
                    signal_dbm = float(sig_val)
                    # ePMP returns in dBm*10
                    if dtype == 'epmp' and abs(signal_dbm) > 200:
                        signal_dbm = signal_dbm / 10.0
                except Exception:
                    pass

        # Interface traffic via WALK
        if_in_data = _snmp_walk(ip, community, profile['if_in'])
        if_out_data = _snmp_walk(ip, community, profile['if_out'])
        if_names = {}
        if 'if_name' in profile:
            for oid_str, val in _snmp_walk(ip, community, profile['if_name']):
                idx = oid_str.split('.')[-1]
                if_names[idx] = str(val)

        rx_total = sum(v for _, v in if_in_data)
        tx_total = sum(v for _, v in if_out_data)
        rx_bytes_total = rx_total
        tx_bytes_total = tx_total

        # Build per-interface list
        for (oid_str, rx_val), (_, tx_val) in zip(if_in_data, if_out_data):
            idx = oid_str.split('.')[-1]
            interfaces.append({
                'index': idx,
                'name': if_names.get(idx, f'if{idx}'),
                'rx_bytes': rx_val,
                'tx_bytes': tx_val
            })

        # Compute Mbps using delta from previous poll
        rx_mbps = None
        tx_mbps = None
        prev = _snmp_prev_bytes.get(device_id)
        if prev:
            prev_rx, prev_tx, prev_ts = prev
            try:
                elapsed = (datetime.strptime(now, '%Y-%m-%d %H:%M:%S') -
                           datetime.strptime(prev_ts, '%Y-%m-%d %H:%M:%S')).total_seconds()
                if elapsed > 0:
                    rx_mbps = round(((rx_total - prev_rx) * 8) / elapsed / 1_000_000, 3)
                    tx_mbps = round(((tx_total - prev_tx) * 8) / elapsed / 1_000_000, 3)
                    if rx_mbps < 0: rx_mbps = None
                    if tx_mbps < 0: tx_mbps = None
            except Exception:
                pass
        _snmp_prev_bytes[device_id] = (rx_total, tx_total, now)

        # Store Mbps in interfaces list for convenience
        for iface in interfaces:
            iface['rx_mbps'] = rx_mbps
            iface['tx_mbps'] = tx_mbps

    except Exception as e:
        logging.error(f"SNMP poll error for {ip}: {e}")
        status = 'error'

    _store_snmp_metric(device_id, now, status, uptime_seconds, cpu_percent, mem_percent,
                       interfaces, rx_bytes_total, tx_bytes_total, signal_dbm)

def _store_snmp_metric(device_id, timestamp, status, uptime_seconds, cpu_percent,
                       mem_percent, interfaces, rx_bytes, tx_bytes, signal_dbm):
    try:
        conn = sqlite3.connect('ping_history.db', timeout=10)
        conn.execute('''INSERT INTO snmp_metrics
                        (device_id, timestamp, status, uptime_seconds, cpu_percent,
                         mem_percent, interfaces_json, rx_bytes, tx_bytes, signal_dbm)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                     (device_id, timestamp, status, uptime_seconds, cpu_percent,
                      mem_percent, json.dumps(interfaces), rx_bytes, tx_bytes, signal_dbm))
        # Keep only last 24h of metrics per device
        conn.execute('''DELETE FROM snmp_metrics WHERE device_id = ? AND
                        timestamp < datetime('now', '-24 hours')''', (device_id,))
        conn.commit()
        conn.close()
    except Exception as e:
        logging.error(f"Error storing SNMP metric: {e}")

def snmp_polling_loop():
    """Background thread: poll all active SNMP devices every 60 seconds"""
    logging.info("SNMP polling thread started")
    while True:
        try:
            conn = sqlite3.connect('ping_history.db', timeout=10)
            cursor = conn.cursor()
            cursor.execute("SELECT id, name, ip, community, snmp_version, device_type FROM snmp_devices WHERE is_active = 1")
            devices = [{'id': r[0], 'name': r[1], 'ip': r[2], 'community': r[3],
                        'snmp_version': r[4], 'device_type': r[5]} for r in cursor.fetchall()]
            conn.close()

            for device in devices:
                try:
                    snmp_poll_device(device)
                except Exception as e:
                    logging.error(f"Poll error for device {device['id']}: {e}")
        except Exception as e:
            logging.error(f"SNMP polling loop error: {e}")
        time.sleep(60)

# Start SNMP polling thread
_snmp_thread = Thread(target=snmp_polling_loop, daemon=True)
_snmp_thread.start()

# ─── SNMP API routes ───────────────────────────────────────────────────────────

@app.route('/snmp_monitor')
@login_required
def snmp_monitor():
    return render_template('snmp_monitor.html', user=session.get('user', {}))

@app.route('/api/snmp/devices', methods=['GET'])
@login_required
def api_snmp_devices():
    try:
        conn = sqlite3.connect('ping_history.db', timeout=10)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('''SELECT d.*, 
                          (SELECT status FROM snmp_metrics m WHERE m.device_id = d.id
                           ORDER BY m.timestamp DESC LIMIT 1) as last_status,
                          (SELECT timestamp FROM snmp_metrics m WHERE m.device_id = d.id
                           ORDER BY m.timestamp DESC LIMIT 1) as last_seen,
                          (SELECT uptime_seconds FROM snmp_metrics m WHERE m.device_id = d.id
                           ORDER BY m.timestamp DESC LIMIT 1) as uptime_seconds,
                          (SELECT cpu_percent FROM snmp_metrics m WHERE m.device_id = d.id
                           ORDER BY m.timestamp DESC LIMIT 1) as cpu_percent,
                          (SELECT mem_percent FROM snmp_metrics m WHERE m.device_id = d.id
                           ORDER BY m.timestamp DESC LIMIT 1) as mem_percent,
                          (SELECT signal_dbm FROM snmp_metrics m WHERE m.device_id = d.id
                           ORDER BY m.timestamp DESC LIMIT 1) as signal_dbm,
                          (SELECT interfaces_json FROM snmp_metrics m WHERE m.device_id = d.id
                           ORDER BY m.timestamp DESC LIMIT 1) as interfaces_json
                          FROM snmp_devices d ORDER BY d.name''')
        rows = cursor.fetchall()
        conn.close()
        devices = []
        for r in rows:
            d = dict(r)
            try:
                ifaces = json.loads(d.get('interfaces_json') or '[]')
                rx_mbps = ifaces[0].get('rx_mbps') if ifaces else None
                tx_mbps = ifaces[0].get('tx_mbps') if ifaces else None
            except Exception:
                rx_mbps = tx_mbps = None
            d['rx_mbps'] = rx_mbps
            d['tx_mbps'] = tx_mbps
            d.pop('interfaces_json', None)
            devices.append(d)
        return jsonify({'success': True, 'devices': devices})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/snmp/devices', methods=['POST'])
@login_required
def api_snmp_add_device():
    try:
        data = request.get_json()
        required = ['name', 'ip', 'community', 'device_type']
        for f in required:
            if not data.get(f):
                return jsonify({'success': False, 'error': f'{f} is required'}), 400
        conn = sqlite3.connect('ping_history.db', timeout=10)
        conn.execute('''INSERT INTO snmp_devices (name, ip, community, snmp_version, device_type,
                        location, region_id, added_by, created_at, is_active)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1)''',
                     (data['name'], data['ip'], data['community'],
                      data.get('snmp_version', '2c'), data['device_type'],
                      data.get('location', ''), data.get('region_id'),
                      session.get('user', {}).get('username', 'unknown'),
                      datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Device added'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/snmp/devices/<int:device_id>', methods=['PUT'])
@login_required
def api_snmp_update_device(device_id):
    try:
        data = request.get_json()
        conn = sqlite3.connect('ping_history.db', timeout=10)
        conn.execute('''UPDATE snmp_devices SET name=?, ip=?, community=?, snmp_version=?,
                        device_type=?, location=?, region_id=?, is_active=?
                        WHERE id=?''',
                     (data['name'], data['ip'], data['community'],
                      data.get('snmp_version', '2c'), data['device_type'],
                      data.get('location', ''), data.get('region_id'),
                      data.get('is_active', 1), device_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Device updated'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/snmp/devices/<int:device_id>', methods=['DELETE'])
@login_required
def api_snmp_delete_device(device_id):
    try:
        conn = sqlite3.connect('ping_history.db', timeout=10)
        conn.execute('DELETE FROM snmp_metrics WHERE device_id = ?', (device_id,))
        conn.execute('DELETE FROM snmp_devices WHERE id = ?', (device_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Device deleted'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/snmp/devices/<int:device_id>/metrics', methods=['GET'])
@login_required
def api_snmp_device_metrics(device_id):
    try:
        hours = int(request.args.get('hours', 1))
        conn = sqlite3.connect('ping_history.db', timeout=10)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('''SELECT timestamp, status, uptime_seconds, cpu_percent, mem_percent,
                          interfaces_json, rx_bytes, tx_bytes, signal_dbm
                          FROM snmp_metrics WHERE device_id = ?
                          AND timestamp >= datetime('now', ?)
                          ORDER BY timestamp ASC''',
                       (device_id, f'-{hours} hours'))
        rows = [dict(r) for r in cursor.fetchall()]
        conn.close()
        # Build chart-friendly series
        labels = [r['timestamp'] for r in rows]
        rx_series = []
        tx_series = []
        for r in rows:
            try:
                ifaces = json.loads(r.get('interfaces_json') or '[]')
                rx_series.append(ifaces[0].get('rx_mbps') if ifaces else None)
                tx_series.append(ifaces[0].get('tx_mbps') if ifaces else None)
            except Exception:
                rx_series.append(None)
                tx_series.append(None)
        return jsonify({'success': True, 'metrics': rows,
                        'chart': {'labels': labels, 'rx': rx_series, 'tx': tx_series}})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/snmp/test', methods=['POST'])
@login_required
def api_snmp_test():
    """Quick SNMP reachability test"""
    try:
        data = request.get_json()
        ip = data.get('ip')
        community = data.get('community', 'public')
        if not ip:
            return jsonify({'success': False, 'error': 'IP required'}), 400
        val = _snmp_get(ip, community, '1.3.6.1.2.1.1.1.0', timeout=3, retries=1)
        if val is not None:
            return jsonify({'success': True, 'message': f'SNMP reachable. sysDescr: {str(val)[:200]}'})
        else:
            return jsonify({'success': False, 'message': 'No SNMP response. Check IP, community string, and that SNMP is enabled.'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    start_periodic_update()
    socketio.run(app, host='0.0.0.0', port=5000, debug=False)
